from __future__ import annotations

import argparse
import json
import logging
import math
import os
import re
import shutil
import sys
import tempfile
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from functools import lru_cache
from pathlib import Path
from typing import Iterable

import fitz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pypdf import PdfReader


logger = logging.getLogger(__name__)


def _get_env_path(name: str, default: Path) -> Path:
    raw_value = os.getenv(name, "").strip()
    if not raw_value:
        return default
    return Path(raw_value).expanduser()


def _get_env_float(name: str, default: float) -> float:
    raw_value = os.getenv(name, "").strip()
    if not raw_value:
        return default
    try:
        return float(raw_value.replace(",", "."))
    except ValueError:
        logger.warning("Gecersiz ortam degiskeni degeri kullanildi: %s=%r", name, raw_value)
        return default


def _get_env_text(name: str, default: str) -> str:
    raw_value = os.getenv(name, "").strip()
    return raw_value or default


def _runtime_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def _resource_base_dir() -> Path:
    return Path(getattr(sys, "_MEIPASS", _runtime_base_dir()))


def configure_logging(level: str | None = None) -> None:
    root_logger = logging.getLogger()
    if root_logger.handlers:
        if level:
            root_logger.setLevel(getattr(logging, level.upper(), logging.INFO))
        return

    resolved_level = level or os.getenv("TEKLIF_KONTROL_LOG_LEVEL", "INFO")
    logging.basicConfig(
        level=getattr(logging, resolved_level.upper(), logging.INFO),
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    )


HEADER_PRODUCT = "URUN"
HEADER_PRODUCT_ALIASES = {
    "URUN",
    "URUN ADI",
    "MALZEME",
    "PRODUCT",
    "PRODUCT NAME",
}
HEADER_DEFAULT_PRICE = "2026 KURUMSAL NAKIT"
SUMMARY_SHEET = "Ozet"
RESULT_SHEET = "Sonuc"
CORRECTED_SHEET = "DuzeltilmisTeklif"
FINANCIAL_SHEET = "FinansalKontrol"
PRICE_HEADER_HINTS = ("KURUMSAL", "PERAKENDE", "NAKIT", "TAKSIT", "KDV")

HEADER_SKIP_LINES = {
    "MALZEME MIKTAR BIRIM FIYAT",
    "KURUMSAL",
    "INDIRIMLI",
    "FIYAT",
    "TOPLAM",
    "TUTAR",
}
NON_PRODUCT_LINE_HINTS = {
    "FIYAT",
    "KDV",
    "TEKLIF",
    "GARANTI",
    "TESLIM",
    "ODEME",
    "MONTAJ",
    "MALZEME",
    "FIRMA",
    "BIREYSEL",
    "YETKILI",
    "EMAIL",
    "GSM",
    "KURUMSAL",
    "PERAKENDE",
    "TOPLAM",
    "YATIRIM",
    "SECILEN",
    "ISKONTO",
    "MEBLAGI",
    "TUTARI",
    "INDIRIMLI",
}
SPEC_LINE_HINTS = {
    "KAPASITE",
    "DEPOLAMA",
    "SOGUK",
    "SICAK",
    "LT GUN",
    "LT SAAT",
    "L GUN",
    "L SAAT",
    "GUN",
    "SAAT",
    "WATT",
    "KW",
    "BAR",
    "PSI",
    "HZ",
    "VOLT",
    "AMPER",
    "MICRON",
    "GPD",
}

STOP_WORDS = {
    "VE",
    "ILE",
    "BY",
    "RAINWATER",
    "TAM",
    "OTOMATIK",
    "SISTEM",
    "SISTEMI",
}
CONFLICT_TOKEN_PAIRS = (
    ("ARITMALI", "ARITMASIZ"),
    ("POMPALI", "POMPASIZ"),
)
INCH_QUOTE_PATTERN = re.compile(r"(?P<value>\d+(?:[.,]\d+)?)\s*(?:\"|''|’’|′′|″|“|”)")
MEASUREMENT_PATTERNS = (
    ("LT_GUN", re.compile(r"\b(?P<value>\d+(?:[.,]\d+)?)\s*LT\s*(?:/|\b)\s*GUN\b")),
    ("LT_SAAT", re.compile(r"\b(?P<value>\d+(?:[.,]\d+)?)\s*LT\s*(?:/|\b)\s*SAAT\b")),
    ("GPD", re.compile(r"\b(?P<value>\d+(?:[.,]\d+)?)\s*GPD\b")),
    ("INC", re.compile(r"\b(?P<value>\d+(?:[.,]\d+)?)\s*(?:INC|INCH)\b")),
    ("LT", re.compile(r"\b(?P<value>\d+(?:[.,]\d+)?)\s*(?:LT|LITRE)\b(?!\s*(?:GUN|SAAT)\b)")),
)

ITEM_PATTERN = re.compile(
    r"^(?P<name>.+?)\s+"
    r"(?P<quantity>\d+(?:[.,]\d+)?)\s+ADET\s+"
    r"(?P<unit_price>\d{1,3}(?:\.\d{3})*(?:,\d+)?)\s*TL\s+"
    r"(?:(?P<discount_amount>\d{1,3}(?:\.\d{3})*(?:,\d+)?)\s*TL\s+)?"
    r"(?P<discounted_price>\d{1,3}(?:\.\d{3})*(?:,\d+)?)\s*TL\s+"
    r"(?P<total_price>\d{1,3}(?:\.\d{3})*(?:,\d+)?)\s*TL$",
    re.IGNORECASE,
)
MONEY_TL_PATTERN = re.compile(r"(?P<amount>\d{1,3}(?:\.\d{3})*(?:,\d+)?)\s*TL", re.IGNORECASE)
VAT_RATE_PATTERN = re.compile(r"KDV\s*\(\s*%?\s*(?P<rate>\d+(?:[.,]\d+)?)\s*\)", re.IGNORECASE)
VAT_EXCLUDED_PATTERN = re.compile(
    r"\bKDV\b.{0,48}\b(?:HARIC|HARICTIR|DAHIL\s+DEGIL|DAHIL\s+DEGILDIR)\b",
    re.IGNORECASE,
)
VAT_INCLUDED_PATTERN = re.compile(
    r"\bKDV\b.{0,24}\bDAHIL(?:DIR)?\b",
    re.IGNORECASE,
)

CODE_PATTERNS = (
    re.compile(r"\b[A-Z]{1,6}\s*-\s*\d{1,4}(?:\s*-\s*[A-Z0-9]{1,3})?\b"),
    re.compile(r"\b[A-Z]{1,6}\s+\d{1,4}\b"),
    re.compile(r"\b[A-Z]{1,6}\d{2,5}\b"),
    re.compile(r"(?:RNW|RWS|RO|RW)\s*\d{1,4}(?:\s+[A-Z])?"),
)
BUNDLE_SEPARATOR_PATTERN = re.compile(r"\s*\+\s*")
TURKISH_TRANSLATION = str.maketrans(
    {
        "\u00C7": "C",
        "\u00E7": "c",
        "\u011E": "G",
        "\u011F": "g",
        "\u0130": "I",
        "\u0131": "i",
        "\u00D6": "O",
        "\u00F6": "o",
        "\u015E": "S",
        "\u015F": "s",
        "\u00DC": "U",
        "\u00FC": "u",
    }
)
ARIAL_FONT_PATH = _get_env_path("TEKLIF_KONTROL_FONT_REGULAR", Path(r"C:\Windows\Fonts\arial.ttf"))
ARIAL_BOLD_FONT_PATH = _get_env_path("TEKLIF_KONTROL_FONT_BOLD", Path(r"C:\Windows\Fonts\arialbd.ttf"))
PDF_ACCENT = (11 / 255, 134 / 255, 200 / 255)
PDF_TEXT = (24 / 255, 53 / 255, 76 / 255)
PDF_MUTED = (93 / 255, 119 / 255, 140 / 255)
PDF_LIGHT = (198 / 255, 224 / 255, 238 / 255)
PDF_ROW_SOFT = (235 / 255, 248 / 255, 252 / 255)
MAX_GENERATED_OFFER_ITEMS = 6
DEFAULT_VAT_RATE = _get_env_float("TEKLIF_KONTROL_DEFAULT_VAT_RATE", 20.0)
DISCOUNT_TYPE_NONE = "none"
DISCOUNT_TYPE_AMOUNT = "amount"
DISCOUNT_TYPE_PERCENT = "percent"
DISCOUNT_TYPES = {DISCOUNT_TYPE_NONE, DISCOUNT_TYPE_AMOUNT, DISCOUNT_TYPE_PERCENT}
AUTO_OFFER_NUMBER_PATTERN = re.compile(
    r"^RW[-_](?P<daymonth>\d{4})[-_](?P<year>\d{4})(?:[-_](?P<sequence>\d{3}))?$",
    re.IGNORECASE,
)
TEMPLATE_BADGE_CLIPS = (
    fitz.Rect(187.92, 401.64, 220.2, 423.24),
    fitz.Rect(226.2, 403.44, 255.24, 422.88),
)
TEMPLATE_HEADER_CLIP = fitz.Rect(0, 0, 540, 84.2)
OUTPUT_ROOT_DIRNAME = _get_env_text("TEKLIF_KONTROL_OUTPUT_DIRNAME", "ciktilar")
GENERATED_OFFERS_DIRNAME = "olusturulan_teklifler"
REPORTS_DIRNAME = "raporlar"
CORRECTED_PDFS_DIRNAME = "duzeltilmis_teklifler"
RUNTIME_BASE_DIR = _get_env_path("CALL_PORTAL_OFFER_DATA_DIR", _runtime_base_dir())
RESOURCE_BASE_DIR = _resource_base_dir()
PRODUCT_ALIAS_PATH = _get_env_path(
    "TEKLIF_KONTROL_ALIAS_PATH",
    RUNTIME_BASE_DIR / "veri" / "urun_aliaslari.json",
)
BUNDLED_PRODUCT_ALIAS_PATH = RESOURCE_BASE_DIR / "veri" / "urun_aliaslari.json"
OFFER_HEADER_ASSET_NAME = "offer_header_banner.png"
OFFER_BADGE_ASSET_NAMES = ("offer_badge_belgium.png", "offer_badge_eu.png")
OFFER_SIGNATURE_CLEAR_RECT = fitz.Rect(320, 640, 525, 704)


@lru_cache(maxsize=None)
def _offer_asset_path(filename: str) -> Path | None:
    asset_path = RESOURCE_BASE_DIR / "assets" / filename
    if asset_path.exists():
        return asset_path
    fallback_path = RUNTIME_BASE_DIR / "assets" / filename
    if fallback_path.exists():
        return fallback_path
    return None


@dataclass(slots=True)
class PriceRow:
    row_number: int
    product_name: str
    prices: dict[str, float | None]
    note: str | None = None


@dataclass(slots=True)
class OfferItem:
    product_name: str
    quantity: float
    unit_price: float
    discounted_price: float
    total_price: float


@dataclass(slots=True)
class MatchResult:
    offer_item: OfferItem
    matched_row: PriceRow | None
    score: float
    status: str
    selected_column: str
    reference_unit_price: float | None
    reference_total_price: float | None
    suggested_unit_price: float | None
    suggested_total_price: float | None
    difference: float | None
    note: str


@dataclass(slots=True)
class OfferFinancialSummary:
    vat_rate: float
    vat_rate_source: str
    net_total: float | None
    vat_total: float | None
    gross_total: float | None


@dataclass(slots=True)
class FinancialCheck:
    label: str
    status: str
    offer_value: float | None
    calculated_value: float | None
    difference: float | None
    note: str


@dataclass(slots=True)
class FinancialReview:
    vat_rate: float
    vat_rate_source: str
    vat_included: bool
    item_gross_total: float
    expected_net_total: float
    expected_vat_total: float
    expected_gross_total: float
    expected_summary_total: float
    checks: list[FinancialCheck]

    @property
    def overall_status(self) -> str:
        if any(check.status == "DUZELT" for check in self.checks):
            return "DUZELT"
        if any(check.status == "INCELE" for check in self.checks):
            return "INCELE"
        return "ONAY"


@dataclass(slots=True)
class BundleComponentMatch:
    requested_name: str
    matched_row: PriceRow
    score: float
    reference_unit_price: float
    reference_column: str | None
    reference_source: str
    ambiguous: bool = False


@dataclass(slots=True)
class OfferLineItem:
    row_number: int
    product_name: str
    quantity: float
    reference_unit_price: float | None
    base_unit_price: float | None
    unit_price: float
    total_price: float
    discount_amount: float = 0.0
    price_source: str = "list"


@dataclass(slots=True)
class OfferSelection:
    row_number: int
    quantity: float
    manual_price: float | None = None
    discount_type: str = DISCOUNT_TYPE_NONE
    discount_value: float | None = None


def normalize_text(value: str) -> str:
    text = (value or "").replace("\u00A0", " ").replace("\u00AD", " ").translate(TURKISH_TRANSLATION)
    text = unicodedata.normalize("NFKD", text)
    text = INCH_QUOTE_PATTERN.sub(r"\g<value> INC ", text)
    text = text.encode("ascii", "ignore").decode("ascii")
    text = text.upper()
    text = INCH_QUOTE_PATTERN.sub(r"\g<value> INC ", text)
    # Some PDFs merge the brand and model code into a single token like
    # "RAINWATERRNW-2200"; split these back out before punctuation cleanup.
    text = re.sub(r"(RAINWATER)(?=(?:RNW|RWS|RO|RW)[-–—]?\d)", r"\1 ", text)
    text = re.sub(r"([A-Z])((?:RNW|RWS|RO|RW)[-–—]?\d)", r"\1 \2", text)
    text = text.replace("'", " ")
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_money(value: str | int | float | None) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return round(float(value), 2)
    text = str(value).strip().upper().replace("TL", "").replace(" ", "")
    if not text:
        return None
    text = text.replace(".", "").replace(",", ".")
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def _ensure_safe_output_path(
    output_path: Path,
    *,
    expected_suffix: str,
    label: str,
    source_paths: Iterable[Path] = (),
) -> Path:
    if output_path.suffix.lower() != expected_suffix:
        raise ValueError(f"{label} {expected_suffix} uzantili bir dosya olmali: {output_path.name}")
    if output_path.exists() and output_path.is_dir():
        raise ValueError(f"{label} bir klasor olamaz: {output_path}")

    try:
        resolved_output = output_path.resolve()
    except Exception:
        resolved_output = output_path.absolute()

    for source_path in source_paths:
        try:
            resolved_source = source_path.resolve()
        except Exception:
            resolved_source = source_path.absolute()
        if resolved_output == resolved_source:
            raise ValueError(f"{label} kaynak dosyanin uzerine yazilamaz: {output_path}")

    return output_path


def format_money(value: float | None) -> str:
    if value is None:
        return ""
    formatted = f"{value:,.2f}"
    formatted = formatted.replace(",", "X").replace(".", ",").replace("X", ".")
    if formatted.endswith(",00"):
        formatted = formatted[:-3]
    return f"{formatted} TL"


def format_pdf_money(value: float | None) -> str:
    if value is None:
        return ""
    rounded = int(round(value))
    return f"{rounded:,}".replace(",", ".") + " TL"


def excel_column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def find_single_file(
    base_dir: Path,
    suffix: str,
    exclude_prefixes: Iterable[str] | None = None,
    exclude_suffixes: Iterable[str] | None = None,
) -> Path:
    exclude_prefixes = tuple(exclude_prefixes or ())
    exclude_suffixes = tuple(exclude_suffixes or ())
    built_in_excludes = ("~$",)
    files = [
        path
        for path in sorted(base_dir.glob(f"*{suffix}"))
        if not any(path.name.startswith(prefix) for prefix in built_in_excludes + exclude_prefixes)
        and not any(path.name.endswith(excluded_suffix) for excluded_suffix in exclude_suffixes)
    ]
    if not files:
        raise FileNotFoundError(f"{suffix} uzantili dosya bulunamadı: {base_dir}")
    if len(files) > 1:
        names = ", ".join(path.name for path in files)
        raise FileExistsError(
            f"Birden fazla {suffix} dosyası bulundu. Lütfen dosya yolunu parametreyle verin: {names}"
        )
    return files[0]


def _find_price_sheet_layout(sheet) -> tuple[int, int, dict[int, str], dict[int, str], list[int]] | None:
    best_layout: tuple[int, int, dict[int, str], dict[int, str], list[int]] | None = None
    best_score = -1

    for row_index in range(1, sheet.max_row + 1):
        product_col_index = None
        for col_index in range(1, sheet.max_column + 1):
            cell_value = normalize_text(str(sheet.cell(row_index, col_index).value or ""))
            if cell_value in HEADER_PRODUCT_ALIASES:
                product_col_index = col_index
                break
        if product_col_index is None:
            continue

        headers_by_col: dict[int, str] = {}
        for col_index in range(1, sheet.max_column + 1):
            header_value = sheet.cell(row_index, col_index).value
            if header_value is None:
                continue
            header_text = str(header_value).strip()
            if not header_text:
                continue
            headers_by_col[col_index] = header_text

        price_headers_by_col = {
            col_index: header
            for col_index, header in headers_by_col.items()
            if col_index != product_col_index and any(hint in normalize_text(header) for hint in PRICE_HEADER_HINTS)
        }
        if not price_headers_by_col:
            continue

        note_columns = [
            col_index
            for col_index, header in headers_by_col.items()
            if col_index != product_col_index and normalize_text(header) in {"NOT", "ACIKLAMA", "NOTE", "NOTLAR"}
        ]
        score = (len(price_headers_by_col) * 1000) - row_index
        if score > best_score:
            best_score = score
            best_layout = (row_index, product_col_index, headers_by_col, price_headers_by_col, note_columns)

    return best_layout


def load_price_rows(workbook_path: Path, sheet_name: str | None = None) -> tuple[list[PriceRow], list[str]]:
    if not workbook_path.exists():
        raise FileNotFoundError(f"Excel dosyasi bulunamadi: {workbook_path}")
    if workbook_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(f"Desteklenmeyen Excel formati: {workbook_path.name}")

    workbook = None
    temp_workbook_path: Path | None = None
    try:
        workbook = load_workbook(workbook_path, data_only=True)
    except PermissionError:
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix=workbook_path.suffix) as temp_file:
                temp_workbook_path = Path(temp_file.name)
            shutil.copy2(workbook_path, temp_workbook_path)
            workbook = load_workbook(temp_workbook_path, data_only=True)
        except Exception as exc:
            if temp_workbook_path and temp_workbook_path.exists():
                temp_workbook_path.unlink(missing_ok=True)
            logger.exception("Excel dosyasi kilitli veya okunamadi: %s", workbook_path)
            raise ValueError(f"Excel dosyasi okunamadi: {workbook_path.name}") from exc
    except Exception as exc:
        logger.exception("Excel dosyasi okunamadi: %s", workbook_path)
        raise ValueError(f"Excel dosyasi okunamadi: {workbook_path.name}") from exc

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            available_sheets = ", ".join(workbook.sheetnames)
            workbook.close()
            if temp_workbook_path and temp_workbook_path.exists():
                temp_workbook_path.unlink(missing_ok=True)
            raise ValueError(
                f"Excel sayfasi bulunamadi: {sheet_name}. Kullanilabilir sayfalar: {available_sheets}"
            )
        sheet = workbook[sheet_name]
        layout = _find_price_sheet_layout(sheet)
    else:
        sheet = None
        layout = None
        for candidate_sheet in workbook.worksheets:
            candidate_layout = _find_price_sheet_layout(candidate_sheet)
            if candidate_layout is None:
                continue
            sheet = candidate_sheet
            layout = candidate_layout
            break

    if sheet is None or layout is None:
        workbook.close()
        if temp_workbook_path and temp_workbook_path.exists():
            temp_workbook_path.unlink(missing_ok=True)
        raise ValueError("Excel dosyasinda urun basligi veya fiyat kolonlari bulunamadi.")

    header_row_index, product_col_index, headers_by_col, price_headers_by_col, note_columns = layout
    ordered_headers: list[str] = [headers_by_col[product_col_index], *price_headers_by_col.values()]

    rows: list[PriceRow] = []
    for row_index in range(header_row_index + 1, sheet.max_row + 1):
        product_name = sheet.cell(row_index, product_col_index).value
        if product_name is None or not str(product_name).strip():
            continue

        prices: dict[str, float | None] = {}
        note: str | None = None
        for col_index, header in price_headers_by_col.items():
            value = sheet.cell(row_index, col_index).value
            prices[header] = parse_money(value)

        for col_index in note_columns:
            value = sheet.cell(row_index, col_index).value
            if isinstance(value, str) and value.strip():
                note = value.strip()
                break

        if not any(price is not None for price in prices.values()):
            continue

        rows.append(
            PriceRow(
                row_number=row_index,
                product_name=str(product_name).strip(),
                prices=prices,
                note=note,
            )
        )

    if not rows:
        workbook.close()
        if temp_workbook_path and temp_workbook_path.exists():
            temp_workbook_path.unlink(missing_ok=True)
        raise ValueError(f"Excel dosyasında kullanilabilir ürün satırı bulunamadı: {workbook_path.name}")

    workbook.close()
    if temp_workbook_path and temp_workbook_path.exists():
        temp_workbook_path.unlink(missing_ok=True)
    return rows, ordered_headers


def extract_offer_text(pdf_path: Path) -> str:
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF dosyasi bulunamadi: {pdf_path}")
    if pdf_path.suffix.lower() != ".pdf":
        raise ValueError(f"Desteklenmeyen teklif formati: {pdf_path.name}")

    try:
        with pdf_path.open("rb") as stream:
            reader = PdfReader(stream)
            text_parts: list[str] = []
            for page_index, page in enumerate(reader.pages, start=1):
                try:
                    text = page.extract_text() or ""
                except Exception as exc:
                    logger.warning("PDF sayfa metni okunamadi: %s sayfa=%s hata=%s", pdf_path, page_index, exc)
                    continue
                cleaned_text = text.replace("\uf0a7", " ")
                if cleaned_text.strip():
                    text_parts.append(cleaned_text)
    except Exception as exc:
        logger.exception("PDF dosyasi okunamadi: %s", pdf_path)
        raise ValueError(f"PDF dosyasi okunamadi veya bozuk: {pdf_path.name}") from exc

    fitz_parts: list[str] = []
    try:
        with fitz.open(pdf_path) as doc:
            for page in doc:
                page_text = (page.get_text("text") or "").replace("\uf0a7", " ")
                if page_text.strip():
                    fitz_parts.append(page_text)
    except Exception as exc:
        logger.warning("PyMuPDF ile ek metin okunamadi: %s hata=%s", pdf_path, exc)

    merged_lines: list[str] = []
    seen_lines: set[str] = set()
    for source_text in [*text_parts, *fitz_parts]:
        for raw_line in source_text.splitlines():
            line = raw_line.strip()
            if not line:
                continue
            normalized_line = normalize_text(line)
            if not normalized_line or normalized_line in seen_lines:
                continue
            seen_lines.add(normalized_line)
            merged_lines.append(line)

    if not merged_lines:
        raise ValueError(f"PDF icinden okunabilir metin cikarilamadi: {pdf_path.name}")
    return "\n".join(merged_lines)


def parse_offer_items(pdf_path: Path) -> tuple[list[OfferItem], str]:
    text = extract_offer_text(pdf_path)
    lines = [line.strip() for line in text.splitlines()]
    items: list[OfferItem] = []
    buffer: list[str] = []
    inside_items = False

    for raw_line in lines:
        if not raw_line:
            continue
        normalized = normalize_text(raw_line)
        if "MALZEME" in normalized and "MIKTAR" in normalized:
            inside_items = True
            buffer.clear()
            continue
        if inside_items and _is_item_section_terminator(normalized):
            inside_items = False
            buffer.clear()
            continue
        if not inside_items:
            continue
        if normalized == "FIYATTEKLIFI":
            inside_items = False
            buffer.clear()
            continue
        if not normalized or normalized in HEADER_SKIP_LINES:
            continue

        if "ADET" in normalized:
            candidates = _build_item_parse_candidates(buffer, raw_line)

            for candidate in candidates:
                match = ITEM_PATTERN.match(candidate)
                if not match:
                    continue
                items.append(
                    OfferItem(
                        product_name=match.group("name").strip(),
                        quantity=parse_money(match.group("quantity")) or 0,
                        unit_price=parse_money(match.group("unit_price")) or 0,
                        discounted_price=parse_money(match.group("discounted_price")) or 0,
                        total_price=parse_money(match.group("total_price")) or 0,
                    )
                )
                buffer.clear()
                break
            else:
                buffer.clear()
            continue

        if _is_possible_product_prefix(normalized):
            buffer.append(raw_line)
        elif buffer:
            buffer.append(raw_line)

    if not items:
        logger.warning("PDF teklif satirlari ayristrilamadi: %s", pdf_path)
        raise ValueError("PDF teklif satirlari ayristrilamadi.")
    return items, text


def _is_item_section_terminator(normalized_line: str) -> bool:
    if not normalized_line:
        return False
    return "YATIRIM MALIYETI" in normalized_line or "KURUMSAL SATIS SORUMLUSU" in normalized_line


def _is_possible_product_prefix(normalized_line: str) -> bool:
    if not normalized_line or "ADET" in normalized_line:
        return False
    return not any(hint in normalized_line for hint in NON_PRODUCT_LINE_HINTS)


def _build_item_parse_candidates(buffer: list[str], raw_line: str) -> list[str]:
    candidates: list[str] = [raw_line]
    if not buffer:
        return candidates

    preferred_prefix = _select_preferred_product_prefix(buffer)
    if preferred_prefix:
        candidates.insert(0, f"{preferred_prefix} {raw_line}".strip())

    full_prefix = " ".join(buffer).strip()
    if full_prefix and full_prefix != preferred_prefix:
        candidates.insert(1 if preferred_prefix else 0, f"{full_prefix} {raw_line}".strip())

    if _is_possible_product_prefix(normalize_text(buffer[-1])):
        trailing_one = " ".join([buffer[-1], raw_line]).strip()
        if trailing_one not in candidates:
            candidates.append(trailing_one)

    if len(buffer) >= 2:
        tail = buffer[-2:]
        if all(_is_possible_product_prefix(normalize_text(line)) for line in tail):
            trailing_two = " ".join(tail + [raw_line]).strip()
            if trailing_two not in candidates:
                candidates.append(trailing_two)

    return candidates


def _select_preferred_product_prefix(buffer: list[str]) -> str:
    selected: list[str] = []
    for index, line in enumerate(buffer):
        normalized = normalize_text(line)
        if index > 0 and _looks_like_technical_spec_line(normalized):
            break
        selected.append(line)
    if not selected:
        selected = buffer[:1]
    return " ".join(selected).strip()


def _looks_like_technical_spec_line(normalized_line: str) -> bool:
    if not normalized_line:
        return False

    digit_count = sum(character.isdigit() for character in normalized_line)
    if digit_count == 0:
        return False

    compact = normalized_line.replace("/", " ").replace("-", " ")
    has_spec_hint = any(hint in compact for hint in SPEC_LINE_HINTS)
    if has_spec_hint:
        return True

    if any(pattern.search(normalized_line) for pattern in CODE_PATTERNS):
        return False

    return False


def detect_vat_rate(offer_text: str) -> tuple[float, str]:
    match = VAT_RATE_PATTERN.search(offer_text)
    if match:
        detected_rate = parse_money(match.group("rate"))
        if detected_rate is not None and detected_rate >= 0:
            return detected_rate, "PDF"
    return 20.0, "Varsayilan"


def calculate_vat_breakdown(gross_total: float, vat_rate: float) -> tuple[float, float]:
    divisor = 1 + (vat_rate / 100)
    if divisor <= 0:
        return round(gross_total, 2), 0.0
    net_total = round(gross_total / divisor, 2)
    vat_total = round(gross_total - net_total, 2)
    return net_total, vat_total


def calculate_gross_from_net(net_total: float, vat_rate: float) -> tuple[float, float]:
    vat_total = round(net_total * (vat_rate / 100), 2)
    gross_total = round(net_total + vat_total, 2)
    return gross_total, vat_total


def detect_offer_vat_included(offer_text: str) -> bool | None:
    normalized = normalize_text(offer_text)

    lines = [normalize_text(raw_line) for raw_line in offer_text.splitlines()]
    for line in lines:
        if not line or "KDV" not in line:
            continue
        if VAT_EXCLUDED_PATTERN.search(line):
            return False

    if VAT_EXCLUDED_PATTERN.search(normalized):
        return False

    for line in lines:
        if not line or "KDV" not in line:
            continue
        if VAT_INCLUDED_PATTERN.search(line):
            return True

    if VAT_INCLUDED_PATTERN.search(normalized):
        return True
    return None


def is_kdv_excluded_header(header: str) -> bool:
    return "KDV HARIC" in normalize_text(header)


def try_resolve_price_column(requested_header: str, available_headers: Iterable[str]) -> str | None:
    normalized_target = normalize_text(requested_header)
    for header in available_headers:
        if normalize_text(header) == normalized_target:
            return header
    return None


def get_kdv_excluded_header(base_header: str, available_headers: Iterable[str]) -> str | None:
    if is_kdv_excluded_header(base_header):
        return try_resolve_price_column(base_header, available_headers)
    return try_resolve_price_column(f"KDV HARİÇ {base_header}", available_headers)


def resolve_price_for_vat_mode(
    price_row: PriceRow,
    selected_column: str,
    *,
    vat_included: bool,
    vat_rate: float = DEFAULT_VAT_RATE,
) -> tuple[float | None, str | None, str]:
    available_headers = list(price_row.prices.keys())
    direct_value = price_row.prices.get(selected_column)

    if vat_included:
        if direct_value is not None and not is_kdv_excluded_header(selected_column):
            return direct_value, selected_column, "gross_list"
        if direct_value is not None and is_kdv_excluded_header(selected_column):
            gross_value, _ = calculate_gross_from_net(direct_value, vat_rate)
            return gross_value, selected_column, "gross_derived"
        return None, selected_column, "missing"

    excluded_header = get_kdv_excluded_header(selected_column, available_headers)
    if excluded_header:
        excluded_value = price_row.prices.get(excluded_header)
        if excluded_value is not None:
            return excluded_value, excluded_header, "net_list"

    if direct_value is not None:
        if is_kdv_excluded_header(selected_column):
            return direct_value, selected_column, "net_list"
        net_value, _ = calculate_vat_breakdown(direct_value, vat_rate)
        return net_value, selected_column, "net_derived"

    return None, excluded_header or selected_column, "missing"


def derive_standard_price_header(selected_column: str, available_headers: Iterable[str]) -> str | None:
    normalized = normalize_text(selected_column)
    candidates: list[str] = []

    if "PERAKENDE" in normalized:
        if has_taksit_count(selected_column, 6):
            candidates.append("2026 PERAKENDE 6 TAKSİT")
        elif has_taksit_count(selected_column, 4):
            candidates.append("2026 PERAKENDE 4 TAKSİT")
        elif "NAKIT" in normalized:
            candidates.append("2026 PERAKENDE NAKİT")
        candidates.append(selected_column)
    else:
        if has_taksit_count(selected_column, 6):
            candidates.append("2026 PERAKENDE 6 TAKSİT")
        elif has_taksit_count(selected_column, 4):
            candidates.append("2026 PERAKENDE 4 TAKSİT")
        elif "NAKIT" in normalized:
            candidates.append("2026 PERAKENDE NAKİT")

    for candidate in candidates:
        resolved = try_resolve_price_column(candidate, available_headers)
        if resolved:
            return resolved
    return None


def calculate_offer_totals(
    line_total_sum: float,
    *,
    vat_rate: float,
    vat_included: bool,
) -> tuple[float, float, float]:
    if vat_included:
        gross_total = round(line_total_sum, 2)
        net_total, vat_total = calculate_vat_breakdown(gross_total, vat_rate)
        return net_total, vat_total, gross_total

    net_total = round(line_total_sum, 2)
    gross_total, vat_total = calculate_gross_from_net(net_total, vat_rate)
    return net_total, vat_total, gross_total


def _find_summary_amount(
    lines: list[str],
    label_matcher,
    *,
    max_lookahead: int = 3,
) -> float | None:
    for index in range(len(lines) - 1, -1, -1):
        normalized = normalize_text(lines[index])
        if not label_matcher(normalized):
            continue
        for offset in range(max_lookahead + 1):
            candidate_index = index + offset
            if candidate_index >= len(lines):
                break
            match = MONEY_TL_PATTERN.search(lines[candidate_index])
            if match:
                return parse_money(match.group("amount"))
    return None


def parse_offer_financial_summary(offer_text: str) -> OfferFinancialSummary:
    lines = [line.strip() for line in offer_text.splitlines() if line.strip()]
    vat_rate, vat_rate_source = detect_vat_rate(offer_text)
    label_matchers = {
        "net_total": lambda normalized: normalized.startswith("YATIRIM MALIYETI"),
        "vat_total": lambda normalized: normalized.startswith("KDV"),
        "gross_total": lambda normalized: normalized.startswith("TOPLAM YATIRIM MALIYETI"),
    }
    label_indexes: dict[str, int] = {}
    for index in range(len(lines) - 1, -1, -1):
        normalized = normalize_text(lines[index])
        for key, matcher in label_matchers.items():
            if key not in label_indexes and matcher(normalized):
                label_indexes[key] = index

    net_total = None
    vat_total = None
    gross_total = None
    if len(label_indexes) == 3:
        ordered_labels = sorted((index, key) for key, index in label_indexes.items())
        found_amounts: list[float] = []
        for candidate_index in range(ordered_labels[0][0], len(lines)):
            match = MONEY_TL_PATTERN.search(lines[candidate_index])
            if match:
                parsed_amount = parse_money(match.group("amount"))
                if parsed_amount is not None:
                    found_amounts.append(parsed_amount)
                if len(found_amounts) == 3:
                    break
        if len(found_amounts) == 3:
            grouped_values = {
                key: amount for (_, key), amount in zip(ordered_labels, found_amounts, strict=False)
            }
            candidate_net_total = grouped_values.get("net_total")
            candidate_vat_total = grouped_values.get("vat_total")
            candidate_gross_total = grouped_values.get("gross_total")
            if (
                candidate_net_total is not None
                and candidate_vat_total is not None
                and candidate_gross_total is not None
                and abs((candidate_net_total + candidate_vat_total) - candidate_gross_total) <= 2.0
                and candidate_gross_total >= candidate_net_total
                and candidate_gross_total >= candidate_vat_total
            ):
                net_total = candidate_net_total
                vat_total = candidate_vat_total
                gross_total = candidate_gross_total

    if net_total is None:
        net_total = _find_summary_amount(lines, label_matchers["net_total"])
    if vat_total is None:
        vat_total = _find_summary_amount(lines, label_matchers["vat_total"])
    if gross_total is None:
        gross_total = _find_summary_amount(lines, label_matchers["gross_total"])

    return OfferFinancialSummary(
        vat_rate=vat_rate,
        vat_rate_source=vat_rate_source,
        net_total=net_total,
        vat_total=vat_total,
        gross_total=gross_total,
    )


def _build_financial_check(
    label: str,
    offer_value: float | None,
    calculated_value: float | None,
    *,
    tolerance: float,
    success_note: str,
    mismatch_note: str,
    missing_note: str,
    extra_note: str | None = None,
) -> FinancialCheck:
    if offer_value is None or calculated_value is None:
        note = missing_note
        if extra_note:
            note = f"{note} {extra_note}"
        return FinancialCheck(
            label=label,
            status="INCELE",
            offer_value=offer_value,
            calculated_value=calculated_value,
            difference=None,
            note=note,
        )

    difference = round(offer_value - calculated_value, 2)
    if abs(difference) <= tolerance:
        note = success_note
        status = "ONAY"
    else:
        note = mismatch_note
        status = "DUZELT"

    if extra_note:
        note = f"{note} {extra_note}"

    return FinancialCheck(
        label=label,
        status=status,
        offer_value=offer_value,
        calculated_value=calculated_value,
        difference=difference,
        note=note,
    )


def _build_line_item_consistency_check(offer_items: list[OfferItem], *, tolerance: float) -> FinancialCheck:
    mismatches: list[str] = []
    declared_total = round(sum(item.total_price for item in offer_items), 2)
    calculated_total = 0.0

    for item in offer_items:
        expected_total = round(item.discounted_price * item.quantity, 2)
        calculated_total = round(calculated_total + expected_total, 2)
        if abs(item.total_price - expected_total) > tolerance:
            mismatches.append(item.product_name)

    if not mismatches:
        return FinancialCheck(
            label="Kalem Toplam Tutarlılığı",
            status="ONAY",
            offer_value=declared_total,
            calculated_value=calculated_total,
            difference=round(declared_total - calculated_total, 2),
            note="Kalem toplamları birim fiyat x adet hesabıyla uyumlu.",
        )

    sample = ", ".join(mismatches[:3])
    if len(mismatches) > 3:
        sample = f"{sample} ve {len(mismatches) - 3} satir daha"

    return FinancialCheck(
        label="Kalem Toplam Tutarlılığı",
        status="DUZELT",
        offer_value=declared_total,
        calculated_value=calculated_total,
        difference=round(declared_total - calculated_total, 2),
        note=f"Bazi satirlarda birim fiyat x adet ile toplam tutar uyusmuyor: {sample}",
    )


def _build_vat_mode_check(offer_text: str, *, vat_included: bool | None, vat_rate: float) -> FinancialCheck:
    if vat_included is None:
        return FinancialCheck(
            label="KDV Modu",
            status="INCELE",
            offer_value=None,
            calculated_value=None,
            difference=None,
            note=f"PDF icinde KDV dahil/haric bilgisi net okunamadi; finansal kontrolde %{vat_rate:g} KDV varsayimi kullanildi.",
        )

    if vat_included:
        return FinancialCheck(
            label="KDV Modu",
            status="ONAY",
            offer_value=None,
            calculated_value=None,
            difference=None,
            note=f"PDF KDV dahil olarak algilandi; kalem toplamindan net, KDV ve brut toplam kontrol edildi. KDV orani: %{vat_rate:g}.",
        )

    return FinancialCheck(
        label="KDV Modu",
        status="ONAY",
        offer_value=None,
        calculated_value=None,
        difference=None,
        note=f"PDF KDV haric olarak algilandi; kalem toplami net toplam kabul edilerek kontrol edildi. KDV orani: %{vat_rate:g}.",
    )


def build_financial_review(
    offer_items: list[OfferItem],
    offer_text: str,
    *,
    tolerance: float = 1.0,
) -> FinancialReview:
    summary = parse_offer_financial_summary(offer_text)
    line_total_sum = round(sum(item.total_price for item in offer_items), 2)
    detected_vat_included = detect_offer_vat_included(offer_text)
    vat_included = detected_vat_included
    if vat_included is None:
        vat_included = True
    expected_net_total, expected_vat_total, expected_gross_total = calculate_offer_totals(
        line_total_sum,
        vat_rate=summary.vat_rate,
        vat_included=vat_included,
    )
    expected_summary_total = expected_gross_total if vat_included else expected_net_total
    vat_rate_note = (
        None
        if summary.vat_rate_source == "PDF"
        else f"KDV oran\u0131 PDF'ten okunamad\u0131, %{summary.vat_rate:g} varsay\u0131ld\u0131."
    )
    total_label = "Toplam Yat\u0131r\u0131m Maliyeti" if vat_included else "Toplam Yat\u0131r\u0131m Maliyeti (KDV Hari\u00e7)"
    total_success_note = (
        "Kalem toplam\u0131 ile uyumlu."
        if vat_included
        else "KDV hari\u00e7 toplam alan\u0131 kalem toplam\u0131 ile uyumlu."
    )
    total_mismatch_note = (
        "Toplam yat\u0131r\u0131m maliyeti kalem toplam\u0131 ile uyu\u015fmuyor."
        if vat_included
        else "KDV hari\u00e7 toplam alan\u0131 kalem toplam\u0131 ile uyu\u015fmuyor."
    )

    checks = [
        _build_line_item_consistency_check(offer_items, tolerance=tolerance),
        _build_vat_mode_check(offer_text, vat_included=detected_vat_included, vat_rate=summary.vat_rate),
    ]

    if vat_included:
        checks = checks + [
            _build_financial_check(
                "Yat\u0131r\u0131m Maliyeti",
                summary.net_total,
                expected_net_total,
                tolerance=tolerance,
                success_note="Kalem toplamlar\u0131 ve KDV oran\u0131 ile uyumlu.",
                mismatch_note="Teklif \u00f6zeti kalem toplamlar\u0131na g\u00f6re farkl\u0131 hesaplanm\u0131\u015f.",
                missing_note="PDF \u00f6zetinde yat\u0131r\u0131m maliyeti alan\u0131 bulunamad\u0131.",
                extra_note=vat_rate_note,
            ),
            _build_financial_check(
                "KDV",
                summary.vat_total,
                expected_vat_total,
                tolerance=tolerance,
                success_note="KDV tutar\u0131 hesapla uyumlu.",
                mismatch_note="KDV tutar\u0131 kalem toplamlar\u0131na g\u00f6re farkl\u0131.",
                missing_note="PDF \u00f6zetinde KDV alan\u0131 bulunamad\u0131.",
                extra_note=vat_rate_note,
            ),
            _build_financial_check(
                total_label,
                summary.gross_total,
                expected_summary_total,
                tolerance=tolerance,
                success_note=total_success_note,
                mismatch_note=total_mismatch_note,
                missing_note="PDF \u00f6zetinde toplam yat\u0131r\u0131m maliyeti alan\u0131 bulunamad\u0131.",
            ),
        ]
    else:
        checks = checks + [
            _build_financial_check(
                "Yat\u0131r\u0131m Maliyeti (KDV Hari\u00e7)",
                summary.net_total,
                expected_net_total,
                tolerance=tolerance,
                success_note="KDV hari\u00e7 yat\u0131r\u0131m maliyeti kalem toplam\u0131 ile uyumlu.",
                mismatch_note="KDV hari\u00e7 yat\u0131r\u0131m maliyeti kalem toplam\u0131na g\u00f6re farkl\u0131.",
                missing_note="PDF \u00f6zetinde KDV hari\u00e7 yat\u0131r\u0131m maliyeti alan\u0131 bulunamad\u0131.",
                extra_note=vat_rate_note,
            ),
            _build_financial_check(
                "KDV",
                summary.vat_total,
                expected_vat_total,
                tolerance=tolerance,
                success_note="KDV tutar\u0131 hesapla uyumlu.",
                mismatch_note="KDV tutar\u0131 kalem toplamlar\u0131na g\u00f6re farkl\u0131.",
                missing_note="PDF \u00f6zetinde KDV alan\u0131 bulunamad\u0131.",
                extra_note=vat_rate_note,
            ),
            _build_financial_check(
                "Toplam Yat\u0131r\u0131m Maliyeti (KDV Dahil)",
                summary.gross_total,
                expected_gross_total,
                tolerance=tolerance,
                success_note="KDV dahil genel toplam net tutar ve KDV ile uyumlu.",
                mismatch_note="KDV dahil genel toplam net tutar ve KDV hesab\u0131na g\u00f6re farkl\u0131.",
                missing_note="PDF \u00f6zetinde KDV dahil toplam yat\u0131r\u0131m maliyeti alan\u0131 bulunamad\u0131.",
            ),
        ]

    return FinancialReview(
        vat_rate=summary.vat_rate,
        vat_rate_source=summary.vat_rate_source,
        vat_included=vat_included,
        item_gross_total=expected_gross_total,
        expected_net_total=expected_net_total,
        expected_vat_total=expected_vat_total,
        expected_gross_total=expected_gross_total,
        expected_summary_total=expected_summary_total,
        checks=checks,
    )


def tokenize(value: str) -> set[str]:
    normalized = normalize_text(value)
    return {
        token
        for token in normalized.split()
        if len(token) > 1 and token not in STOP_WORDS and not token.isdigit()
    }


def find_descriptor_conflicts(left: str, right: str) -> tuple[str, ...]:
    left_tokens = tokenize(left)
    right_tokens = tokenize(right)
    conflicts: list[str] = []
    for first, second in CONFLICT_TOKEN_PAIRS:
        if (first in left_tokens and second in right_tokens) or (second in left_tokens and first in right_tokens):
            conflicts.append(f"{first}/{second}")
    return tuple(conflicts)


def extract_measurements(value: str) -> dict[str, set[str]]:
    normalized = normalize_text(value)
    measurements: dict[str, set[str]] = {}
    for unit, pattern in MEASUREMENT_PATTERNS:
        values = {match.group("value").replace(",", ".") for match in pattern.finditer(normalized)}
        if values:
            measurements[unit] = values
    return measurements


def find_measurement_conflicts(left: str, right: str) -> tuple[str, ...]:
    left_measurements = extract_measurements(left)
    right_measurements = extract_measurements(right)
    conflicts: list[str] = []
    for unit in sorted(left_measurements.keys() & right_measurements.keys()):
        left_values = left_measurements[unit]
        right_values = right_measurements[unit]
        if left_values and right_values and left_values.isdisjoint(right_values):
            left_label = "/".join(sorted(left_values))
            right_label = "/".join(sorted(right_values))
            conflicts.append(f"{unit}:{left_label}!={right_label}")
    return tuple(conflicts)


def extract_codes(value: str) -> set[str]:
    normalized = normalize_text(value)
    codes = set()
    for pattern in CODE_PATTERNS:
        for match in pattern.findall(normalized):
            codes.add(re.sub(r"[^A-Z0-9]", "", match))
    return codes


@lru_cache(maxsize=1)
def load_product_alias_groups() -> tuple[tuple[str, tuple[str, ...]], ...]:
    alias_path = PRODUCT_ALIAS_PATH
    if not alias_path.exists() and BUNDLED_PRODUCT_ALIAS_PATH.exists():
        try:
            alias_path.parent.mkdir(parents=True, exist_ok=True)
            alias_path.write_text(BUNDLED_PRODUCT_ALIAS_PATH.read_text(encoding="utf-8"), encoding="utf-8")
        except Exception as exc:
            logger.warning("Urun alias dosyasi calisma klasorune kopyalanamadi: %s", exc)
            alias_path = BUNDLED_PRODUCT_ALIAS_PATH

    if not alias_path.exists():
        logger.warning("Urun alias dosyasi bulunamadi: %s", alias_path)
        return ()

    try:
        payload = json.loads(alias_path.read_text(encoding="utf-8"))
    except Exception as exc:
        logger.warning("Urun alias dosyasi okunamadi: %s hata=%s", alias_path, exc)
        return ()

    if not isinstance(payload, dict):
        logger.warning("Urun alias dosyasi beklenen yapida degil: %s", alias_path)
        return ()

    raw_groups = payload.get("groups", [])
    if not isinstance(raw_groups, list):
        logger.warning("Urun alias gruplari liste yapisinda degil: %s", alias_path)
        return ()

    groups: list[tuple[str, tuple[str, ...]]] = []
    for entry in raw_groups:
        if not isinstance(entry, dict):
            continue
        canonical = str(entry.get("canonical") or "").strip()
        raw_aliases = entry.get("aliases", [])
        if not isinstance(raw_aliases, (list, tuple)):
            continue
        aliases = tuple(str(alias).strip() for alias in raw_aliases if str(alias).strip())
        if not canonical or not aliases:
            continue
        groups.append((canonical, aliases))
    return tuple(groups)


def get_alias_candidates(value: str) -> list[str]:
    normalized_value = normalize_text(value)
    candidates = [value]
    for primary_name, aliases in load_product_alias_groups():
        group_values = (primary_name, *aliases)
        normalized_group = {normalize_text(candidate) for candidate in group_values}
        if normalized_value not in normalized_group:
            continue
        for candidate in group_values:
            if normalize_text(candidate) == normalized_value:
                continue
            candidates.append(candidate)
    return candidates


def is_alias_match(left: str, right: str) -> bool:
    left_normalized = normalize_text(left)
    right_normalized = normalize_text(right)
    if not left_normalized or not right_normalized or left_normalized == right_normalized:
        return False
    return bool({normalize_text(candidate) for candidate in get_alias_candidates(left)} & {normalize_text(candidate) for candidate in get_alias_candidates(right)})


def _base_similarity_score(left: str, right: str) -> float:
    left_normalized = normalize_text(left)
    right_normalized = normalize_text(right)
    if left_normalized and left_normalized == right_normalized:
        return 1.0
    left_tokens = tokenize(left)
    right_tokens = tokenize(right)
    token_overlap = (
        len(left_tokens & right_tokens) / len(left_tokens | right_tokens)
        if left_tokens or right_tokens
        else 0.0
    )
    sequence_score = SequenceMatcher(None, left_normalized, right_normalized).ratio()
    token_sequence = SequenceMatcher(
        None,
        " ".join(sorted(left_tokens)),
        " ".join(sorted(right_tokens)),
    ).ratio()
    left_codes = extract_codes(left)
    right_codes = extract_codes(right)
    shared_codes = left_codes & right_codes
    code_score = 1.0 if shared_codes else 0.0
    conflicts = find_descriptor_conflicts(left, right)
    measurement_conflicts = find_measurement_conflicts(left, right)
    containment_bonus = (
        0.1
        if not conflicts
        and not measurement_conflicts
        and (left_normalized in right_normalized or right_normalized in left_normalized)
        else 0.0
    )

    score = (sequence_score * 0.18) + (token_overlap * 0.22) + (token_sequence * 0.15) + (code_score * 0.45)
    if shared_codes and left_tokens & right_tokens:
        score += 0.1
    if conflicts:
        score -= 0.34 + (0.08 if shared_codes else 0.0)
    if measurement_conflicts:
        score -= 0.36 + (0.08 if shared_codes else 0.0)
    return min(1.0, max(0.0, score + containment_bonus))


def similarity_score(left: str, right: str) -> float:
    best_score = 0.0
    for left_candidate in get_alias_candidates(left):
        for right_candidate in get_alias_candidates(right):
            best_score = max(best_score, _base_similarity_score(left_candidate, right_candidate))
    if is_alias_match(left, right):
        best_score = max(best_score, 0.9)
    return best_score


def find_unique_code_match(product_name: str, price_rows: Iterable[PriceRow]) -> tuple[PriceRow, float] | None:
    product_codes = extract_codes(product_name)
    if not product_codes:
        return None

    ranked_candidates: list[tuple[int, float, PriceRow]] = []
    for price_row in price_rows:
        row_codes = extract_codes(price_row.product_name)
        shared_codes = product_codes & row_codes
        if not shared_codes:
            continue
        longest_code_length = max(len(code) for code in shared_codes)
        ranked_candidates.append((longest_code_length, similarity_score(product_name, price_row.product_name), price_row))

    if not ranked_candidates:
        return None

    ranked_candidates.sort(key=lambda item: (item[0], item[1]), reverse=True)
    best_length, best_similarity, best_row = ranked_candidates[0]
    second_length = ranked_candidates[1][0] if len(ranked_candidates) > 1 else -1
    if best_length == second_length and len(ranked_candidates) > 1:
        second_similarity = ranked_candidates[1][1]
        if abs(best_similarity - second_similarity) < 0.03:
            return None

    return best_row, max(best_similarity, 0.96)


def split_bundle_product_name(product_name: str) -> list[str]:
    if "+" not in product_name:
        return []
    raw_parts = [
        part.strip(" -\t\r\n")
        for part in BUNDLE_SEPARATOR_PATTERN.split(product_name)
        if part.strip(" -\t\r\n")
    ]
    if len(raw_parts) < 2:
        return []
    parts = [_compact_bundle_component_name(part) for part in raw_parts]
    return [part for part in parts if part]


def _compact_bundle_component_name(component_name: str) -> str:
    cleaned = " ".join(str(component_name or "").split()).strip()
    if not cleaned:
        return ""
    tokens = cleaned.split()
    normalized_tokens = [normalize_text(token) for token in tokens]

    # "40 LT TANK Yuksek Kapasiteli..." -> "40 LT TANK"
    if "TANK" in normalized_tokens:
        tank_index = normalized_tokens.index("TANK")
        if tank_index >= 1:
            return " ".join(tokens[: tank_index + 1]).strip()

    # Keep short model-oriented prefixes before the generic description starts.
    cut_markers = {"ARITMALI", "ARITMASIZ", "YUKSEK", "TAM", "OTOMATIK", "SISTEMI", "SISTEM"}
    compact_tokens: list[str] = []
    for token, normalized in zip(tokens, normalized_tokens, strict=False):
        if compact_tokens and normalized in cut_markers:
            break
        compact_tokens.append(token)
    return " ".join(compact_tokens).strip()


def _component_name_variants(component_name: str, full_product_name: str) -> list[str]:
    variants: list[str] = []

    def add_variant(value: str) -> None:
        cleaned = " ".join(str(value or "").split())
        if cleaned and cleaned not in variants:
            variants.append(cleaned)

    add_variant(component_name)
    normalized_component = normalize_text(component_name)
    if "LITRE" in normalized_component:
        add_variant(re.sub(r"\bLITRE\b", "LT", normalized_component))
    if "INC" in normalized_component:
        add_variant(re.sub(r"\bINC\b", "INÇ", normalized_component))

    full_tokens = full_product_name.split()
    brand_prefix = ""
    if full_tokens:
        first_token = normalize_text(full_tokens[0])
        if first_token in {"RAINWATER", "AXEON", "BLACK"}:
            brand_prefix = full_tokens[0]
    if brand_prefix and not normalize_text(component_name).startswith(normalize_text(brand_prefix)):
        add_variant(f"{brand_prefix} {component_name}")
        if "LITRE" in normalized_component:
            litre_variant = re.sub(r"\bLITRE\b", "LT", normalized_component)
            add_variant(f"{brand_prefix} {litre_variant}")

    return variants


def _component_similarity_score(component_name: str, full_product_name: str, catalog_name: str) -> float:
    best_score = 0.0
    catalog_tokens = tokenize(catalog_name)
    catalog_numbers = set(re.findall(r"\d+", normalize_text(catalog_name)))
    catalog_codes = extract_codes(catalog_name)
    for variant in _component_name_variants(component_name, full_product_name):
        score = similarity_score(variant, catalog_name)
        variant_tokens = tokenize(variant)
        variant_numbers = set(re.findall(r"\d+", normalize_text(variant)))
        variant_codes = extract_codes(variant)
        measurement_conflicts = find_measurement_conflicts(variant, catalog_name)
        if measurement_conflicts:
            score = min(score, 0.34)
        elif variant_codes and variant_codes & catalog_codes:
            score = max(score, 0.95)
        elif (
            not variant_codes
            and variant_tokens
            and variant_tokens <= catalog_tokens
            and variant_numbers
            and variant_numbers & catalog_numbers
        ):
            score = max(score, 0.88)
        best_score = max(best_score, score)
    return best_score


def _find_bundle_component_match(
    component_name: str,
    full_product_name: str,
    price_rows: list[PriceRow],
    selected_column: str,
    *,
    vat_included: bool,
    vat_rate: float,
    min_match_score: float,
) -> BundleComponentMatch | None:
    ranked_rows = sorted(
        (
            (_component_similarity_score(component_name, full_product_name, price_row.product_name), price_row)
            for price_row in price_rows
        ),
        key=lambda item: item[0],
        reverse=True,
    )
    if not ranked_rows:
        return None
    best_score, best_row = ranked_rows[0]
    if best_score < min_match_score:
        return None

    reference_unit_price, reference_column, reference_source = resolve_price_for_vat_mode(
        best_row,
        selected_column,
        vat_included=vat_included,
        vat_rate=vat_rate,
    )
    if reference_unit_price is None:
        return None

    second_score = ranked_rows[1][0] if len(ranked_rows) > 1 else 0.0
    return BundleComponentMatch(
        requested_name=component_name,
        matched_row=best_row,
        score=best_score,
        reference_unit_price=reference_unit_price,
        reference_column=reference_column,
        reference_source=reference_source,
        ambiguous=(best_score - second_score) < 0.06,
    )


def resolve_price_column(requested_header: str, available_headers: Iterable[str]) -> str:
    normalized_target = normalize_text(requested_header)
    for header in available_headers:
        if normalize_text(header) == normalized_target:
            return header
    available = ", ".join(available_headers)
    raise ValueError(f"'{requested_header}' sütunu bulunamadı. Kullanılabilir başlıklar: {available}")


def get_price_columns(available_headers: Iterable[str]) -> list[str]:
    price_headers: list[str] = []
    for header in available_headers:
        normalized = normalize_text(header)
        if any(hint in normalized for hint in PRICE_HEADER_HINTS):
            price_headers.append(header)
    return price_headers


def has_taksit_count(value: str, count: int) -> bool:
    normalized = f" {normalize_text(value)} "
    return f" {count} TAKSIT " in normalized


def extract_selected_price_column(
    offer_text: str,
    available_headers: Iterable[str],
    *,
    vat_included: bool | None = None,
) -> str | None:
    normalized_available = {normalize_text(header): header for header in available_headers}
    for raw_line in offer_text.splitlines():
        line = (raw_line or "").strip()
        if not line:
            continue
        normalized_line = normalize_text(line)
        if not normalized_line.startswith("SECILEN FIYAT TIPI"):
            continue
        if ":" in line:
            line = line.split(":", 1)[1].strip()
        if "|" in line:
            line = line.split("|", 1)[0].strip()
        normalized_line = normalize_text(line)
        if not normalized_line:
            continue
        if vat_included is False:
            excluded = normalized_available.get(normalize_text(f"KDV HARIC {line}"))
            if excluded:
                return excluded
        exact = normalized_available.get(normalized_line)
        if exact:
            return exact
    return None


def extract_payment_mode_hint(offer_text: str) -> str | None:
    for raw_line in offer_text.splitlines():
        line = (raw_line or "").strip()
        if not line:
            continue
        normalized_line = normalize_text(line)
        if "ODEME BILGISI" not in normalized_line:
            continue
        if "NAKIT" in normalized_line or "BANKA HAVALESI" in normalized_line:
            return "NAKIT"
        if "6 TAKSIT" in normalized_line:
            return "6 TAKSIT"
        if "4 TAKSIT" in normalized_line:
            return "4 TAKSIT"
    return None


def detect_price_column(offer_text: str, available_headers: list[str]) -> str:
    normalized_text = normalize_text(offer_text)
    vat_included = detect_offer_vat_included(offer_text)
    normalized_available = {normalize_text(header): header for header in available_headers}
    selected_price_column = extract_selected_price_column(
        offer_text,
        available_headers,
        vat_included=vat_included,
    )
    if selected_price_column:
        return selected_price_column
    payment_mode_hint = extract_payment_mode_hint(offer_text)
    exact_matches = [
        header
        for normalized_header, header in normalized_available.items()
        if normalized_header and normalized_header in normalized_text
    ]
    if vat_included is False:
        for header in exact_matches:
            if is_kdv_excluded_header(header):
                return header
    for header in exact_matches:
        if not is_kdv_excluded_header(header):
            return header

    preferred_headers = [
        "2026 KURUMSAL 6 TAKSİT",
        "2026 KURUMSAL 4 TAKSİT",
        HEADER_DEFAULT_PRICE,
        "2026 PERAKENDE NAKİT",
    ]

    if "PERAKENDE" in normalized_text and has_taksit_count(offer_text, 6):
        preferred_headers = [
            "2026 PERAKENDE 6 TAKSİT",
            "2026 PERAKENDE 4 TAKSİT",
            "2026 PERAKENDE NAKİT",
        ]
    elif "PERAKENDE" in normalized_text and has_taksit_count(offer_text, 4):
        preferred_headers = [
            "2026 PERAKENDE 4 TAKSİT",
            "2026 PERAKENDE 6 TAKSİT",
            "2026 PERAKENDE NAKİT",
        ]
    elif "PERAKENDE" in normalized_text:
        preferred_headers = [
            "2026 PERAKENDE NAKİT",
            "2026 PERAKENDE 4 TAKSİT",
            "2026 PERAKENDE 6 TAKSİT",
        ]
    elif payment_mode_hint == "NAKIT" and "PERAKENDE" not in normalized_text:
        preferred_headers = [
            HEADER_DEFAULT_PRICE,
            "2026 KURUMSAL 4 TAKSİT",
            "2026 KURUMSAL 6 TAKSİT",
        ]
    elif payment_mode_hint == "NAKIT" and "PERAKENDE" in normalized_text:
        preferred_headers = [
            "2026 PERAKENDE NAKİT",
            "2026 PERAKENDE 4 TAKSİT",
            "2026 PERAKENDE 6 TAKSİT",
        ]
    elif has_taksit_count(offer_text, 6) or payment_mode_hint == "6 TAKSIT":
        preferred_headers = [
            "2026 KURUMSAL 6 TAKSİT",
            HEADER_DEFAULT_PRICE,
        ]
    elif has_taksit_count(offer_text, 4) or payment_mode_hint == "4 TAKSIT":
        preferred_headers = [
            "2026 KURUMSAL 4 TAKSİT",
            HEADER_DEFAULT_PRICE,
        ]
    elif "KURUMSAL INDIRIMLI FIYAT" in normalized_text:
        preferred_headers = [
            "2026 KURUMSAL 6 TAKSİT",
            "2026 KURUMSAL 4 TAKSİT",
            HEADER_DEFAULT_PRICE,
        ]

    for preferred in preferred_headers:
        match = None
        if vat_included is False:
            match = normalized_available.get(normalize_text(f"KDV HARİÇ {preferred}"))
        if match is None:
            match = normalized_available.get(normalize_text(preferred))
        if match:
            return match

    return available_headers[0]


def run_comparison(
    price_list_path: Path,
    offer_path: Path,
    sheet_name: str | None = None,
    price_column: str | None = None,
    output_path: Path | None = None,
    min_match_score: float = 0.55,
    tolerance: float = 1.0,
) -> tuple[list[MatchResult], str, Path, list[str], FinancialReview]:
    price_rows, available_headers = load_price_rows(price_list_path, sheet_name)
    price_columns = get_price_columns(available_headers)
    if not price_columns:
        raise ValueError("Excel dosyasında kullanılabilir fiyat sütunu bulunamadı.")

    offer_items, offer_text = parse_offer_items(offer_path)
    detected_vat_included = detect_offer_vat_included(offer_text)
    if detected_vat_included is None:
        detected_vat_included = not is_kdv_excluded_header(price_column or "")
    detected_vat_rate, _ = detect_vat_rate(offer_text)

    if price_column:
        selected_column = resolve_price_column(price_column, price_columns)
    else:
        selected_column = detect_price_column(offer_text, price_columns)

    results = compare_offer_to_catalog(
        offer_items=offer_items,
        price_rows=price_rows,
        selected_column=selected_column,
        tolerance=tolerance,
        min_match_score=min_match_score,
        vat_included=detected_vat_included,
        vat_rate=detected_vat_rate,
    )
    financial_review = build_financial_review(offer_items, offer_text, tolerance=tolerance)

    final_output_path = output_path or build_report_output_path(price_list_path.parent, offer_path)
    write_report(
        final_output_path,
        results,
        price_list_path,
        offer_path,
        selected_column,
        financial_review=financial_review,
    )
    return results, selected_column, final_output_path, price_columns, financial_review


def build_corrected_pdf_path(offer_path: Path) -> Path:
    target_dir = ensure_output_subdir(offer_path.parent, CORRECTED_PDFS_DIRNAME)
    cleaned_stem = sanitize_filename_part(offer_path.stem) or "TEKLIF"
    return target_dir / f"{cleaned_stem}_duzeltilmis.pdf"


def derive_payment_info(selected_column: str) -> str:
    normalized = normalize_text(selected_column)
    if has_taksit_count(selected_column, 6):
        return "6 Taksit"
    if has_taksit_count(selected_column, 4):
        return "4 Taksit"
    if "NAKIT" in normalized:
        return "Nakit veya Banka Havalesi"
    return selected_column


def resolve_payment_info(selected_column: str, payment_info: str | None = None) -> str:
    explicit_payment_info = (payment_info or "").strip()
    if explicit_payment_info:
        return explicit_payment_info
    return derive_payment_info(selected_column)


def resolve_price_display_label(selected_column: str, price_label: str | None = None) -> str:
    explicit_price_label = (price_label or "").strip()
    if explicit_price_label:
        return explicit_price_label
    return selected_column


def _offer_number_prefix(for_date: date) -> str:
    return f"RW-{for_date:%d%m}-{for_date:%Y}"


def ensure_output_subdir(base_dir: Path, *parts: str) -> Path:
    path = base_dir / OUTPUT_ROOT_DIRNAME
    for part in parts:
        path /= part
    path.mkdir(parents=True, exist_ok=True)
    return path


def build_report_output_path(base_dir: Path, offer_path: Path | None = None) -> Path:
    target_dir = ensure_output_subdir(base_dir, REPORTS_DIRNAME)
    if offer_path is None:
        return target_dir / "teklif_kontrol_raporu.xlsx"
    cleaned_stem = sanitize_filename_part(offer_path.stem) or "TEKLIF"
    return target_dir / f"{cleaned_stem}_rapor.xlsx"


def _count_existing_offer_numbers(base_dir: Path, prefix: str) -> tuple[int, int]:
    sanitized_prefix = sanitize_filename_part(prefix)
    max_sequence = 0
    match_count = 0
    for path in base_dir.rglob("*.pdf"):
        if path.name.endswith("_duzeltilmis.pdf"):
            continue
        sanitized_stem = sanitize_filename_part(path.stem)
        if not (sanitized_stem == sanitized_prefix or sanitized_stem.startswith(f"{sanitized_prefix}_")):
            continue
        match_count += 1
        suffix = sanitized_stem[len(sanitized_prefix):].lstrip("_")
        if suffix:
            maybe_sequence = suffix.split("_", 1)[0]
            if maybe_sequence.isdigit():
                max_sequence = max(max_sequence, int(maybe_sequence))
    return match_count, max_sequence


def offer_number_exists(base_dir: Path, offer_number: str) -> bool:
    prefix = sanitize_filename_part(offer_number)
    if not prefix:
        return False
    for path in base_dir.rglob("*.pdf"):
        if path.name.endswith("_duzeltilmis.pdf"):
            continue
        sanitized_stem = sanitize_filename_part(path.stem)
        if sanitized_stem == prefix or sanitized_stem.startswith(f"{prefix}_"):
            return True
    return False


def default_offer_number(for_date: date | None = None, base_dir: Path | None = None) -> str:
    target_date = for_date or date.today()
    working_dir = base_dir or Path(__file__).resolve().parent
    prefix = _offer_number_prefix(target_date)
    match_count, max_sequence = _count_existing_offer_numbers(working_dir, prefix)
    next_sequence = (max(max_sequence, match_count) + 1) if (match_count or max_sequence) else 1
    return f"{prefix}-{next_sequence:03d}"


def default_valid_until(for_date: date | None = None, days: int = 7) -> date:
    target_date = for_date or date.today()
    return target_date + timedelta(days=days)


def sanitize_filename_part(value: str) -> str:
    normalized = normalize_text(value)
    if not normalized:
        return ""
    return re.sub(r"[^A-Z0-9]+", "_", normalized).strip("_")


def build_offer_output_path(base_dir: Path, offer_number: str, contact_name: str | None = None) -> Path:
    target_dir = ensure_output_subdir(base_dir, GENERATED_OFFERS_DIRNAME)
    parts = [sanitize_filename_part(offer_number or "TEKLIF")]
    contact_part = sanitize_filename_part(contact_name or "")
    if contact_part:
        parts.append(contact_part)
    filename = "_".join(parts) or "TEKLIF"
    return target_dir / f"{filename}.pdf"


def _format_quantity(quantity: float) -> str:
    if float(quantity).is_integer():
        return f"{int(quantity)} ADET"
    return f"{quantity:.2f}".replace(".", ",") + " ADET"


def _register_offer_fonts(page: fitz.Page) -> tuple[tuple[str, str | None], tuple[str, str | None]]:
    if ARIAL_FONT_PATH.exists():
        regular = ("rw_regular", str(ARIAL_FONT_PATH))
        bold = ("rw_bold", str(ARIAL_BOLD_FONT_PATH if ARIAL_BOLD_FONT_PATH.exists() else ARIAL_FONT_PATH))
        return regular, bold
    return ("helv", None), ("helv", None)


def _fill_rect(page: fitz.Page, rect: fitz.Rect, fill: tuple[float, float, float] = (1, 1, 1)) -> None:
    page.add_redact_annot(rect, fill=fill)


def _draw_textbox(
    page: fitz.Page,
    rect: fitz.Rect,
    text: str,
    *,
    fontname: str,
    fontfile: str | None = None,
    fontsize: float,
    color: tuple[float, float, float] = PDF_TEXT,
    align: int = 0,
) -> None:
    page.insert_textbox(
        rect,
        text,
        fontname=fontname,
        fontfile=fontfile,
        fontsize=fontsize,
        color=color,
        align=align,
    )


def _coerce_offer_selection(raw_selection: OfferSelection | tuple) -> OfferSelection:
    if isinstance(raw_selection, OfferSelection):
        return raw_selection
    if not isinstance(raw_selection, tuple):
        raise ValueError("Geçersiz ürün seçimi alındı.")
    if len(raw_selection) == 3:
        row_number, quantity, manual_price = raw_selection
        return OfferSelection(
            row_number=int(row_number),
            quantity=float(quantity),
            manual_price=manual_price,
        )
    if len(raw_selection) == 5:
        row_number, quantity, manual_price, discount_type, discount_value = raw_selection
        return OfferSelection(
            row_number=int(row_number),
            quantity=float(quantity),
            manual_price=manual_price,
            discount_type=str(discount_type or DISCOUNT_TYPE_NONE),
            discount_value=discount_value,
        )
    raise ValueError("Ürün seçimi biçimi desteklenmiyor.")


def _draw_offer_badges(
    target_page: fitz.Page,
    source_doc: fitz.Document,
    badge_clips: tuple[fitz.Rect, fitz.Rect],
    row_top: float,
    *,
    has_discount_layout: bool,
    background_fill: tuple[float, float, float] = (1, 1, 1),
) -> None:
    if has_discount_layout:
        target_rects = (
            fitz.Rect(184, row_top + 4, 212, row_top + 22),
            fitz.Rect(216, row_top + 4, 242, row_top + 22),
        )
    else:
        target_rects = (
            fitz.Rect(188, row_top + 4, 220, row_top + 24),
            fitz.Rect(226, row_top + 4, 255, row_top + 24),
        )

    badge_asset_paths = tuple(_offer_asset_path(name) for name in OFFER_BADGE_ASSET_NAMES)
    for asset_path, clip_rect, target_rect in zip(badge_asset_paths, badge_clips, target_rects, strict=False):
        try:
            target_page.draw_rect(target_rect, color=background_fill, fill=background_fill, overlay=True)
            if asset_path is not None:
                target_page.insert_image(
                    target_rect,
                    filename=str(asset_path),
                    keep_proportion=True,
                    overlay=True,
                )
            else:
                source_page = source_doc[0]
                badge_pixmap = source_page.get_pixmap(
                    matrix=fitz.Matrix(4, 4),
                    clip=clip_rect,
                    alpha=True,
                )
                target_page.insert_image(
                    target_rect,
                    pixmap=badge_pixmap,
                    keep_proportion=True,
                    overlay=True,
                )
        except Exception:
            return


def _draw_offer_header(
    target_page: fitz.Page,
    source_doc: fitz.Document,
    header_clip: fitz.Rect,
) -> None:
    try:
        header_asset_path = _offer_asset_path(OFFER_HEADER_ASSET_NAME)
        if header_asset_path is not None:
            target_page.insert_image(
                header_clip,
                filename=str(header_asset_path),
                keep_proportion=False,
                overlay=True,
            )
        else:
            source_page = source_doc[0]
            header_pixmap = source_page.get_pixmap(
                matrix=fitz.Matrix(3, 3),
                clip=header_clip,
                alpha=False,
            )
            target_page.insert_image(
                header_clip,
                pixmap=header_pixmap,
                keep_proportion=False,
                overlay=True,
            )
    except Exception:
        return


def _extract_badge_clips_from_doc(source_doc: fitz.Document) -> tuple[fitz.Rect, fitz.Rect] | None:
    if source_doc.page_count == 0:
        return None
    page = source_doc[0]
    text_dict = page.get_text("dict")
    candidate_rects: list[fitz.Rect] = []
    for block in text_dict.get("blocks", []):
        if block.get("type") != 1:
            continue
        rect = fitz.Rect(block.get("bbox"))
        if rect.x0 < 170 or rect.x1 > 270:
            continue
        if rect.y0 < 390 or rect.y1 > 460:
            continue
        if rect.width < 14 or rect.width > 40:
            continue
        if rect.height < 10 or rect.height > 28:
            continue
        candidate_rects.append(rect)

    if len(candidate_rects) < 2:
        return None

    unique_rects: list[fitz.Rect] = []
    for rect in sorted(candidate_rects, key=lambda item: (round(item.y0, 1), item.x0)):
        if any(abs(existing.x0 - rect.x0) < 2 and abs(existing.y0 - rect.y0) < 2 for existing in unique_rects):
            continue
        unique_rects.append(rect)

    if len(unique_rects) < 2:
        return None
    visible_rects = [
        _tighten_visible_rect(page, rect)
        for rect in unique_rects
        if _rect_has_visible_artwork(page, rect)
    ]
    if len(visible_rects) < 2:
        return None
    return visible_rects[0], visible_rects[1]


def _rect_has_visible_artwork(page: fitz.Page, rect: fitz.Rect) -> bool:
    try:
        pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), clip=rect, alpha=False)
    except Exception:
        return False
    samples = pix.samples
    if not samples:
        return False
    average = sum(samples) / len(samples)
    minimum = min(samples)
    return average < 240 or minimum < 200


def _tighten_visible_rect(
    page: fitz.Page,
    rect: fitz.Rect,
    *,
    scale: int = 4,
    threshold: int = 245,
) -> fitz.Rect:
    try:
        pix = page.get_pixmap(matrix=fitz.Matrix(scale, scale), clip=rect, alpha=False)
    except Exception:
        return rect

    width = pix.width
    height = pix.height
    channels = pix.n
    samples = pix.samples
    if width <= 0 or height <= 0 or not samples:
        return rect

    min_x = width
    min_y = height
    max_x = -1
    max_y = -1
    for y in range(height):
        row_offset = y * width * channels
        for x in range(width):
            offset = row_offset + x * channels
            pixel = samples[offset:offset + min(3, channels)]
            if any(channel < threshold for channel in pixel):
                min_x = min(min_x, x)
                min_y = min(min_y, y)
                max_x = max(max_x, x)
                max_y = max(max_y, y)

    if max_x < min_x or max_y < min_y:
        return rect

    x_unit = rect.width / width
    y_unit = rect.height / height
    tightened = fitz.Rect(
        rect.x0 + (min_x * x_unit),
        rect.y0 + (min_y * y_unit),
        rect.x0 + ((max_x + 1) * x_unit),
        rect.y0 + ((max_y + 1) * y_unit),
    )
    if tightened.width < 4 or tightened.height < 4:
        return rect
    return tightened


def _extract_header_clip_from_doc(source_doc: fitz.Document) -> fitz.Rect | None:
    if source_doc.page_count == 0:
        return None
    page = source_doc[0]
    if not _rect_has_visible_artwork(page, TEMPLATE_HEADER_CLIP):
        return None
    return TEMPLATE_HEADER_CLIP


def _resolve_badge_source(
    template_path: Path,
    template_doc: fitz.Document,
) -> tuple[fitz.Document, tuple[fitz.Rect, fitz.Rect], bool]:
    template_badges = _extract_badge_clips_from_doc(template_doc)
    if template_badges is not None:
        return template_doc, template_badges, False

    search_roots = [template_path.parent, Path.cwd()]
    seen_paths = {template_path.resolve()}
    for root in search_roots:
        if not root.exists():
            continue
        for candidate in sorted(root.glob("*.pdf")):
            try:
                resolved_candidate = candidate.resolve()
            except Exception:
                resolved_candidate = candidate.absolute()
            if resolved_candidate in seen_paths:
                continue
            seen_paths.add(resolved_candidate)
            if candidate.name.endswith("_duzeltilmis.pdf"):
                continue
            try:
                donor_doc = fitz.open(candidate)
            except Exception:
                continue
            donor_badges = _extract_badge_clips_from_doc(donor_doc)
            if donor_badges is not None:
                return donor_doc, donor_badges, True
            donor_doc.close()

    return template_doc, TEMPLATE_BADGE_CLIPS, False


def _resolve_header_source(
    template_path: Path,
    template_doc: fitz.Document,
) -> tuple[fitz.Document, fitz.Rect, bool]:
    template_header = _extract_header_clip_from_doc(template_doc)
    if template_header is not None:
        return template_doc, template_header, False

    search_roots = [template_path.parent, Path.cwd()]
    seen_paths = {template_path.resolve()}
    for root in search_roots:
        if not root.exists():
            continue
        for candidate in sorted(root.glob("*.pdf")):
            try:
                resolved_candidate = candidate.resolve()
            except Exception:
                resolved_candidate = candidate.absolute()
            if resolved_candidate in seen_paths:
                continue
            seen_paths.add(resolved_candidate)
            if candidate.name.endswith("_duzeltilmis.pdf"):
                continue
            try:
                donor_doc = fitz.open(candidate)
            except Exception:
                continue
            donor_header = _extract_header_clip_from_doc(donor_doc)
            if donor_header is not None:
                return donor_doc, donor_header, True
            donor_doc.close()

    return template_doc, TEMPLATE_HEADER_CLIP, False


def _legacy_build_offer_items_from_selection(
    price_rows: list[PriceRow],
    selected_column: str,
    selected_entries: list[OfferSelection | tuple],
    *,
    vat_included: bool = True,
    vat_rate: float = DEFAULT_VAT_RATE,
) -> list[OfferLineItem]:
    if not selected_entries:
        raise ValueError("Teklif oluşturmak için en az bir ürün seç.")

    combined_quantities: dict[int, float] = {}
    row_config_by_number: dict[int, tuple[float | None, str, float | None]] = {}
    for raw_selection in selected_entries:
        selection = _coerce_offer_selection(raw_selection)
        if selection.quantity <= 0:
            raise ValueError("Ürün adetleri sıfırdan büyük olmalı.")
        normalized_discount_type = str(selection.discount_type or DISCOUNT_TYPE_NONE).strip().lower()
        if normalized_discount_type not in DISCOUNT_TYPES:
            raise ValueError("Geçersiz iskonto tipi seçildi.")

        normalized_manual_price = round(selection.manual_price, 2) if selection.manual_price is not None else None
        if False:
            raise ValueError("Aynı ürün için birden fazla farklı manuel fiyat girildi.")
        normalized_discount_value = round(selection.discount_value, 2) if selection.discount_value is not None else None
        if normalized_discount_type != DISCOUNT_TYPE_NONE and normalized_discount_value is None:
            raise ValueError("İskonto seçildiyse değeri de girilmeli.")

        combined_quantities[selection.row_number] = round(
            combined_quantities.get(selection.row_number, 0) + selection.quantity,
            2,
        )
        row_config = (normalized_manual_price, normalized_discount_type, normalized_discount_value)
        if selection.row_number in row_config_by_number and row_config_by_number[selection.row_number] != row_config:
            raise ValueError("Aynı ürün için farklı fiyat/iskonto kurguları girildi.")
        row_config_by_number[selection.row_number] = row_config

    if len(combined_quantities) > MAX_GENERATED_OFFER_ITEMS:
        raise ValueError(f"Tek sayfa şablonda en fazla {MAX_GENERATED_OFFER_ITEMS} ürün destekleniyor.")

    rows_by_number = {row.row_number: row for row in price_rows}
    offer_items: list[OfferLineItem] = []
    for row_number, quantity in combined_quantities.items():
        price_row = rows_by_number.get(row_number)
        if price_row is None:
            raise ValueError(f"Seçilen ürün satırı bulunamadı: {row_number}")
        standard_header = derive_standard_price_header(selected_column, price_row.prices.keys())
        reference_unit_price, _, _ = resolve_price_for_vat_mode(
            price_row,
            selected_column,
            vat_included=vat_included,
            vat_rate=vat_rate,
        )
        base_unit_price = None
        if standard_header is not None:
            base_unit_price, _, _ = resolve_price_for_vat_mode(
                price_row,
                standard_header,
                vat_included=vat_included,
                vat_rate=vat_rate,
            )
        manual_unit_price = manual_prices_by_row.get(row_number)
        unit_price = manual_unit_price if manual_unit_price is not None else reference_unit_price
        if unit_price is None:
            raise ValueError(f"'{price_row.product_name}' için '{selected_column}' kolonunda fiyat yok.")
        if base_unit_price is None:
            base_unit_price = unit_price
        discount_amount = round(max(base_unit_price - unit_price, 0.0), 2)
        offer_items.append(
            OfferLineItem(
                row_number=row_number,
                product_name=price_row.product_name,
                quantity=quantity,
                reference_unit_price=reference_unit_price,
                base_unit_price=base_unit_price,
                unit_price=unit_price,
                total_price=round(unit_price * quantity, 2),
                discount_amount=discount_amount,
                price_source="manual" if manual_unit_price is not None else "list",
            )
        )

    offer_items.sort(key=lambda item: item.row_number)
    return offer_items


def _legacy_create_offer_from_catalog(
    *,
    template_path: Path,
    price_list_path: Path,
    selected_column: str,
    selected_entries: list[tuple[int, float, float | None]],
    vat_included: bool = True,
    vat_rate: float = DEFAULT_VAT_RATE,
    offer_number: str,
    offer_date: date,
    valid_until: date,
    company_name: str,
    contact_name: str,
    email: str,
    gsm: str,
    note_text: str | None = None,
    payment_info: str | None = None,
    price_label: str | None = None,
    output_path: Path | None = None,
) -> Path:
    price_rows, available_headers = load_price_rows(price_list_path)
    resolved_column = resolve_price_column(selected_column, get_price_columns(available_headers))
    offer_items = build_offer_items_from_selection(
        price_rows,
        resolved_column,
        selected_entries,
        vat_included=vat_included,
        vat_rate=vat_rate,
    )
    final_output_path = output_path or build_offer_output_path(RUNTIME_BASE_DIR, offer_number, contact_name)
    return generate_offer_pdf(
        template_path=template_path,
        output_path=final_output_path,
        offer_number=offer_number,
        offer_date=offer_date,
        valid_until=valid_until,
        company_name=company_name,
        contact_name=contact_name,
        email=email,
        gsm=gsm,
        selected_column=resolved_column,
        vat_included=vat_included,
        vat_rate=vat_rate,
        offer_items=offer_items,
        note_text=note_text,
        payment_info=payment_info,
        price_label=price_label,
    )


def build_offer_items_from_selection(
    price_rows: list[PriceRow],
    selected_column: str,
    selected_entries: list[OfferSelection | tuple],
    *,
    vat_included: bool = True,
    vat_rate: float = DEFAULT_VAT_RATE,
) -> list[OfferLineItem]:
    if not selected_entries:
        raise ValueError("Teklif oluşturmak için en az bir ürün seç.")

    combined_quantities: dict[int, float] = {}
    row_config_by_number: dict[int, tuple[float | None, str, float | None]] = {}
    for raw_selection in selected_entries:
        selection = _coerce_offer_selection(raw_selection)
        if selection.quantity <= 0:
            raise ValueError("Ürün adetleri sıfırdan büyük olmalı.")

        normalized_discount_type = str(selection.discount_type or DISCOUNT_TYPE_NONE).strip().lower()
        if normalized_discount_type not in DISCOUNT_TYPES:
            raise ValueError("Geçersiz iskonto tipi seçildi.")

        normalized_manual_price = round(selection.manual_price, 2) if selection.manual_price is not None else None
        normalized_discount_value = round(selection.discount_value, 2) if selection.discount_value is not None else None
        if normalized_discount_type != DISCOUNT_TYPE_NONE and normalized_discount_value is None:
            raise ValueError("İskonto seçildiyse değeri de girilmeli.")

        combined_quantities[selection.row_number] = round(
            combined_quantities.get(selection.row_number, 0) + selection.quantity,
            2,
        )
        row_config = (normalized_manual_price, normalized_discount_type, normalized_discount_value)
        if selection.row_number in row_config_by_number and row_config_by_number[selection.row_number] != row_config:
            raise ValueError("Aynı ürün için farklı fiyat veya iskonto kurguları girildi.")
        row_config_by_number[selection.row_number] = row_config

    if len(combined_quantities) > MAX_GENERATED_OFFER_ITEMS:
        raise ValueError(f"Tek sayfa şablonda en fazla {MAX_GENERATED_OFFER_ITEMS} ürün destekleniyor.")

    rows_by_number = {row.row_number: row for row in price_rows}
    offer_items: list[OfferLineItem] = []
    for row_number, quantity in combined_quantities.items():
        price_row = rows_by_number.get(row_number)
        if price_row is None:
            raise ValueError(f"Seçilen ürün satırı bulunamadı: {row_number}")

        standard_header = derive_standard_price_header(selected_column, price_row.prices.keys())
        reference_unit_price, _, _ = resolve_price_for_vat_mode(
            price_row,
            selected_column,
            vat_included=vat_included,
            vat_rate=vat_rate,
        )
        base_unit_price = None
        if standard_header is not None:
            base_unit_price, _, _ = resolve_price_for_vat_mode(
                price_row,
                standard_header,
                vat_included=vat_included,
                vat_rate=vat_rate,
            )

        manual_unit_price, discount_type, discount_value = row_config_by_number.get(
            row_number,
            (None, DISCOUNT_TYPE_NONE, None),
        )
        if base_unit_price is None:
            base_unit_price = reference_unit_price

        if discount_type != DISCOUNT_TYPE_NONE:
            if base_unit_price is None:
                raise ValueError(f"'{price_row.product_name}' için iskonto uygulanacak standart fiyat bulunamadı.")
            if discount_value is None or discount_value < 0:
                raise ValueError(f"'{price_row.product_name}' için iskonto değeri geçersiz.")
            if discount_type == DISCOUNT_TYPE_PERCENT:
                if discount_value > 100:
                    raise ValueError(f"'{price_row.product_name}' için iskonto yüzdesi 100'ü aşamaz.")
                discount_amount = round(base_unit_price * discount_value / 100, 2)
            else:
                if discount_value > round(base_unit_price, 2):
                    raise ValueError(f"'{price_row.product_name}' için iskonto tutarı baz fiyatı aşamaz.")
                discount_amount = round(discount_value, 2)
            unit_price = round(base_unit_price - discount_amount, 2)
            if unit_price < 0:
                raise ValueError(f"'{price_row.product_name}' için indirimli fiyat sıfırın altına düşemez.")
            price_source = "discount"
        else:
            unit_price = manual_unit_price if manual_unit_price is not None else reference_unit_price
            if unit_price is None:
                raise ValueError(f"'{price_row.product_name}' için '{selected_column}' kolonunda fiyat yok.")
            if manual_unit_price is not None:
                base_unit_price = unit_price
            elif reference_unit_price is not None:
                base_unit_price = reference_unit_price
            elif base_unit_price is None:
                base_unit_price = unit_price
            discount_amount = 0.0
            price_source = "manual" if manual_unit_price is not None else "list"

        offer_items.append(
            OfferLineItem(
                row_number=row_number,
                product_name=price_row.product_name,
                quantity=quantity,
                reference_unit_price=reference_unit_price,
                base_unit_price=base_unit_price,
                unit_price=unit_price,
                total_price=round(unit_price * quantity, 2),
                discount_amount=discount_amount,
                price_source=price_source,
            )
        )

    offer_items.sort(key=lambda item: item.row_number)
    return offer_items


def create_offer_from_catalog(
    *,
    template_path: Path,
    price_list_path: Path,
    selected_column: str,
    selected_entries: list[OfferSelection | tuple],
    vat_included: bool = True,
    vat_rate: float = DEFAULT_VAT_RATE,
    offer_number: str,
    offer_date: date,
    valid_until: date,
    company_name: str,
    contact_name: str,
    email: str,
    gsm: str,
    note_text: str | None = None,
    payment_info: str | None = None,
    price_label: str | None = None,
    output_path: Path | None = None,
) -> Path:
    price_rows, available_headers = load_price_rows(price_list_path)
    resolved_column = resolve_price_column(selected_column, get_price_columns(available_headers))
    offer_items = build_offer_items_from_selection(
        price_rows,
        resolved_column,
        selected_entries,
        vat_included=vat_included,
        vat_rate=vat_rate,
    )
    final_output_path = output_path or build_offer_output_path(RUNTIME_BASE_DIR, offer_number, contact_name)
    return generate_offer_pdf(
        template_path=template_path,
        output_path=final_output_path,
        offer_number=offer_number,
        offer_date=offer_date,
        valid_until=valid_until,
        company_name=company_name,
        contact_name=contact_name,
        email=email,
        gsm=gsm,
        selected_column=resolved_column,
        vat_included=vat_included,
        vat_rate=vat_rate,
        offer_items=offer_items,
        note_text=note_text,
        payment_info=payment_info,
        price_label=price_label,
    )


def _legacy_generate_offer_pdf(
    *,
    template_path: Path,
    output_path: Path,
    offer_number: str,
    offer_date: date,
    valid_until: date,
    company_name: str,
    contact_name: str,
    email: str,
    gsm: str,
    selected_column: str,
    vat_included: bool = True,
    vat_rate: float = DEFAULT_VAT_RATE,
    offer_items: list[OfferLineItem],
    note_text: str | None = None,
    payment_info: str | None = None,
    price_label: str | None = None,
) -> Path:
    if not offer_items:
        raise ValueError("Teklif oluşturmak için ürün bulunamadı.")
    if len(offer_items) > MAX_GENERATED_OFFER_ITEMS:
        raise ValueError(f"Tek sayfa şablonda en fazla {MAX_GENERATED_OFFER_ITEMS} ürün destekleniyor.")
    if not template_path.exists():
        raise FileNotFoundError(f"Şablon PDF bulunamadı: {template_path}")
    if template_path.suffix.lower() != ".pdf":
        raise ValueError(f"Desteklenmeyen sablon formati: {template_path.name}")

    final_output_path = _ensure_safe_output_path(
        output_path,
        expected_suffix=".pdf",
        label="Teklif cikisi",
        source_paths=(template_path,),
    )
    final_output_path.parent.mkdir(parents=True, exist_ok=True)

    source = fitz.open(template_path)
    badge_source = source
    badge_clips = TEMPLATE_BADGE_CLIPS
    close_badge_source = False
    header_source = source
    header_clip = TEMPLATE_HEADER_CLIP
    close_header_source = False
    doc = fitz.open()
    try:
        header_source, header_clip, close_header_source = _resolve_header_source(template_path, source)
        badge_source, badge_clips, close_badge_source = _resolve_badge_source(template_path, source)
        doc.insert_pdf(source, from_page=0, to_page=0)
        page = doc[0]
        (regular_font, regular_fontfile), (bold_font, bold_fontfile) = _register_offer_fonts(page)

        clear_rects = [
            fitz.Rect(0, 90, 540, 182),
            fitz.Rect(0, 202, 540, 700),
            fitz.Rect(330, 15, 540, 72),
            OFFER_SIGNATURE_CLEAR_RECT,
        ]
        for rect in clear_rects:
            _fill_rect(page, rect)
        page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)
        _draw_offer_header(page, header_source, header_clip)
        _fill_rect(page, fitz.Rect(330, 15, 540, 72))
        page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)

        page.draw_line(fitz.Point(23, 188), fitz.Point(518, 188), color=PDF_ACCENT, width=1)
        page.draw_line(fitz.Point(23, 341), fitz.Point(518, 341), color=PDF_ACCENT, width=1)

        _draw_textbox(
            page,
            fitz.Rect(340, 18, 520, 38),
            "FİYAT TEKLİFİ",
            fontname=bold_font,
            fontfile=bold_fontfile,
            fontsize=15,
            align=2,
        )
        _draw_textbox(
            page,
            fitz.Rect(340, 39, 520, 54),
            f"Teklif No: {offer_number}",
            fontname=regular_font,
            fontfile=regular_fontfile,
            fontsize=10,
            align=2,
        )
        _draw_textbox(
            page,
            fitz.Rect(340, 54, 520, 69),
            f"Teklif Tarihi: {offer_date:%d.%m.%Y}",
            fontname=regular_font,
            fontfile=regular_fontfile,
            fontsize=10,
            align=2,
        )

        customer_labels = [
            ("Firma/ Bireysel", company_name or "-"),
            ("Yetkili Adı", contact_name or "-"),
            ("E Mail", email or "-"),
            ("GSM", gsm or "-"),
        ]
        label_y = 98
        for label, value in customer_labels:
            _draw_textbox(page, fitz.Rect(24, label_y, 122, label_y + 16), label, fontname=bold_font, fontfile=bold_fontfile, fontsize=9.5)
            _draw_textbox(page, fitz.Rect(129, label_y, 132, label_y + 16), ":", fontname=bold_font, fontfile=bold_fontfile, fontsize=10)
            _draw_textbox(page, fitz.Rect(138, label_y, 320, label_y + 16), value, fontname=regular_font, fontfile=regular_fontfile, fontsize=10)
            label_y += 20

        vat_mode_label = "KDV Dahil" if vat_included else "KDV Hariç"
        price_display_label = resolve_price_display_label(selected_column, price_label)
        payment_info_text = resolve_payment_info(selected_column, payment_info)
        bullet_lines = [
            "Garanti : Sistemlerimiz 10 YIL RAINWATER GARANTİSİ altındadır.",
            "Kullanıcı hataları dışında, elektrik motorları 2 yıldır.",
            "Teslim süresi : Katı siparişi 7 iş günü",
            f"Ödeme Bilgisi : {payment_info_text}",
            "Montaj Bilgisi : Montaj alanına elektrik hat çekimi tarafınıza aittir.",
            "Ürün nakliye ve montajı tarafımıza aittir.",
            f"Fiyatlarımıza KDV (%{vat_rate:g}) {'dahildir' if vat_included else 'dahil değildir'}.",
            f"Teklif süresi : {valid_until:%d.%m.%Y} tarihine kadar geçerlidir.",
        ]
        bullet_y = 208
        for line in bullet_lines:
            _draw_textbox(page, fitz.Rect(35, bullet_y, 515, bullet_y + 13), f"•  {line}", fontname=regular_font, fontfile=regular_fontfile, fontsize=9.2)
            bullet_y += 16

        has_discount_layout = any(item.discount_amount > 0.01 for item in offer_items)
        total_discount_amount = round(
            sum(item.discount_amount * item.quantity for item in offer_items),
            2,
        )
        discounted_header = (
            "KURUMSAL\nİNDİRİMLİ\nFİYAT"
            if "KURUMSAL" in normalize_text(selected_column)
            else "İNDİRİMLİ\nFİYAT"
        )
        table_header_top = 350
        page.draw_line(fitz.Point(20, table_header_top - 4), fitz.Point(520, table_header_top - 4), color=PDF_ACCENT, width=1)
        if has_discount_layout:
            _draw_textbox(page, fitz.Rect(23, table_header_top, 226, table_header_top + 18), "MALZEME", fontname=bold_font, fontfile=bold_fontfile, fontsize=10)
            _draw_textbox(page, fitz.Rect(228, table_header_top, 286, table_header_top + 18), "MİKTAR", fontname=bold_font, fontfile=bold_fontfile, fontsize=9.3, align=1)
            _draw_textbox(page, fitz.Rect(288, table_header_top, 352, table_header_top + 26), "BİRİM\nFİYAT", fontname=bold_font, fontfile=bold_fontfile, fontsize=8.6, align=1)
            _draw_textbox(page, fitz.Rect(354, table_header_top, 406, table_header_top + 26), "İSKONTO\nTUTARI", fontname=bold_font, fontfile=bold_fontfile, fontsize=8.2, align=1)
            _draw_textbox(page, fitz.Rect(408, table_header_top - 1, 470, table_header_top + 30), discounted_header, fontname=bold_font, fontfile=bold_fontfile, fontsize=7.8, align=1)
            _draw_textbox(page, fitz.Rect(472, table_header_top, 518, table_header_top + 26), "TOPLAM\nTUTAR", fontname=bold_font, fontfile=bold_fontfile, fontsize=8.3, align=2)
        else:
            _draw_textbox(page, fitz.Rect(23, table_header_top, 250, table_header_top + 16), "MALZEME", fontname=bold_font, fontfile=bold_fontfile, fontsize=10)
            _draw_textbox(page, fitz.Rect(252, table_header_top, 320, table_header_top + 16), "MİKTAR", fontname=bold_font, fontfile=bold_fontfile, fontsize=10, align=1)
            _draw_textbox(page, fitz.Rect(322, table_header_top, 388, table_header_top + 16), "BİRİM FİYAT", fontname=bold_font, fontfile=bold_fontfile, fontsize=9.4, align=1)
            _draw_textbox(page, fitz.Rect(390, table_header_top, 458, table_header_top + 16), "SEÇİLEN FİYAT", fontname=bold_font, fontfile=bold_fontfile, fontsize=9.2, align=1)
            _draw_textbox(page, fitz.Rect(460, table_header_top, 518, table_header_top + 16), "TOPLAM", fontname=bold_font, fontfile=bold_fontfile, fontsize=10, align=2)
        _draw_textbox(page, fitz.Rect(23, table_header_top + 24, 518, table_header_top + 38), f"Seçilen fiyat tipi: {price_display_label} | {vat_mode_label}", fontname=regular_font, fontfile=regular_fontfile, fontsize=8.7, color=PDF_MUTED)
        page.draw_line(fitz.Point(23, table_header_top + 44), fitz.Point(518, table_header_top + 44), color=PDF_LIGHT, width=0.8)

        row_top = table_header_top + 54
        row_height = 32 if has_discount_layout else 30
        for row_index, item in enumerate(offer_items):
            row_bottom = row_top + row_height
            if row_index % 2 == 0:
                page.draw_rect(fitz.Rect(20, row_top - 2, 520, row_bottom), color=PDF_ROW_SOFT, fill=PDF_ROW_SOFT, overlay=True)
            _draw_textbox(
                page,
                fitz.Rect(23, row_top, 182 if has_discount_layout else 212, row_bottom),
                item.product_name,
                fontname=bold_font,
                fontfile=bold_fontfile,
                fontsize=8.8,
            )
            _draw_offer_badges(
                page,
                badge_source,
                badge_clips,
                row_top,
                has_discount_layout=has_discount_layout,
                background_fill=PDF_ROW_SOFT if row_index % 2 == 0 else (1, 1, 1),
            )
            _draw_textbox(
                page,
                fitz.Rect(228 if has_discount_layout else 260, row_top + 2, 286 if has_discount_layout else 320, row_bottom),
                _format_quantity(item.quantity),
                fontname=regular_font,
                fontfile=regular_fontfile,
                fontsize=8.8,
                align=1,
            )
            if has_discount_layout:
                _draw_textbox(page, fitz.Rect(288, row_top + 2, 352, row_bottom), format_pdf_money(item.base_unit_price), fontname=regular_font, fontfile=regular_fontfile, fontsize=8.2, align=1)
                _draw_textbox(page, fitz.Rect(354, row_top + 2, 406, row_bottom), format_pdf_money(item.discount_amount), fontname=regular_font, fontfile=regular_fontfile, fontsize=8.2, align=1)
                _draw_textbox(page, fitz.Rect(408, row_top + 2, 470, row_bottom), format_pdf_money(item.unit_price), fontname=regular_font, fontfile=regular_fontfile, fontsize=8.2, align=1)
                _draw_textbox(page, fitz.Rect(472, row_top + 2, 518, row_bottom), format_pdf_money(item.total_price), fontname=bold_font, fontfile=bold_fontfile, fontsize=8.4, align=2)
            else:
                _draw_textbox(page, fitz.Rect(322, row_top + 2, 388, row_bottom), format_pdf_money(item.base_unit_price or item.unit_price), fontname=regular_font, fontfile=regular_fontfile, fontsize=8.6, align=1)
                _draw_textbox(page, fitz.Rect(390, row_top + 2, 458, row_bottom), format_pdf_money(item.unit_price), fontname=regular_font, fontfile=regular_fontfile, fontsize=8.6, align=1)
                _draw_textbox(page, fitz.Rect(460, row_top + 2, 518, row_bottom), format_pdf_money(item.total_price), fontname=bold_font, fontfile=bold_fontfile, fontsize=8.8, align=2)
            page.draw_line(fitz.Point(23, row_bottom), fitz.Point(518, row_bottom), color=PDF_LIGHT, width=0.6)
            row_top += row_height

        line_total_sum = round(sum(item.total_price for item in offer_items), 2)
        net_total, vat_total, gross_total = calculate_offer_totals(
            line_total_sum,
            vat_rate=vat_rate,
            vat_included=vat_included,
        )
        totals_top = max(500, row_top + 18)
        if vat_included:
            _draw_textbox(page, fitz.Rect(320, totals_top, 445, totals_top + 16), "YATIRIM MALİYETİ", fontname=bold_font, fontfile=bold_fontfile, fontsize=10, align=2)
            _draw_textbox(page, fitz.Rect(448, totals_top, 518, totals_top + 16), f": {format_pdf_money(net_total)}", fontname=regular_font, fontfile=regular_fontfile, fontsize=10, align=2)
            _draw_textbox(page, fitz.Rect(320, totals_top + 20, 445, totals_top + 36), f"KDV (%{vat_rate:g})", fontname=bold_font, fontfile=bold_fontfile, fontsize=10, align=2)
            _draw_textbox(page, fitz.Rect(448, totals_top + 20, 518, totals_top + 36), f": {format_pdf_money(vat_total)}", fontname=regular_font, fontfile=regular_fontfile, fontsize=10, align=2)
            _draw_textbox(page, fitz.Rect(280, totals_top + 42, 445, totals_top + 60), "TOPLAM YATIRIM MALİYETİ", fontname=bold_font, fontfile=bold_fontfile, fontsize=11, color=(1, 0, 0), align=2)
            _draw_textbox(page, fitz.Rect(448, totals_top + 42, 518, totals_top + 60), f": {format_pdf_money(gross_total)}", fontname=bold_font, fontfile=bold_fontfile, fontsize=11, color=(1, 0, 0), align=2)
        else:
            _draw_textbox(page, fitz.Rect(276, totals_top, 445, totals_top + 16), "YATIRIM MALİYETİ (KDV HARİÇ)", fontname=bold_font, fontfile=bold_fontfile, fontsize=9.4, align=2)
            _draw_textbox(page, fitz.Rect(448, totals_top, 518, totals_top + 16), f": {format_pdf_money(net_total)}", fontname=regular_font, fontfile=regular_fontfile, fontsize=10, align=2)
            _draw_textbox(page, fitz.Rect(320, totals_top + 20, 445, totals_top + 36), f"KDV (%{vat_rate:g})", fontname=bold_font, fontfile=bold_fontfile, fontsize=10, align=2)
            _draw_textbox(page, fitz.Rect(448, totals_top + 20, 518, totals_top + 36), f": {format_pdf_money(vat_total)}", fontname=regular_font, fontfile=regular_fontfile, fontsize=10, align=2)
            _draw_textbox(page, fitz.Rect(230, totals_top + 42, 445, totals_top + 60), "TOPLAM YATIRIM MALİYETİ (KDV HARİÇ)", fontname=bold_font, fontfile=bold_fontfile, fontsize=9.1, color=(1, 0, 0), align=2)
            _draw_textbox(page, fitz.Rect(448, totals_top + 42, 518, totals_top + 60), f": {format_pdf_money(net_total)}", fontname=bold_font, fontfile=bold_fontfile, fontsize=11, color=(1, 0, 0), align=2)

        if note_text and note_text.strip():
            note_top = min(totals_top + 86, 612)
            _draw_textbox(page, fitz.Rect(24, note_top, 518, note_top + 42), f"•  {note_text.strip()}", fontname=regular_font, fontfile=regular_fontfile, fontsize=9.2)

        _draw_textbox(page, fitz.Rect(342, 654, 510, 668), "Yetkili Adı", fontname=regular_font, fontfile=regular_fontfile, fontsize=10, align=1)
        _draw_textbox(page, fitz.Rect(342, 674, 510, 688), (contact_name or company_name or "-").upper(), fontname=bold_font, fontfile=bold_fontfile, fontsize=10.5, align=1)

        doc.save(final_output_path)
        return final_output_path
    finally:
        if close_header_source and header_source is not source:
            header_source.close()
        if close_badge_source and badge_source is not source:
            badge_source.close()
        source.close()
        doc.close()


def generate_offer_pdf(
    *,
    template_path: Path,
    output_path: Path,
    offer_number: str,
    offer_date: date,
    valid_until: date,
    company_name: str,
    contact_name: str,
    email: str,
    gsm: str,
    selected_column: str,
    vat_included: bool = True,
    vat_rate: float = DEFAULT_VAT_RATE,
    offer_items: list[OfferLineItem],
    note_text: str | None = None,
    payment_info: str | None = None,
    price_label: str | None = None,
) -> Path:
    if not offer_items:
        raise ValueError("Teklif oluşturmak için ürün bulunamadı.")
    if len(offer_items) > MAX_GENERATED_OFFER_ITEMS:
        raise ValueError(f"Tek sayfa şablonda en fazla {MAX_GENERATED_OFFER_ITEMS} ürün destekleniyor.")
    if not template_path.exists():
        raise FileNotFoundError(f"Şablon PDF bulunamadı: {template_path}")
    if template_path.suffix.lower() != ".pdf":
        raise ValueError(f"Desteklenmeyen sablon formati: {template_path.name}")

    final_output_path = _ensure_safe_output_path(
        output_path,
        expected_suffix=".pdf",
        label="Teklif cikisi",
        source_paths=(template_path,),
    )
    final_output_path.parent.mkdir(parents=True, exist_ok=True)

    source = fitz.open(template_path)
    badge_source = source
    badge_clips = TEMPLATE_BADGE_CLIPS
    close_badge_source = False
    header_source = source
    header_clip = TEMPLATE_HEADER_CLIP
    close_header_source = False
    doc = fitz.open()
    try:
        header_source, header_clip, close_header_source = _resolve_header_source(template_path, source)
        badge_source, badge_clips, close_badge_source = _resolve_badge_source(template_path, source)
        doc.insert_pdf(source, from_page=0, to_page=0)
        page = doc[0]
        (regular_font, regular_fontfile), (bold_font, bold_fontfile) = _register_offer_fonts(page)

        clear_rects = [
            fitz.Rect(0, 90, 540, 182),
            fitz.Rect(0, 202, 540, 700),
            fitz.Rect(330, 15, 540, 72),
            OFFER_SIGNATURE_CLEAR_RECT,
        ]
        for rect in clear_rects:
            _fill_rect(page, rect)
        page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)
        _draw_offer_header(page, header_source, header_clip)
        _fill_rect(page, fitz.Rect(330, 15, 540, 72))
        page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)

        page.draw_line(fitz.Point(23, 188), fitz.Point(518, 188), color=PDF_ACCENT, width=1)
        page.draw_line(fitz.Point(23, 341), fitz.Point(518, 341), color=PDF_ACCENT, width=1)

        _draw_textbox(
            page,
            fitz.Rect(340, 18, 520, 38),
            "FİYAT TEKLİFİ",
            fontname=bold_font,
            fontfile=bold_fontfile,
            fontsize=15,
            align=2,
        )
        _draw_textbox(
            page,
            fitz.Rect(340, 39, 520, 54),
            f"Teklif No: {offer_number}",
            fontname=bold_font,
            fontfile=bold_fontfile,
            fontsize=10.5,
            align=2,
        )
        _draw_textbox(
            page,
            fitz.Rect(340, 54, 520, 69),
            f"Tarih: {offer_date:%d.%m.%Y}",
            fontname=regular_font,
            fontfile=regular_fontfile,
            fontsize=9.5,
            align=2,
        )

        customer_labels = [
            ("Firma/ Bireysel", company_name or "-"),
            ("Yetkili Adı", contact_name or "-"),
            ("E Mail", email or "-"),
            ("GSM", gsm or "-"),
        ]
        label_y = 98
        for label, value in customer_labels:
            _draw_textbox(page, fitz.Rect(24, label_y, 122, label_y + 16), label, fontname=bold_font, fontfile=bold_fontfile, fontsize=9.5)
            _draw_textbox(page, fitz.Rect(129, label_y, 132, label_y + 16), ":", fontname=bold_font, fontfile=bold_fontfile, fontsize=10)
            _draw_textbox(page, fitz.Rect(138, label_y, 320, label_y + 16), value, fontname=regular_font, fontfile=regular_fontfile, fontsize=10)
            label_y += 20

        vat_mode_label = "KDV Dahil" if vat_included else "KDV Hariç"
        price_display_label = resolve_price_display_label(selected_column, price_label)
        payment_info_text = resolve_payment_info(selected_column, payment_info)
        bullet_lines = [
            "Garanti : Sistemlerimiz 10 YIL RAINWATER GARANTİSİ altındadır.",
            "Kullanıcı hataları dışında, elektrik motorları 2 yıldır.",
            "Teslim süresi : Katı siparişi 7 iş günü",
            f"Ödeme Bilgisi : {payment_info_text}",
            "Montaj Bilgisi : Montaj alanına elektrik hat çekimi tarafınıza aittir.",
            "Ürün nakliye ve montajı tarafımıza aittir.",
            f"Fiyatlarımıza KDV (%{vat_rate:g}) {'dahildir' if vat_included else 'dahil değildir'}.",
            f"Teklif süresi : {valid_until:%d.%m.%Y} tarihine kadar geçerlidir.",
        ]

        bullet_y = 208
        for line in bullet_lines:
            _draw_textbox(page, fitz.Rect(35, bullet_y, 515, bullet_y + 13), f"•  {line}", fontname=regular_font, fontfile=regular_fontfile, fontsize=9.2)
            bullet_y += 16

        has_discount_layout = any(item.discount_amount > 0.01 for item in offer_items)
        total_discount_amount = round(
            sum(item.discount_amount * item.quantity for item in offer_items),
            2,
        )
        discounted_header = (
            "KURUMSAL\nİNDİRİMLİ\nFİYAT"
            if "KURUMSAL" in normalize_text(selected_column)
            else "İNDİRİMLİ\nFİYAT"
        )
        table_header_top = 350
        page.draw_line(fitz.Point(20, table_header_top - 4), fitz.Point(520, table_header_top - 4), color=PDF_ACCENT, width=1)
        if has_discount_layout:
            _draw_textbox(page, fitz.Rect(23, table_header_top, 226, table_header_top + 18), "MALZEME", fontname=bold_font, fontfile=bold_fontfile, fontsize=10)
            _draw_textbox(page, fitz.Rect(228, table_header_top, 286, table_header_top + 18), "MİKTAR", fontname=bold_font, fontfile=bold_fontfile, fontsize=9.3, align=1)
            _draw_textbox(page, fitz.Rect(288, table_header_top, 352, table_header_top + 26), "BİRİM\nFİYAT", fontname=bold_font, fontfile=bold_fontfile, fontsize=8.6, align=1)
            _draw_textbox(page, fitz.Rect(354, table_header_top, 406, table_header_top + 26), "İSKONTO\nTUTARI", fontname=bold_font, fontfile=bold_fontfile, fontsize=8.2, align=1)
            _draw_textbox(page, fitz.Rect(408, table_header_top - 1, 470, table_header_top + 30), discounted_header, fontname=bold_font, fontfile=bold_fontfile, fontsize=7.8, align=1)
            _draw_textbox(page, fitz.Rect(472, table_header_top, 518, table_header_top + 26), "TOPLAM\nTUTAR", fontname=bold_font, fontfile=bold_fontfile, fontsize=8.3, align=2)
        else:
            _draw_textbox(page, fitz.Rect(23, table_header_top, 250, table_header_top + 16), "MALZEME", fontname=bold_font, fontfile=bold_fontfile, fontsize=10)
            _draw_textbox(page, fitz.Rect(252, table_header_top, 320, table_header_top + 16), "MİKTAR", fontname=bold_font, fontfile=bold_fontfile, fontsize=10, align=1)
            _draw_textbox(page, fitz.Rect(322, table_header_top, 388, table_header_top + 16), "BİRİM FİYAT", fontname=bold_font, fontfile=bold_fontfile, fontsize=9.4, align=1)
            _draw_textbox(page, fitz.Rect(390, table_header_top, 458, table_header_top + 16), "SEÇİLEN FİYAT", fontname=bold_font, fontfile=bold_fontfile, fontsize=9.2, align=1)
            _draw_textbox(page, fitz.Rect(460, table_header_top, 518, table_header_top + 16), "TOPLAM", fontname=bold_font, fontfile=bold_fontfile, fontsize=10, align=2)
        _draw_textbox(page, fitz.Rect(23, table_header_top + 24, 518, table_header_top + 38), f"Seçilen fiyat tipi: {price_display_label} | {vat_mode_label}", fontname=regular_font, fontfile=regular_fontfile, fontsize=8.7, color=PDF_MUTED)
        page.draw_line(fitz.Point(23, table_header_top + 44), fitz.Point(518, table_header_top + 44), color=PDF_LIGHT, width=0.8)

        row_top = table_header_top + 54
        row_height = 32 if has_discount_layout else 30
        for row_index, item in enumerate(offer_items):
            row_bottom = row_top + row_height
            if row_index % 2 == 0:
                page.draw_rect(fitz.Rect(20, row_top - 2, 520, row_bottom), color=PDF_ROW_SOFT, fill=PDF_ROW_SOFT, overlay=True)
            _draw_textbox(
                page,
                fitz.Rect(23, row_top, 182 if has_discount_layout else 212, row_bottom),
                item.product_name,
                fontname=bold_font,
                fontfile=bold_fontfile,
                fontsize=8.8,
            )
            _draw_offer_badges(
                page,
                badge_source,
                badge_clips,
                row_top,
                has_discount_layout=has_discount_layout,
                background_fill=PDF_ROW_SOFT if row_index % 2 == 0 else (1, 1, 1),
            )
            _draw_textbox(
                page,
                fitz.Rect(228 if has_discount_layout else 260, row_top + 2, 286 if has_discount_layout else 320, row_bottom),
                _format_quantity(item.quantity),
                fontname=regular_font,
                fontfile=regular_fontfile,
                fontsize=8.8,
                align=1,
            )
            if has_discount_layout:
                _draw_textbox(page, fitz.Rect(288, row_top + 2, 352, row_bottom), format_pdf_money(item.base_unit_price), fontname=regular_font, fontfile=regular_fontfile, fontsize=8.2, align=1)
                _draw_textbox(page, fitz.Rect(354, row_top + 2, 406, row_bottom), format_pdf_money(item.discount_amount), fontname=regular_font, fontfile=regular_fontfile, fontsize=8.2, align=1)
                _draw_textbox(page, fitz.Rect(408, row_top + 2, 470, row_bottom), format_pdf_money(item.unit_price), fontname=regular_font, fontfile=regular_fontfile, fontsize=8.2, align=1)
                _draw_textbox(page, fitz.Rect(472, row_top + 2, 518, row_bottom), format_pdf_money(item.total_price), fontname=bold_font, fontfile=bold_fontfile, fontsize=8.4, align=2)
            else:
                _draw_textbox(page, fitz.Rect(322, row_top + 2, 388, row_bottom), format_pdf_money(item.base_unit_price or item.unit_price), fontname=regular_font, fontfile=regular_fontfile, fontsize=8.6, align=1)
                _draw_textbox(page, fitz.Rect(390, row_top + 2, 458, row_bottom), format_pdf_money(item.unit_price), fontname=regular_font, fontfile=regular_fontfile, fontsize=8.6, align=1)
                _draw_textbox(page, fitz.Rect(460, row_top + 2, 518, row_bottom), format_pdf_money(item.total_price), fontname=bold_font, fontfile=bold_fontfile, fontsize=8.8, align=2)
            page.draw_line(fitz.Point(23, row_bottom), fitz.Point(518, row_bottom), color=PDF_LIGHT, width=0.6)
            row_top += row_height

        line_total_sum = round(sum(item.total_price for item in offer_items), 2)
        net_total, vat_total, gross_total = calculate_offer_totals(
            line_total_sum,
            vat_rate=vat_rate,
            vat_included=vat_included,
        )
        totals_top = max(500, row_top + 18)
        show_discount_total = total_discount_amount > 0.01
        if vat_included:
            _draw_textbox(page, fitz.Rect(320, totals_top, 445, totals_top + 16), "YATIRIM MALİYETİ", fontname=bold_font, fontfile=bold_fontfile, fontsize=10, align=2)
            _draw_textbox(page, fitz.Rect(448, totals_top, 518, totals_top + 16), f": {format_pdf_money(net_total)}", fontname=regular_font, fontfile=regular_fontfile, fontsize=10, align=2)
            _draw_textbox(page, fitz.Rect(320, totals_top + 20, 445, totals_top + 36), f"KDV (%{vat_rate:g})", fontname=bold_font, fontfile=bold_fontfile, fontsize=10, align=2)
            _draw_textbox(page, fitz.Rect(448, totals_top + 20, 518, totals_top + 36), f": {format_pdf_money(vat_total)}", fontname=regular_font, fontfile=regular_fontfile, fontsize=10, align=2)

            total_line_top = totals_top + 42
            if show_discount_total:
                _draw_textbox(page, fitz.Rect(300, total_line_top, 445, total_line_top + 16), "İSKONTO TUTARI", fontname=bold_font, fontfile=bold_fontfile, fontsize=10, align=2)
                _draw_textbox(page, fitz.Rect(448, total_line_top, 518, total_line_top + 16), f": {format_pdf_money(total_discount_amount)}", fontname=regular_font, fontfile=regular_fontfile, fontsize=10, align=2)
                total_line_top += 22

            _draw_textbox(page, fitz.Rect(280, total_line_top, 445, total_line_top + 18), "TOPLAM YATIRIM MALİYETİ", fontname=bold_font, fontfile=bold_fontfile, fontsize=11, color=(1, 0, 0), align=2)
            _draw_textbox(page, fitz.Rect(448, total_line_top, 518, total_line_top + 18), f": {format_pdf_money(gross_total)}", fontname=bold_font, fontfile=bold_fontfile, fontsize=11, color=(1, 0, 0), align=2)
            totals_bottom = total_line_top + 18
        else:
            total_line_top = totals_top + 16
            if show_discount_total:
                _draw_textbox(page, fitz.Rect(300, totals_top, 445, totals_top + 16), "İSKONTO TUTARI", fontname=bold_font, fontfile=bold_fontfile, fontsize=10, align=2)
                _draw_textbox(page, fitz.Rect(448, totals_top, 518, totals_top + 16), f": {format_pdf_money(total_discount_amount)}", fontname=regular_font, fontfile=regular_fontfile, fontsize=10, align=2)
                total_line_top = totals_top + 22

            _draw_textbox(page, fitz.Rect(220, total_line_top, 445, total_line_top + 18), "TOPLAM YATIRIM MALİYETİ (KDV HARİÇ)", fontname=bold_font, fontfile=bold_fontfile, fontsize=10.2, color=(1, 0, 0), align=2)
            _draw_textbox(page, fitz.Rect(448, total_line_top, 518, total_line_top + 18), f": {format_pdf_money(net_total)}", fontname=bold_font, fontfile=bold_fontfile, fontsize=11, color=(1, 0, 0), align=2)
            totals_bottom = total_line_top + 18

        if note_text and note_text.strip():
            note_top = min(totals_bottom + 26, 612)
            _draw_textbox(page, fitz.Rect(24, note_top, 518, note_top + 42), f"•  {note_text.strip()}", fontname=regular_font, fontfile=regular_fontfile, fontsize=9.2)

        _draw_textbox(page, fitz.Rect(342, 654, 510, 668), "Yetkili Adı", fontname=regular_font, fontfile=regular_fontfile, fontsize=10, align=1)
        _draw_textbox(page, fitz.Rect(342, 674, 510, 688), (contact_name or company_name or "-").upper(), fontname=bold_font, fontfile=bold_fontfile, fontsize=10.5, align=1)

        doc.save(final_output_path)
        return final_output_path
    finally:
        if close_header_source and header_source is not source:
            header_source.close()
        if close_badge_source and badge_source is not source:
            badge_source.close()
        source.close()
        doc.close()


def compare_offer_to_catalog(
    offer_items: list[OfferItem],
    price_rows: list[PriceRow],
    selected_column: str,
    tolerance: float,
    min_match_score: float,
    *,
    vat_included: bool = True,
    vat_rate: float = DEFAULT_VAT_RATE,
) -> list[MatchResult]:
    results: list[MatchResult] = []

    for item in offer_items:
        bundle_components = split_bundle_product_name(item.product_name)
        bundle_result = build_bundle_match_result(
            item,
            price_rows,
            selected_column,
            tolerance=tolerance,
            min_match_score=min_match_score,
            vat_included=vat_included,
            vat_rate=vat_rate,
        )
        if bundle_result is not None:
            results.append(bundle_result)
            continue

        best_row = None
        best_score = 0.0
        for price_row in price_rows:
            score = similarity_score(item.product_name, price_row.product_name)
            if score > best_score:
                best_score = score
                best_row = price_row

        code_match = find_unique_code_match(item.product_name, price_rows)
        if code_match is not None:
            code_row, code_score = code_match
            if code_score > best_score:
                best_row = code_row
                best_score = code_score

        if best_row is None or best_score < min_match_score:
            results.append(
                MatchResult(
                    offer_item=item,
                    matched_row=None,
                    score=best_score,
                    status="ESLESMEDI",
                    selected_column=selected_column,
                    reference_unit_price=None,
                    reference_total_price=None,
                    suggested_unit_price=None,
                    suggested_total_price=None,
                    difference=None,
                    note=(
                        "Satirda birden fazla urun olabilir; bilesen toplami otomatik dogrulanamadi. "
                        "Excel listesinde guvenilir eslesme bulunamadi."
                        if bundle_components
                        else "Excel listesinde guvenilir eslesme bulunamadi."
                    ),
                )
            )
            continue

        result = build_match_result(
            item,
            best_row,
            selected_column,
            tolerance=tolerance,
            score=best_score,
            vat_included=vat_included,
            vat_rate=vat_rate,
        )
        if bundle_components:
            result.note = (
                "Satirda birden fazla urun olabilir; bilesen toplami otomatik dogrulanamadi. "
                f"{result.note}"
            )
            if result.status == "ONAY":
                result.status = "INCELE"
                result.suggested_unit_price = result.reference_unit_price
                result.suggested_total_price = result.reference_total_price
        results.append(result)

    return results


def build_bundle_match_result(
    item: OfferItem,
    price_rows: list[PriceRow],
    selected_column: str,
    *,
    tolerance: float,
    min_match_score: float,
    vat_included: bool = True,
    vat_rate: float = DEFAULT_VAT_RATE,
) -> MatchResult | None:
    components = split_bundle_product_name(item.product_name)
    if len(components) < 2:
        return None

    component_matches: list[BundleComponentMatch] = []
    for component_name in components:
        component_match = _find_bundle_component_match(
            component_name,
            item.product_name,
            price_rows,
            selected_column,
            vat_included=vat_included,
            vat_rate=vat_rate,
            min_match_score=min_match_score,
        )
        if component_match is None:
            return None
        component_matches.append(component_match)

    reference_unit_price = round(sum(match.reference_unit_price for match in component_matches), 2)
    reference_total_price = round(reference_unit_price * item.quantity, 2)
    difference = round(item.discounted_price - reference_unit_price, 2)
    total_difference = round(item.total_price - reference_total_price, 2)
    low_confidence = any(match.score < 0.68 or match.ambiguous for match in component_matches)

    if abs(difference) <= tolerance and abs(total_difference) <= tolerance and not low_confidence:
        status = "ONAY"
        suggested_unit_price = None
        suggested_total_price = None
        note = "Satir birden fazla urunden olusuyor; bilesen fiyatlarinin toplami teklif ile uyumlu."
    elif abs(difference) <= tolerance and abs(total_difference) <= tolerance:
        status = "INCELE"
        suggested_unit_price = reference_unit_price
        suggested_total_price = reference_total_price
        note = "Satir birden fazla urunden olusuyor; bilesen toplami tutuyor ama eslesme manuel kontrol edilmeli."
    else:
        status = "DUZELT"
        suggested_unit_price = reference_unit_price
        suggested_total_price = reference_total_price
        note = "Satir birden fazla urunden olusuyor; bilesen fiyatlari toplami teklif satiriyla uyusmuyor."

    component_summary = "; ".join(
        f"{match.requested_name} -> {match.matched_row.product_name} ({format_pdf_money(match.reference_unit_price)})"
        for match in component_matches
    )
    if any(match.reference_source == "net_derived" for match in component_matches):
        note = f"{note} KDV haric bilesen fiyati brut listeden hesaplandi."
    elif any(match.reference_source == "gross_derived" for match in component_matches):
        note = f"{note} KDV dahil bilesen fiyati net listeden hesaplandi."
    note = f"{note} Bilesenler: {component_summary}"

    return MatchResult(
        offer_item=item,
        matched_row=component_matches[0].matched_row,
        score=min(match.score for match in component_matches),
        status=status,
        selected_column=selected_column,
        reference_unit_price=reference_unit_price,
        reference_total_price=reference_total_price,
        suggested_unit_price=suggested_unit_price,
        suggested_total_price=suggested_total_price,
        difference=difference,
        note=note,
    )


def build_match_result(
    item: OfferItem,
    matched_row: PriceRow,
    selected_column: str,
    *,
    tolerance: float,
    score: float | None = None,
    vat_included: bool = True,
    vat_rate: float = DEFAULT_VAT_RATE,
    manual_override: bool = False,
) -> MatchResult:
    resolved_score = score if score is not None else similarity_score(item.product_name, matched_row.product_name)
    descriptor_conflicts = find_descriptor_conflicts(item.product_name, matched_row.product_name)
    measurement_conflicts = find_measurement_conflicts(item.product_name, matched_row.product_name)
    reference_unit_price, resolved_reference_column, reference_source = resolve_price_for_vat_mode(
        matched_row,
        selected_column,
        vat_included=vat_included,
        vat_rate=vat_rate,
    )
    if reference_unit_price is None:
        note_prefix = "Elle seçilen üründe" if manual_override else "Eşleşen üründe"
        return MatchResult(
            offer_item=item,
            matched_row=matched_row,
            score=resolved_score,
            status="INCELE",
            selected_column=selected_column,
            reference_unit_price=None,
            reference_total_price=None,
            suggested_unit_price=None,
            suggested_total_price=None,
            difference=None,
            note=f"{note_prefix} '{resolved_reference_column or selected_column}' sütununda fiyat yok.",
        )

    reference_total_price = round(reference_unit_price * item.quantity, 2)
    difference = round(item.discounted_price - reference_unit_price, 2)
    total_difference = round(item.total_price - reference_total_price, 2)

    if abs(item.discounted_price) <= tolerance and reference_unit_price > tolerance:
        status = "INCELE"
        note = "Teklifte bedelsiz geçilmiş; otomatik onay verilmedi."
    elif not manual_override and resolved_score < 0.72:
        status = "INCELE"
        note = "Eşleşme skoru düşük olduğu için manuel kontrol önerilir."
    elif abs(difference) <= tolerance and abs(total_difference) <= tolerance:
        status = "ONAY"
        note = "Teklif fiyatı liste ile uyumlu."
    else:
        status = "DUZELT"
        note = "Teklif fiyatı seçilen liste fiyatından farklı."

    if is_alias_match(item.product_name, matched_row.product_name):
        note = f"Alias eşleşmesi kullanıldı. {note}"

    if descriptor_conflicts:
        conflict_note = "Urun adinda celisen ifade bulundu: " + ", ".join(descriptor_conflicts) + "."
        if status == "ONAY":
            status = "INCELE"
            note = f"{conflict_note} Otomatik onay verilmedi."
        else:
            note = f"{conflict_note} {note}"

    if measurement_conflicts:
        conflict_note = "Olcu uyusmazligi bulundu: " + ", ".join(measurement_conflicts) + "."
        if status == "ONAY":
            status = "INCELE"
            note = f"{conflict_note} Otomatik onay verilmedi."
        else:
            note = f"{conflict_note} {note}"

    if manual_override:
        note = f"Ürün elle seçildi. {note}"

    if reference_source == "net_derived":
        note = f"{note} KDV haric referans fiyat brut listeden hesaplandi."
    elif reference_source == "gross_derived":
        note = f"{note} KDV dahil referans fiyat net listeden hesaplandi."

    if matched_row.note:
        note = f"{note} Not: {matched_row.note}"

    return MatchResult(
        offer_item=item,
        matched_row=matched_row,
        score=resolved_score,
        status=status,
        selected_column=selected_column,
        reference_unit_price=reference_unit_price,
        reference_total_price=reference_total_price,
        suggested_unit_price=reference_unit_price if status in {"DUZELT", "INCELE"} else None,
        suggested_total_price=reference_total_price if status in {"DUZELT", "INCELE"} else None,
        difference=difference,
        note=note,
    )


def _collect_offer_row_rects(doc: fitz.Document) -> list[tuple[int, fitz.Rect, str]]:
    row_rects: list[tuple[int, fitz.Rect, str]] = []
    for page_index, page in enumerate(doc):
        blocks = page.get_text("blocks")
        for block_index, block in enumerate(blocks):
            x0, y0, x1, y1, text, *_ = block
            normalized = normalize_text(str(text or ""))
            if "ADET" not in normalized:
                continue
            if normalized.count("TL") < 2:
                continue
            row_text_parts = [str(text or "")]
            scan_index = block_index - 1
            while scan_index >= 0:
                prev_x0, prev_y0, prev_x1, prev_y1, prev_text, *_ = blocks[scan_index]
                prev_text = str(prev_text or "")
                prev_normalized = normalize_text(prev_text)
                if not prev_normalized:
                    scan_index -= 1
                    continue
                if prev_normalized in HEADER_SKIP_LINES or "MALZEME" in prev_normalized:
                    break
                if "KURUMSAL SATIS SORUMLUSU" in prev_normalized or "YATIRIM MALIYETI" in prev_normalized:
                    break
                if "ADET" in prev_normalized:
                    break

                vertically_related = prev_y1 >= y0 - 42 and prev_y0 <= y1 + 20
                horizontally_related = prev_x0 <= x0 + 24
                if not (vertically_related and horizontally_related):
                    if prev_y1 < y0 - 42:
                        break
                    scan_index -= 1
                    continue

                row_text_parts.insert(0, prev_text)
                scan_index -= 1
            row_rects.append((page_index, fitz.Rect(x0, y0, x1, y1), " ".join(row_text_parts)))
    row_rects.sort(key=lambda item: (item[0], item[1].y0, item[1].x0))
    return row_rects


def _rect_overlaps_row(candidate: fitz.Rect, row_rect: fitz.Rect, tolerance: float = 3.0) -> bool:
    return candidate.y0 >= row_rect.y0 - tolerance and candidate.y1 <= row_rect.y1 + tolerance


def _pick_price_rects_for_row(
    page: fitz.Page,
    row_rect: fitz.Rect,
    discounted_text: str,
    total_text: str,
) -> tuple[fitz.Rect | None, fitz.Rect | None]:
    discounted_hits = [rect for rect in page.search_for(discounted_text) if _rect_overlaps_row(rect, row_rect)]
    total_hits = [rect for rect in page.search_for(total_text) if _rect_overlaps_row(rect, row_rect)]

    if discounted_text == total_text:
        shared_hits = sorted(discounted_hits, key=lambda rect: rect.x0)
        if len(shared_hits) >= 2:
            return shared_hits[0], shared_hits[-1]
        return None, None

    discounted_rect = min(discounted_hits, key=lambda rect: rect.x0) if discounted_hits else None
    total_rect = max(total_hits, key=lambda rect: rect.x0) if total_hits else None
    return discounted_rect, total_rect


def _row_text_matches_result(row_text: str, result: MatchResult) -> bool:
    candidate_names = [result.offer_item.product_name]
    if result.matched_row and result.matched_row.product_name not in candidate_names:
        candidate_names.append(result.matched_row.product_name)

    row_text_normalized = normalize_text(row_text)
    expected_price_texts = {
        format_pdf_money(result.offer_item.discounted_price),
        format_pdf_money(result.offer_item.total_price),
    }
    for price_text in expected_price_texts:
        if price_text and normalize_text(price_text) not in row_text_normalized:
            return False

    row_codes = extract_codes(row_text)
    row_tokens = tokenize(row_text)
    for candidate_name in candidate_names:
        candidate_codes = extract_codes(candidate_name)
        if candidate_codes and not (row_codes & candidate_codes):
            continue

        candidate_tokens = tokenize(candidate_name)
        if candidate_tokens:
            shared_tokens = row_tokens & candidate_tokens
            minimum_shared = max(1, min(2, len(candidate_tokens)))
            if len(shared_tokens) < minimum_shared:
                continue
        return True

    return False


def _replace_text_in_rect(
    page: fitz.Page,
    rect: fitz.Rect,
    text: str,
    *,
    color: tuple[float, float, float] = (0, 0, 0),
    align: int = 1,
    fontname: str = "helv",
    fontfile: str | None = None,
    fontsize: float | None = None,
    baseline_y: float | None = None,
    x_hint: float | None = None,
) -> None:
    padded = fitz.Rect(rect.x0 - 2, rect.y0 - 1, rect.x1 + 10, rect.y1 + 1)
    page.add_redact_annot(padded, fill=(1, 1, 1))
    page.apply_redactions(images=fitz.PDF_REDACT_IMAGE_NONE)
    resolved_fontsize = fontsize if fontsize is not None else max(7.5, min(10, padded.height * 0.78))
    if fontfile:
        try:
            text_width = fitz.Font(fontfile=fontfile).text_length(text, fontsize=resolved_fontsize)
        except Exception:
            text_width = fitz.get_text_length(text, fontname=fontname, fontsize=resolved_fontsize)
    else:
        text_width = fitz.get_text_length(text, fontname=fontname, fontsize=resolved_fontsize)
    if align == 2:
        x_position = max(padded.x0, padded.x1 - text_width)
    elif align == 1:
        x_position = max(padded.x0, padded.x0 + ((padded.width - text_width) / 2))
    else:
        x_position = x_hint if x_hint is not None else padded.x0
    y_position = baseline_y if baseline_y is not None else padded.y1 - 1.5
    page.insert_text(
        fitz.Point(x_position, y_position),
        text,
        fontname=fontname,
        fontfile=fontfile,
        fontsize=resolved_fontsize,
        color=color,
    )


def _find_summary_value_rect(
    page: fitz.Page,
    value_text: str,
    *,
    highest: bool = False,
) -> fitz.Rect | None:
    rects = page.search_for(value_text)
    if not rects:
        return None
    if highest:
        return max(rects, key=lambda rect: rect.y0)
    return min(rects, key=lambda rect: rect.y0)


def _find_line_rect_by_text_prefix(
    page: fitz.Page,
    prefix_text: str,
    *,
    highest: bool = False,
) -> tuple[fitz.Rect | None, str]:
    rect, line_text, _style = _find_line_render_spec_by_text_prefix(page, prefix_text, highest=highest)
    return rect, line_text


def _resolve_replacement_font(
    page: fitz.Page,
    source_font_name: str,
) -> tuple[str, str | None]:
    regular_font, bold_font = _register_offer_fonts(page)
    normalized_font = normalize_text(source_font_name)
    if any(token in normalized_font for token in ("BOLD", "BLACK", "SEMIBOLD", "DEMI", "MEDIUM")):
        return bold_font
    return regular_font


def _find_line_render_spec_by_text_prefix(
    page: fitz.Page,
    prefix_text: str,
    *,
    highest: bool = False,
) -> tuple[fitz.Rect | None, str, dict[str, float | str | None]]:
    normalized_prefix = normalize_text(prefix_text)
    if not normalized_prefix:
        return None, "", {}

    matching_lines: list[tuple[fitz.Rect, str, dict[str, float | str | None]]] = []
    text_dict = page.get_text("dict")
    for block in text_dict.get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            spans = line.get("spans", [])
            line_text = "".join(str(span.get("text", "")) for span in spans)
            if normalized_prefix not in normalize_text(line_text):
                continue
            line_rect = fitz.Rect(line["bbox"])
            meaningful_spans = [span for span in spans if str(span.get("text", "")).strip()]
            dominant_span = max(
                meaningful_spans,
                key=lambda span: float(span.get("size", 0) or 0),
                default=None,
            )
            span_sizes = [
                float(span.get("size", 0) or 0)
                for span in meaningful_spans
                if float(span.get("size", 0) or 0) > 0
            ]
            inferred_fontsize = max(7.5, min(11, (max(span_sizes) if span_sizes else line_rect.height * 0.78)))
            source_font_name = str(dominant_span.get("font", "")) if dominant_span else ""
            fontname, fontfile = _resolve_replacement_font(page, source_font_name)
            baseline_y = max(
                (
                    float(span.get("origin", (0, line_rect.y1))[1])
                    for span in meaningful_spans
                    if len(span.get("origin", ())) >= 2
                ),
                default=line_rect.y1 - 1.5,
            )
            x_hint = min(
                (
                    float(span.get("bbox", (line_rect.x0,))[0])
                    for span in meaningful_spans
                    if len(span.get("bbox", ())) >= 1
                ),
                default=line_rect.x0,
            )
            matching_lines.append(
                (
                    line_rect,
                    line_text,
                    {
                        "fontname": fontname,
                        "fontfile": fontfile,
                        "fontsize": inferred_fontsize,
                        "baseline_y": baseline_y,
                        "x_hint": x_hint,
                    },
                )
            )

    if not matching_lines:
        return None, "", {}
    if highest:
        return max(matching_lines, key=lambda item: item[0].y0)
    return min(matching_lines, key=lambda item: item[0].y0)


def _find_block_rect_by_text_prefix(
    page: fitz.Page,
    prefix_text: str,
    *,
    highest: bool = False,
) -> fitz.Rect | None:
    rect, _line_text = _find_line_rect_by_text_prefix(page, prefix_text, highest=highest)
    return rect


def apply_approved_corrections_to_pdf(
    offer_path: Path,
    results: list[MatchResult],
    approved_indexes: list[int],
    output_path: Path | None = None,
) -> Path:
    if not offer_path.exists():
        raise FileNotFoundError(f"Teklif PDF bulunamadi: {offer_path}")
    if offer_path.suffix.lower() != ".pdf":
        raise ValueError(f"Desteklenmeyen teklif formati: {offer_path.name}")
    if not approved_indexes:
        raise ValueError("Uygulanacak onaylı satır seçilmedi.")

    approved_indexes = sorted(set(approved_indexes))
    invalid_indexes = [index for index in approved_indexes if index < 0 or index >= len(results)]
    if invalid_indexes:
        raise ValueError(f"Uygulanacak satir indeksleri gecersiz: {invalid_indexes}")

    final_output_path = _ensure_safe_output_path(
        output_path or build_corrected_pdf_path(offer_path),
        expected_suffix=".pdf",
        label="Duzeltilmis PDF cikisi",
        source_paths=(offer_path,),
    )

    doc = fitz.open(offer_path)
    try:
        row_rects = _collect_offer_row_rects(doc)
        if len(row_rects) < len(results):
            raise ValueError("PDF içindeki teklif satırları bulunamadı; düzeltme uygulanamadı.")

        operations: list[tuple[int, fitz.Rect, str, tuple[float, float, float], int, dict[str, float | str | None] | None]] = []

        for result_index in approved_indexes:
            result = results[result_index]
            if result.status != "DUZELT":
                raise ValueError(
                    f"Sadece DÜZELT durumundaki satırlar uygulanabilir: {result.offer_item.product_name}"
                )
            if result.suggested_unit_price is None or result.suggested_total_price is None:
                raise ValueError(f"Seçilen satır için uygulanabilir fiyat önerisi yok: {result.offer_item.product_name}")

            page_index, row_rect, row_text = row_rects[result_index]
            if not _row_text_matches_result(row_text, result):
                raise ValueError(
                    f"PDF satırı beklenen ürünle doğrulanamadı, işlem durduruldu: {result.offer_item.product_name}"
                )

            page = doc[page_index]
            discounted_text = format_pdf_money(result.offer_item.discounted_price)
            total_text = format_pdf_money(result.offer_item.total_price)
            new_discounted_text = format_pdf_money(result.suggested_unit_price)
            new_total_text = format_pdf_money(result.suggested_total_price)

            discounted_rect, total_rect = _pick_price_rects_for_row(page, row_rect, discounted_text, total_text)
            if discounted_rect is None or total_rect is None:
                raise ValueError(
                    f"PDF satırında hem birim hem toplam fiyat alanı bulunamadı: {result.offer_item.product_name}"
                )

            operations.append((page_index, discounted_rect, new_discounted_text, (0, 0, 0), 1, None))
            operations.append((page_index, total_rect, new_total_text, (0, 0, 0), 1, None))

        current_line_total = round(sum(result.offer_item.total_price for result in results), 2)
        corrected_line_total = round(
            sum(
                (
                    results[index].suggested_total_price
                    if index in approved_indexes and results[index].suggested_total_price is not None
                    else results[index].offer_item.total_price
                )
                for index in range(len(results))
            ),
            2,
        )
        offer_text = extract_offer_text(offer_path)
        offer_financial_summary = parse_offer_financial_summary(offer_text)
        vat_included = detect_offer_vat_included(offer_text)
        if vat_included is None:
            vat_included = True
        selected_column = results[approved_indexes[0]].selected_column
        price_display_label = resolve_price_display_label(selected_column)
        payment_info_text = resolve_payment_info(selected_column)
        vat_mode_label = "KDV Dahil" if vat_included else "KDV Hariç"
        current_net_total, current_vat_total, current_gross_total = calculate_offer_totals(
            current_line_total,
            vat_rate=offer_financial_summary.vat_rate,
            vat_included=vat_included,
        )
        corrected_net_total, corrected_vat_total, corrected_gross_total = calculate_offer_totals(
            corrected_line_total,
            vat_rate=offer_financial_summary.vat_rate,
            vat_included=vat_included,
        )
        displayed_current_net = (
            offer_financial_summary.net_total
            if offer_financial_summary.net_total is not None
            else current_net_total
        )
        displayed_current_vat = (
            offer_financial_summary.vat_total
            if offer_financial_summary.vat_total is not None
            else current_vat_total
        )
        displayed_current_gross = (
            offer_financial_summary.gross_total
            if offer_financial_summary.gross_total is not None
            else current_gross_total
        )

        last_page_index = len(doc) - 1
        last_page = doc[last_page_index]
        summary_specs = [
            (format_pdf_money(displayed_current_net), format_pdf_money(corrected_net_total), (0, 0, 0), 2, False),
            (format_pdf_money(displayed_current_vat), format_pdf_money(corrected_vat_total), (0, 0, 0), 2, False),
            (format_pdf_money(displayed_current_gross), format_pdf_money(corrected_gross_total), (1, 0, 0), 2, True),
        ]
        for current_text, new_text, color, align, highest in summary_specs:
            rect = _find_summary_value_rect(last_page, current_text, highest=highest)
            if rect is not None:
                operations.append((last_page_index, rect, new_text, color, align, None))

        for page_index, page in enumerate(doc):
            payment_rect, payment_line_text, payment_style = _find_line_render_spec_by_text_prefix(
                page,
                "Odeme Bilgisi",
            )
            if payment_rect is not None:
                payment_prefix = "•  "
                for marker in ("Ödeme Bilgisi", "Odeme Bilgisi"):
                    marker_index = payment_line_text.find(marker)
                    if marker_index >= 0:
                        payment_prefix = payment_line_text[:marker_index]
                        break
                operations.append(
                    (
                        page_index,
                        payment_rect,
                        f"{payment_prefix}Ödeme Bilgisi : {payment_info_text}",
                        (0, 0, 0),
                        0,
                        payment_style,
                    )
                )

            price_type_rect, _price_type_text, price_type_style = _find_line_render_spec_by_text_prefix(
                page,
                "Secilen fiyat tipi",
            )
            if price_type_rect is not None:
                operations.append(
                    (
                        page_index,
                        price_type_rect,
                        f"Seçilen fiyat tipi: {price_display_label} | {vat_mode_label}",
                        PDF_MUTED,
                        0,
                        price_type_style,
                    )
                )

        for page_index, rect, text, color, align, text_style in operations:
            _replace_text_in_rect(doc[page_index], rect, text, color=color, align=align, **(text_style or {}))

        final_output_path.parent.mkdir(parents=True, exist_ok=True)
        doc.save(final_output_path)
        return final_output_path
    finally:
        doc.close()


def autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        if not values:
            continue
        max_length = max(len(value) for value in values)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max_length + 2, 45)


def write_report(
    output_path: Path,
    results: list[MatchResult],
    price_list_path: Path,
    offer_path: Path,
    selected_column: str,
    *,
    financial_review: FinancialReview,
) -> None:
    final_output_path = _ensure_safe_output_path(
        output_path,
        expected_suffix=".xlsx",
        label="Excel raporu",
        source_paths=(price_list_path,),
    )

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = SUMMARY_SHEET
    result_sheet = workbook.create_sheet(RESULT_SHEET)
    corrected_sheet = workbook.create_sheet(CORRECTED_SHEET)
    financial_sheet = workbook.create_sheet(FINANCIAL_SHEET)

    summary_sheet["A1"] = "Fiyat Listesi"
    summary_sheet["B1"] = price_list_path.name
    summary_sheet["A2"] = "Teklif Dosyasi"
    summary_sheet["B2"] = offer_path.name
    summary_sheet["A3"] = "Referans Sutun"
    summary_sheet["B3"] = selected_column
    summary_sheet["A4"] = "Rapor Tarihi"
    summary_sheet["B4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    counts = {
        "ONAY": sum(result.status == "ONAY" for result in results),
        "DUZELT": sum(result.status == "DUZELT" for result in results),
        "INCELE": sum(result.status == "INCELE" for result in results),
        "ESLESMEDI": sum(result.status == "ESLESMEDI" for result in results),
    }

    summary_sheet["A6"] = "Durum"
    summary_sheet["B6"] = "Adet"
    current_row = 7
    for status, count in counts.items():
        summary_sheet[f"A{current_row}"] = status
        summary_sheet[f"B{current_row}"] = count
        current_row += 1

    summary_sheet["A12"] = "Toplam Teklif"
    summary_sheet["B12"] = sum(result.offer_item.total_price for result in results)
    summary_sheet["A13"] = "Toplam Referans"
    summary_sheet["B13"] = sum((result.reference_total_price or 0) for result in results)
    summary_sheet["A15"] = "Finansal Kontrol"
    summary_sheet["B15"] = financial_review.overall_status
    summary_sheet["A16"] = "KDV Orani"
    summary_sheet["B16"] = financial_review.vat_rate / 100
    summary_sheet["A17"] = "KDV Orani Kaynagi"
    summary_sheet["B17"] = financial_review.vat_rate_source

    headers = [
        "Durum",
        "Teklif Urunu",
        "Eslesen Liste Urunu",
        "Eslesme Skoru",
        "Adet",
        "Teklif Birim Fiyat",
        "Teklif Indirimli Fiyat",
        "Teklif Toplam",
        "Referans Sutun",
        "Referans Birim Fiyat",
        "Referans Toplam",
        "Onerilen Birim Fiyat",
        "Onerilen Toplam",
        "Birim Fark",
        "Not",
        "Liste Satiri",
    ]
    result_sheet.append(headers)

    fills = {
        "ONAY": PatternFill(fill_type="solid", fgColor="C6EFCE"),
        "DUZELT": PatternFill(fill_type="solid", fgColor="FFC7CE"),
        "INCELE": PatternFill(fill_type="solid", fgColor="FFEB9C"),
        "ESLESMEDI": PatternFill(fill_type="solid", fgColor="D9E1F2"),
    }

    for result in results:
        result_sheet.append(
            [
                result.status,
                result.offer_item.product_name,
                result.matched_row.product_name if result.matched_row else "",
                round(result.score, 4),
                result.offer_item.quantity,
                result.offer_item.unit_price,
                result.offer_item.discounted_price,
                result.offer_item.total_price,
                result.selected_column,
                result.reference_unit_price,
                result.reference_total_price,
                result.suggested_unit_price,
                result.suggested_total_price,
                result.difference,
                result.note,
                result.matched_row.row_number if result.matched_row else "",
            ]
        )
        status_cell = result_sheet[f"A{result_sheet.max_row}"]
        status_cell.fill = fills[result.status]

    corrected_headers = [
        "Durum",
        "Teklif Urunu",
        "Eslesen Liste Urunu",
        "Adet",
        "Mevcut Indirimli Fiyat",
        "Onerilen Birim Fiyat",
        "Onerilen Toplam",
        "Not",
    ]
    corrected_sheet.append(corrected_headers)

    for result in results:
        corrected_unit = result.suggested_unit_price
        corrected_total = result.suggested_total_price
        if corrected_unit is None and result.status == "ONAY":
            corrected_unit = result.offer_item.discounted_price
            corrected_total = result.offer_item.total_price

        corrected_sheet.append(
            [
                result.status,
                result.offer_item.product_name,
                result.matched_row.product_name if result.matched_row else "",
                result.offer_item.quantity,
                result.offer_item.discounted_price,
                corrected_unit,
                corrected_total,
                result.note,
            ]
        )
        corrected_sheet[f"A{corrected_sheet.max_row}"].fill = fills[result.status]

    financial_sheet["A1"] = "Kontrol"
    financial_sheet["B1"] = "Durum"
    financial_sheet["C1"] = "Teklif Ozeti"
    financial_sheet["D1"] = "Hesaplanan"
    financial_sheet["E1"] = "Fark"
    financial_sheet["F1"] = "Not"
    financial_sheet["A2"] = "Tespit Edilen KDV Orani"
    financial_sheet["B2"] = financial_review.vat_rate / 100
    financial_sheet["C2"] = financial_review.vat_rate_source
    financial_sheet["A3"] = (
        "Kalem Toplamindan Beklenen Teklif Toplami (KDV Dahil)"
        if financial_review.vat_included
        else "Kalem Toplamindan Beklenen Teklif Toplami (KDV Haric)"
    )
    financial_sheet["B3"] = financial_review.expected_summary_total

    for check in financial_review.checks:
        financial_sheet.append(
            [
                check.label,
                check.status,
                check.offer_value,
                check.calculated_value,
                check.difference,
                check.note,
            ]
        )
        financial_sheet[f"B{financial_sheet.max_row}"].fill = fills[check.status]

    bold_font = Font(bold=True)
    for cell in summary_sheet[1] + summary_sheet[6]:
        cell.font = bold_font
    for cell in result_sheet[1]:
        cell.font = bold_font
    for cell in corrected_sheet[1]:
        cell.font = bold_font
    for cell in financial_sheet[1]:
        cell.font = bold_font

    for row in range(12, 14):
        summary_sheet[f"B{row}"].number_format = '#,##0.00 "TL"'
    summary_sheet["B16"].number_format = '0%'

    money_columns = ["F", "G", "H", "J", "K", "L", "M", "N"]
    for column_name in money_columns:
        for row in range(2, result_sheet.max_row + 1):
            result_sheet[f"{column_name}{row}"].number_format = '#,##0.00 "TL"'

    autosize_columns(summary_sheet)
    autosize_columns(result_sheet)
    for column_name in ["E", "F", "G"]:
        for row in range(2, corrected_sheet.max_row + 1):
            corrected_sheet[f"{column_name}{row}"].number_format = '#,##0.00 "TL"'
    autosize_columns(corrected_sheet)
    financial_sheet["B2"].number_format = '0%'
    financial_sheet["B3"].number_format = '#,##0.00 "TL"'
    for column_name in ["C", "D", "E"]:
        for row in range(4, financial_sheet.max_row + 1):
            financial_sheet[f"{column_name}{row}"].number_format = '#,##0.00 "TL"'
    autosize_columns(financial_sheet)
    final_output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(final_output_path)
    except Exception as exc:
        logger.exception("Excel raporu kaydedilemedi: %s", final_output_path)
        raise ValueError(f"Excel raporu kaydedilemedi: {final_output_path}") from exc


def print_console_summary(
    results: list[MatchResult],
    selected_column: str,
    output_path: Path,
    financial_review: FinancialReview,
) -> None:
    print(f"Referans sutun: {selected_column}")
    print(f"Rapor: {output_path.name}")
    print(
        f"Finansal kontrol: {financial_review.overall_status} | "
        f"KDV oran\u0131: %{financial_review.vat_rate:g} ({financial_review.vat_rate_source})"
    )
    for check in financial_review.checks:
        print(
            f"  - {check.label}: {check.status} | "
            f"Teklif: {format_money(check.offer_value)} | "
            f"Hesaplanan: {format_money(check.calculated_value)} | "
            f"Fark: {format_money(check.difference)}"
        )
    print("")
    for result in results:
        matched_name = result.matched_row.product_name if result.matched_row else "-"
        print(
            f"[{result.status}] {result.offer_item.product_name} -> {matched_name} "
            f"(skor={result.score:.2f})"
        )
        print(
            f"  Teklif: {format_money(result.offer_item.discounted_price)} | "
            f"Referans: {format_money(result.reference_unit_price)} | "
            f"Fark: {format_money(result.difference)}"
        )
        if result.suggested_unit_price is not None:
            print(f"  Oneri: {format_money(result.suggested_unit_price)}")
        print(f"  Not: {result.note}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Excel fiyat listesi ile PDF teklifini karsilastirip rapor uretir."
    )
    parser.add_argument("--price-list", type=Path, help="Fiyat listesi Excel dosya yolu.")
    parser.add_argument("--offer", type=Path, help="Kontrol edilecek teklif PDF dosya yolu.")
    parser.add_argument("--sheet", help="Excel icindeki sayfa adi.")
    parser.add_argument(
        "--price-column",
        help="Karsilastirmada kullanilacak fiyat sutunu. Verilmezse teklif metninden tahmin edilir.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Uretilecek rapor dosyasi. Varsayilan: teklif_kontrol_raporu.xlsx",
    )
    parser.add_argument(
        "--min-match-score",
        type=float,
        default=0.55,
        help="Otomatik eslesme alt siniri. Varsayilan: 0.55",
    )
    parser.add_argument(
        "--tolerance",
        type=float,
        default=1.0,
        help="Fiyat farki toleransi. Varsayilan: 1 TL",
    )
    return parser


def main() -> None:
    configure_logging()
    parser = build_parser()
    args = parser.parse_args()

    try:
        base_dir = Path.cwd()
        price_list_path = args.price_list or find_single_file(base_dir, ".xlsx", exclude_prefixes=("teklif_kontrol_",))
        offer_path = args.offer or find_single_file(base_dir, ".pdf", exclude_suffixes=("_duzeltilmis.pdf",))
        output_path = args.output or build_report_output_path(base_dir, offer_path)
        results, selected_column, _, _, financial_review = run_comparison(
            price_list_path=price_list_path,
            offer_path=offer_path,
            sheet_name=args.sheet,
            price_column=args.price_column,
            output_path=output_path,
            min_match_score=args.min_match_score,
            tolerance=args.tolerance,
        )
        print_console_summary(results, selected_column, output_path, financial_review)
    except Exception as exc:
        logger.exception("Teklif kontrol islemi basarisiz oldu.")
        raise SystemExit(f"Hata: {exc}") from exc


if __name__ == "__main__":
    main()
