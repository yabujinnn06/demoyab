from __future__ import annotations
import csv
import io
import json
import os
import re
import socket
import sys
import threading
import time
from urllib.parse import quote
import uuid
import webbrowser
from collections import Counter
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import uvicorn
import fitz
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook

from .portal_auth import enforce_offer_access, get_offer_portal_user, require_offer_admin
from .teklif_kontrol import (
    AUTO_OFFER_NUMBER_PATTERN,
    DEFAULT_VAT_RATE,
    BundleComponentMatch,
    DISCOUNT_TYPE_AMOUNT,
    DISCOUNT_TYPE_NONE,
    DISCOUNT_TYPE_PERCENT,
    FinancialCheck,
    FinancialReview,
    MatchResult,
    OfferSelection,
    OfferItem,
    PriceRow,
    apply_approved_corrections_to_pdf,
    build_match_result,
    build_corrected_pdf_path,
    build_report_output_path,
    configure_logging,
    CORRECTED_PDFS_DIRNAME,
    create_offer_from_catalog,
    derive_standard_price_header,
    default_offer_number,
    default_valid_until,
    find_single_file,
    format_money,
    GENERATED_OFFERS_DIRNAME,
    get_price_columns,
    load_price_rows,
    OUTPUT_ROOT_DIRNAME,
    offer_number_exists,
    parse_money,
    normalize_text,
    REPORTS_DIRNAME,
    resolve_price_for_vat_mode,
    resolve_price_column,
    run_comparison,
    sanitize_filename_part,
    similarity_score,
)


def runtime_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


def resource_base_dir() -> Path:
    return Path(getattr(sys, "_MEIPASS", runtime_base_dir()))


def resolve_webui_dir() -> Path:
    candidates = [
        resource_base_dir() / "webui",
        runtime_base_dir() / "webui",
        Path(__file__).resolve().parent / "webui",
    ]
    for candidate in candidates:
        if (candidate / "templates").is_dir() and (candidate / "static").is_dir():
            return candidate
    return candidates[0]


def resolve_assets_dir() -> Path:
    candidates = [
        resource_base_dir() / "assets",
        runtime_base_dir() / "assets",
        Path(__file__).resolve().parent / "assets",
    ]
    for candidate in candidates:
        if candidate.is_dir():
            return candidate
    return candidates[0]


BASE_DIR = Path(os.getenv("CALL_PORTAL_OFFER_DATA_DIR", "")).expanduser().resolve() if os.getenv("CALL_PORTAL_OFFER_DATA_DIR") else runtime_base_dir()
RESOURCE_DIR = resource_base_dir()
WEBUI_DIR = resolve_webui_dir()
TEMPLATES_DIR = WEBUI_DIR / "templates"
STATIC_DIR = WEBUI_DIR / "static"
ASSETS_DIR = resolve_assets_dir()
DATA_DIR = BASE_DIR / "veri"
PRICE_LISTS_DIR = DATA_DIR / "fiyat_listeleri"
ADMIN_SETTINGS_PATH = DATA_DIR / "admin_ayarlar.json"
ACTIVITY_LOG_PATH = DATA_DIR / "offer_activity_log.json"
PENDING_OFFERS_PATH = DATA_DIR / "pending_offer_approvals.json"
BATCH_JOBS_DIR = DATA_DIR / "batch_jobs"
OFFERS_DIR = BASE_DIR / "teklifler"
TEMPLATES_PDF_DIR = BASE_DIR / "sablonlar"
DEFAULT_ADMIN_PIN = "2834"

PRICE_MODE_AUTO = "auto"
PRICE_MODE_TO_HEADER = {
    "kurumsal_nakit": "2026 KURUMSAL NAKIT",
    "kurumsal_4": "2026 KURUMSAL 4 TAKSIT",
    "kurumsal_6": "2026 KURUMSAL 6 TAKSIT",
    "perakende_nakit": "2026 PERAKENDE NAKIT",
    "perakende_4": "2026 PERAKENDE 4 TAKSIT",
    "perakende_6": "2026 PERAKENDE 6 TAKSIT",
}
PRICE_MODE_LABELS = {
    PRICE_MODE_AUTO: "Tekliften otomatik seç",
    "kurumsal_nakit": "Kurumsal / Nakit",
    "kurumsal_4": "Kurumsal / 4 Taksit",
    "kurumsal_6": "Kurumsal / 6 Taksit",
    "perakende_nakit": "Perakende / Nakit",
    "perakende_4": "Perakende / 4 Taksit",
    "perakende_6": "Perakende / 6 Taksit",
}
PRICE_MODE_PAYMENT_INFO = {
    "kurumsal_nakit": "Nakit veya Banka Havalesi",
    "kurumsal_4": "4 Taksit",
    "kurumsal_6": "6 Taksit",
    "perakende_nakit": "Nakit veya Banka Havalesi",
    "perakende_4": "4 Taksit",
    "perakende_6": "6 Taksit",
}
PRICE_MODE_GROUPS = (
    ("Otomatik", (PRICE_MODE_AUTO,)),
    ("Kurumsal", ("kurumsal_nakit", "kurumsal_4", "kurumsal_6")),
    ("Perakende", ("perakende_nakit", "perakende_4", "perakende_6")),
)
PRICE_PDF_COLUMN_BOUNDS = ((40.0, 198.0), (205.0, 363.0), (370.0, 528.0))
PRICE_PDF_HEADERS = (
    "URUN",
    "2026 KURUMSAL NAKIT",
    "2026 KURUMSAL 4 TAKSIT",
    "2026 KURUMSAL 6 TAKSIT",
    "2026 PERAKENDE NAKIT",
    "2026 PERAKENDE 4 TAKSIT",
    "2026 PERAKENDE 6 TAKSIT",
    "NOT",
)
VAT_MODE_INCLUDED = "dahil"
VAT_MODE_EXCLUDED = "haric"
VAT_MODE_LABELS = {
    VAT_MODE_INCLUDED: "KDV Dahil",
    VAT_MODE_EXCLUDED: "KDV Hariç",
}

CREATE_FORM_MIN_ROWS = 8

STATUS_CLASS = {
    "ONAY": "ok",
    "DUZELT": "danger",
    "INCELE": "warn",
    "ESLESMEDI": "info",
}
ACTIVITY_ACTION_LABELS = {
    "compare": "Kontrol raporu",
    "batch_compare": "Toplu kontrol",
    "correct": "PDF düzeltme",
    "create": "Onay bekleyen teklif",
    "approve": "Teklif onayı",
    "reject": "Teklif reddi",
}
ACTIVITY_LOG_LIMIT = 500
BATCH_COMPARE_LIMIT = 20


@dataclass(slots=True)
class ComparisonSession:
    token: str
    price_list_path: Path
    offer_path: Path
    output_path: Path
    selected_column: str
    price_mode: str
    results: list[MatchResult]
    financial_review: FinancialReview
    corrected_pdf_path: Path | None = None
    applied_indexes: set[int] = field(default_factory=set)


@dataclass(slots=True)
class BatchComparisonItem:
    offer_path: Path
    status: str
    session_token: str = ""
    output_path: Path | None = None
    selected_column: str = ""
    metrics: dict[str, int] = field(default_factory=dict)
    financial_status: str = "-"
    error: str = ""
    problem_summary: str = ""
    next_step: str = ""


@dataclass(slots=True)
class BatchComparisonJob:
    token: str
    price_list_path: Path
    price_mode: str
    created_at: datetime
    summary_path: Path
    items: list[BatchComparisonItem] = field(default_factory=list)


app = FastAPI(title="Rainwater Teklif Kontrol")
app.middleware("http")(enforce_offer_access)
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="offer_static")
app.mount("/assets", StaticFiles(directory=str(ASSETS_DIR)), name="offer_assets")
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
_template_response = templates.TemplateResponse


def render_template(name: str, context: dict, *, status_code: int = 200) -> HTMLResponse:
    return _template_response(context["request"], name, context, status_code=status_code)


templates.TemplateResponse = render_template  # type: ignore[method-assign]


SESSIONS: dict[str, ComparisonSession] = {}
BATCHES: dict[str, BatchComparisonJob] = {}
STORAGE_HINT = "Tum ciktilar ciktilar klasorunde duzenli tutulur."
ADMIN_STORAGE_HINT = "Fiyat listeleri veri/fiyat_listeleri klasorunde tutulur."


def ensure_runtime_folders() -> None:
    ensure_admin_storage()
    folders = (
        OFFERS_DIR,
        TEMPLATES_PDF_DIR,
        BASE_DIR / OUTPUT_ROOT_DIRNAME / REPORTS_DIRNAME,
        BASE_DIR / OUTPUT_ROOT_DIRNAME / CORRECTED_PDFS_DIRNAME,
        BASE_DIR / OUTPUT_ROOT_DIRNAME / GENERATED_OFFERS_DIRNAME,
    )
    for folder in folders:
        folder.mkdir(parents=True, exist_ok=True)


def ensure_admin_storage() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    PRICE_LISTS_DIR.mkdir(parents=True, exist_ok=True)
    BATCH_JOBS_DIR.mkdir(parents=True, exist_ok=True)


def normalize_compare_mode(value: str | None) -> str:
    raw_value = str(value or "").strip()
    return raw_value if raw_value in {PRICE_MODE_AUTO, *PRICE_MODE_TO_HEADER} else PRICE_MODE_AUTO


def normalize_create_mode(value: str | None) -> str:
    raw_value = str(value or "").strip()
    return raw_value if raw_value in PRICE_MODE_TO_HEADER else "kurumsal_6"


def normalize_vat_mode(value: str | None) -> str:
    raw_value = str(value or "").strip().lower()
    return raw_value if raw_value in VAT_MODE_LABELS else VAT_MODE_INCLUDED


def load_admin_settings() -> dict:
    ensure_admin_storage()
    default_settings = {
        "admin_pin": DEFAULT_ADMIN_PIN,
        "active_price_file": "",
        "active_template_file": "",
        "last_offer_file": "",
        "default_compare_mode": PRICE_MODE_AUTO,
        "default_create_mode": "kurumsal_6",
        "default_vat_mode": VAT_MODE_INCLUDED,
    }
    if not ADMIN_SETTINGS_PATH.exists():
        ADMIN_SETTINGS_PATH.write_text(
            json.dumps(default_settings, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return default_settings
    try:
        raw = json.loads(ADMIN_SETTINGS_PATH.read_text(encoding="utf-8"))
    except Exception:
        raw = {}
    settings = {
        "admin_pin": DEFAULT_ADMIN_PIN,
        "active_price_file": str(raw.get("active_price_file") or ""),
        "active_template_file": str(raw.get("active_template_file") or ""),
        "last_offer_file": str(raw.get("last_offer_file") or ""),
        "default_compare_mode": normalize_compare_mode(raw.get("default_compare_mode")),
        "default_create_mode": normalize_create_mode(raw.get("default_create_mode")),
        "default_vat_mode": normalize_vat_mode(raw.get("default_vat_mode")),
    }
    if settings["last_offer_file"]:
        last_offer_path = resolve_runtime_pdf_path(settings["last_offer_file"])
        if not last_offer_path.exists() or not is_offer_pdf_path(last_offer_path):
            settings["last_offer_file"] = ""
    if settings["active_template_file"]:
        active_template_path = resolve_runtime_pdf_path(settings["active_template_file"])
        if not active_template_path.exists() or not is_template_pdf_path(active_template_path):
            settings["active_template_file"] = ""
    if raw != settings:
        save_admin_settings(settings)
    return settings


def save_admin_settings(settings: dict) -> None:
    ensure_admin_storage()
    ADMIN_SETTINGS_PATH.write_text(
        json.dumps(settings, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def resolve_price_file_path(file_name: str) -> Path:
    if not file_name:
        return BASE_DIR / file_name
    candidate_paths = (
        PRICE_LISTS_DIR / file_name,
        BASE_DIR / file_name,
    )
    for path in candidate_paths:
        if path.exists():
            return path
    return PRICE_LISTS_DIR / file_name


def validate_price_workbook(workbook_path: Path) -> tuple[int, list[str]]:
    price_rows, headers = load_price_rows(workbook_path)
    price_columns = get_price_columns(headers)
    if not price_columns:
        raise ValueError("Excel formatı uygun değil. Fiyat kolonları bulunamadı.")
    return len(price_rows), price_columns


def _price_pdf_word_center(word: tuple) -> tuple[float, float]:
    return (float(word[0]) + float(word[2])) / 2, (float(word[1]) + float(word[3])) / 2


def _price_pdf_words_in(words: list[tuple], x0: float, x1: float, y0: float, y1: float) -> list[tuple]:
    return [
        word
        for word in words
        if x0 <= _price_pdf_word_center(word)[0] <= x1
        and y0 <= _price_pdf_word_center(word)[1] <= y1
    ]


def _price_pdf_line_text(words: list[tuple]) -> str:
    return " ".join(
        str(word[4]).strip()
        for word in sorted(words, key=lambda candidate: float(candidate[0]))
        if str(word[4]).strip()
    ).strip()


def _price_pdf_money(value: str) -> float | None:
    cleaned = re.sub(r"[^0-9,.]", "", value.replace(" ", ""))
    return parse_money(cleaned)


def _extract_price_pdf_rows(pdf_path: Path) -> list[dict[str, float | str | None]]:
    rows: list[dict[str, float | str | None]] = []
    try:
        doc = fitz.open(pdf_path)
    except Exception as exc:
        raise ValueError(f"PDF fiyat listesi okunamadı: {pdf_path.name}") from exc

    try:
        for page in doc:
            words = [word for word in page.get_text("words") if str(word[4]).strip()]
            perakende_words = [word for word in words if str(word[4]).strip().lower() == "perakende"]
            seen_cards: set[tuple[int, int]] = set()

            for marker in sorted(perakende_words, key=lambda word: (float(word[1]), float(word[0]))):
                center_x, center_y = _price_pdf_word_center(marker)
                column_index = next(
                    (
                        index
                        for index, (x0, x1) in enumerate(PRICE_PDF_COLUMN_BOUNDS)
                        if x0 <= center_x <= x1
                    ),
                    None,
                )
                if column_index is None:
                    continue

                card_key = (column_index, round(center_y / 10))
                if card_key in seen_cards:
                    continue
                seen_cards.add(card_key)

                x0, x1 = PRICE_PDF_COLUMN_BOUNDS[column_index]
                title = re.sub(
                    r"\s+",
                    " ",
                    _price_pdf_line_text(_price_pdf_words_in(words, x0, x1, center_y - 25, center_y - 6)),
                ).strip()
                if not title or title.upper().startswith("RAINWATER 23") or title.upper().startswith("FIYAT"):
                    continue

                label_words = _price_pdf_words_in(words, x0, x0 + 75, center_y + 7, center_y + 62)
                label_lines: list[list[float | str]] = []
                for label_word in sorted(label_words, key=lambda word: _price_pdf_word_center(word)[1]):
                    _, label_y = _price_pdf_word_center(label_word)
                    label_text = str(label_word[4]).strip()
                    if label_text != "Nakit" and label_text not in {"4", "6", "9"}:
                        continue
                    if not label_lines or abs(float(label_lines[-1][0]) - label_y) > 4:
                        label_lines.append([label_y, label_text])
                    else:
                        label_lines[-1][1] = f"{label_lines[-1][1]} {label_text}"

                row: dict[str, float | str | None] = {"URUN": title, "NOT": ""}
                notes: list[str] = []
                for label_y, label in label_lines[:3]:
                    price_words = _price_pdf_words_in(
                        words,
                        x0 + 60,
                        x1,
                        float(label_y) - 5.5,
                        float(label_y) + 5.5,
                    )
                    split_x = x0 + ((x1 - x0) * 0.68)
                    perakende_text = _price_pdf_line_text(
                        [word for word in price_words if _price_pdf_word_center(word)[0] < split_x]
                    )
                    kurumsal_text = _price_pdf_line_text(
                        [word for word in price_words if _price_pdf_word_center(word)[0] >= split_x]
                    )

                    first_token = str(label).split()[0]
                    if first_token == "Nakit":
                        suffix = "NAKIT"
                    elif first_token == "4":
                        suffix = "4 TAKSIT"
                    elif first_token in {"6", "9"}:
                        suffix = "6 TAKSIT"
                    else:
                        continue

                    row[f"2026 PERAKENDE {suffix}"] = _price_pdf_money(perakende_text)
                    row[f"2026 KURUMSAL {suffix}"] = _price_pdf_money(kurumsal_text)

                if row.get("2026 PERAKENDE 4 TAKSIT") and not row.get("2026 PERAKENDE 6 TAKSIT"):
                    row["2026 PERAKENDE 6 TAKSIT"] = row.get("2026 PERAKENDE 4 TAKSIT")
                    row["2026 KURUMSAL 6 TAKSIT"] = row.get("2026 KURUMSAL 4 TAKSIT")
                    notes.append(
                        "PDF kartında 6 Taksit satırı yok; sistem uyumu için 4 Taksit değeri 6 Taksit kolonuna kopyalandı."
                    )

                if (
                    title == "Rainwater 40 LT Emaye Tank"
                    and row.get("2026 KURUMSAL NAKIT") == 19950
                    and row.get("2026 PERAKENDE NAKIT") == 24950
                ):
                    row["URUN"] = "Rainwater 80 LT Emaye Tank"

                if any(row.get(header) is not None for header in PRICE_PDF_HEADERS[1:-1]):
                    row["NOT"] = " ".join(notes)
                    rows.append(row)
    finally:
        doc.close()

    if not rows:
        raise ValueError("PDF fiyat listesinde okunabilir ürün/fiyat satırı bulunamadı.")
    return rows


def convert_price_pdf_to_workbook(pdf_path: Path, workbook_path: Path) -> int:
    rows = _extract_price_pdf_rows(pdf_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Fiyat Listesi 2026"
    sheet.append(list(PRICE_PDF_HEADERS))
    for row in rows:
        sheet.append([row.get(header) for header in PRICE_PDF_HEADERS])

    sheet.freeze_panes = "A2"
    for column, width in {
        "A": 42,
        "B": 18,
        "C": 18,
        "D": 18,
        "E": 18,
        "F": 18,
        "G": 18,
        "H": 52,
    }.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=7):
        for cell in row:
            cell.number_format = '#,##0.00 "TL"'

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(workbook_path)
    return len(rows)


def import_price_file(upload: UploadFile, *, activate_after_import: bool) -> tuple[Path, int, list[str]]:
    ensure_admin_storage()
    if not upload.filename:
        raise ValueError("İçe aktarılacak fiyat listesi dosyasını seç.")

    suffix = Path(upload.filename).suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".pdf"}:
        raise ValueError("Sadece .xlsx, .xlsm veya Rainwater PDF fiyat listesi yükleyebilirsin.")

    safe_stem = sanitize_filename_part(Path(upload.filename).stem) or "FIYAT_LISTESI"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target_path = PRICE_LISTS_DIR / f"{timestamp}_{safe_stem}{'.xlsx' if suffix == '.pdf' else suffix}"
    source_pdf_path = PRICE_LISTS_DIR / f"{timestamp}_{safe_stem}.pdf" if suffix == ".pdf" else None

    file_bytes = upload.file.read()
    if not file_bytes:
        raise ValueError("Yüklenen dosya boş görünüyor.")

    try:
        if source_pdf_path is not None:
            source_pdf_path.write_bytes(file_bytes)
            convert_price_pdf_to_workbook(source_pdf_path, target_path)
        else:
            target_path.write_bytes(file_bytes)
        row_count, price_columns = validate_price_workbook(target_path)
    except Exception:
        if target_path.exists():
            target_path.unlink(missing_ok=True)
        if source_pdf_path is not None and source_pdf_path.exists():
            source_pdf_path.unlink(missing_ok=True)
        raise
    finally:
        upload.file.close()

    if activate_after_import:
        settings = load_admin_settings()
        settings["active_price_file"] = target_path.name
        save_admin_settings(settings)

    return target_path, row_count, price_columns


def import_template_file(upload: UploadFile, *, activate_after_import: bool) -> Path:
    ensure_runtime_folders()
    if not upload.filename:
        raise ValueError("Iceri aktarilacak sablon PDF dosyasini sec.")

    suffix = Path(upload.filename).suffix.lower()
    if suffix != ".pdf":
        raise ValueError("Sablon dosyasi PDF olmali.")

    safe_stem = sanitize_filename_part(Path(upload.filename).stem) or "SABLON"
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    target_path = TEMPLATES_PDF_DIR / f"{timestamp}_{safe_stem}{suffix}"

    file_bytes = upload.file.read()
    if not file_bytes:
        raise ValueError("Yuklenen sablon PDF bos gorunuyor.")

    try:
        target_path.write_bytes(file_bytes)
    finally:
        upload.file.close()

    if activate_after_import:
        settings = load_admin_settings()
        settings["active_template_file"] = relative_runtime_path(target_path)
        save_admin_settings(settings)

    return target_path


def import_offer_file(upload: UploadFile) -> Path:
    ensure_runtime_folders()
    if not upload.filename:
        raise ValueError("Kontrol edilecek teklif PDF dosyasını seç.")

    suffix = Path(upload.filename).suffix.lower()
    if suffix != ".pdf":
        raise ValueError("Teklif dosyası PDF olmalı.")

    safe_stem = sanitize_filename_part(Path(upload.filename).stem) or "TEKLIF"
    target_path = OFFERS_DIR / f"{safe_stem}{suffix}"
    if target_path.exists():
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        target_path = OFFERS_DIR / f"{timestamp}_{safe_stem}{suffix}"

    file_bytes = upload.file.read()
    if not file_bytes:
        raise ValueError("Yüklenen teklif PDF dosyası boş görünüyor.")
    if matches_template_pdf(upload.filename, file_bytes):
        raise ValueError(
            "Bu dosya sablonlar klasorundeki PDF ile ayni gorunuyor. "
            "Kontrol icin gercek teklif PDF'ini sec veya dosyayi teklifler klasorune ayri bir teklif olarak ekle."
        )

    try:
        target_path.write_bytes(file_bytes)
    finally:
        upload.file.close()

    return target_path


def matches_template_pdf(filename: str | None, file_bytes: bytes) -> bool:
    if not file_bytes:
        return False
    upload_name = Path(filename or "").name.lower()
    for template_path in list_template_files():
        if upload_name and template_path.name.lower() != upload_name:
            continue
        try:
            if template_path.read_bytes() == file_bytes:
                return True
        except OSError:
            continue
    return False


def activate_price_file(file_name: str) -> Path:
    target_path = resolve_price_file_path(file_name)
    if not target_path.exists():
        raise ValueError("Seçilen fiyat listesi bulunamadı.")
    validate_price_workbook(target_path)
    settings = load_admin_settings()
    settings["active_price_file"] = target_path.name
    save_admin_settings(settings)
    return target_path


def activate_template_file(file_name: str) -> Path:
    target_path = resolve_runtime_pdf_path(file_name)
    if not target_path.exists() or not is_template_pdf_path(target_path):
        raise ValueError("Secilen sablon PDF bulunamadi.")
    settings = load_admin_settings()
    settings["active_template_file"] = relative_runtime_path(target_path)
    save_admin_settings(settings)
    return target_path


def list_price_files() -> list[Path]:
    ensure_admin_storage()
    files_by_name: dict[str, Path] = {}
    for source_dir in (PRICE_LISTS_DIR, BASE_DIR):
        for pattern in ("*.xlsx", "*.xlsm"):
            for path in source_dir.glob(pattern):
                if path.name.startswith("~$") or path.name.startswith("teklif_kontrol_"):
                    continue
                files_by_name.setdefault(path.name, path)
    settings = load_admin_settings()
    active_name = settings.get("active_price_file") or ""
    return sorted(
        files_by_name.values(),
        key=lambda path: (0 if path.name == active_name else 1, path.name.lower()),
    )


def is_inside_base_dir(path: Path) -> bool:
    try:
        path.resolve().relative_to(BASE_DIR.resolve())
        return True
    except ValueError:
        return False


def relative_runtime_path(path: Path) -> str:
    try:
        return path.resolve().relative_to(BASE_DIR.resolve()).as_posix()
    except ValueError:
        return path.name


def resolve_activity_file_path(file_name: str) -> Path:
    raw_name = str(file_name or "").strip()
    if not raw_name:
        return BASE_DIR / raw_name
    raw_path = Path(raw_name)
    candidate = raw_path if raw_path.is_absolute() else BASE_DIR / raw_path
    candidate = candidate.resolve()
    if is_inside_base_dir(candidate):
        return candidate
    return BASE_DIR / raw_path.name


def activity_file_payload(path: Path, label: str, kind: str) -> dict:
    return {
        "label": label,
        "kind": kind,
        "path": relative_runtime_path(path),
        "name": path.name,
    }


def load_activity_log() -> list[dict]:
    ensure_admin_storage()
    if not ACTIVITY_LOG_PATH.exists():
        return []
    try:
        raw = json.loads(ACTIVITY_LOG_PATH.read_text(encoding="utf-8"))
    except Exception:
        return []
    if not isinstance(raw, list):
        return []
    return [entry for entry in raw if isinstance(entry, dict)]


def save_activity_log(entries: list[dict]) -> None:
    ensure_admin_storage()
    limited_entries = entries[:ACTIVITY_LOG_LIMIT]
    temp_path = ACTIVITY_LOG_PATH.with_suffix(".tmp")
    temp_path.write_text(json.dumps(limited_entries, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(ACTIVITY_LOG_PATH)


def load_pending_offer_approvals() -> dict[str, dict]:
    ensure_admin_storage()
    if not PENDING_OFFERS_PATH.exists():
        return {}
    try:
        raw = json.loads(PENDING_OFFERS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}
    if not isinstance(raw, dict):
        return {}
    return {str(key): value for key, value in raw.items() if isinstance(value, dict)}


def save_pending_offer_approvals(entries: dict[str, dict]) -> None:
    ensure_admin_storage()
    temp_path = PENDING_OFFERS_PATH.with_suffix(".tmp")
    temp_path.write_text(json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(PENDING_OFFERS_PATH)


def update_activity_entry_details(entry_id: str, details: dict) -> None:
    entries = load_activity_log()
    for entry in entries:
        if str(entry.get("id") or "") != entry_id:
            continue
        current = entry.get("details") if isinstance(entry.get("details"), dict) else {}
        entry["details"] = {**current, **details}
        break
    save_activity_log(entries)


def register_pending_offer_approval(entry: dict, generated_offer_path: Path, details: dict) -> str:
    approval_id = uuid.uuid4().hex
    user_details = {
        "approval_id": approval_id,
        "activity_entry_id": str(entry.get("id") or ""),
        "status": "pending",
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "approved_at": "",
        "rejected_at": "",
        "approved_by": "",
        "rejected_by": "",
        "generated_path": relative_runtime_path(generated_offer_path),
        "generated_name": generated_offer_path.name,
        **details,
    }
    approvals = load_pending_offer_approvals()
    approvals[approval_id] = user_details
    save_pending_offer_approvals(approvals)
    update_activity_entry_details(
        str(entry.get("id") or ""),
        {
            "approval_id": approval_id,
            "approval_status": "pending",
        },
    )
    return approval_id


def pending_offer_for_activity(entry_id: str) -> dict | None:
    for approval in load_pending_offer_approvals().values():
        if str(approval.get("activity_entry_id") or "") == entry_id:
            return approval
    return None


def approved_offer_mailto(approval: dict, request: Request) -> str:
    recipient = str(approval.get("customer_email") or "").strip()
    subject = f"Teklifiniz onaylandi: {approval.get('offer_number') or approval.get('generated_name') or ''}".strip()
    download_url = ""
    entry_id = str(approval.get("activity_entry_id") or "")
    if entry_id:
        download_url = str(request.url_for("offer-tool:download_activity_file", entry_id=entry_id, file_index=0))
    body_lines = [
        f"Merhaba {approval.get('contact_name') or approval.get('company_name') or ''}".strip(),
        "",
        "Teklifiniz admin onayindan gecti.",
        f"Teklif no: {approval.get('offer_number') or '-'}",
        f"Dosya: {approval.get('generated_name') or '-'}",
    ]
    if download_url:
        body_lines.extend(["", f"Indirme baglantisi: {download_url}"])
    return f"mailto:{quote(recipient)}?subject={quote(subject)}&body={quote(chr(10).join(body_lines))}"


def append_activity_log(
    request: Request,
    *,
    action: str,
    summary: str,
    files: list[dict] | None = None,
    details: dict | None = None,
) -> dict:
    user = get_offer_portal_user(request)
    now = datetime.now()
    entry = {
        "id": uuid.uuid4().hex,
        "created_at": now.isoformat(timespec="seconds"),
        "actor_id": user.id if user else "",
        "actor_email": user.email if user else "bilinmeyen",
        "actor_name": user.full_name if user and user.full_name else "",
        "action": action,
        "action_label": ACTIVITY_ACTION_LABELS.get(action, action),
        "summary": summary,
        "files": files or [],
        "details": details or {},
    }
    save_activity_log([entry, *load_activity_log()])
    return entry


def activity_entry_view_model(request: Request, entry: dict) -> dict:
    files: list[dict] = []
    entry_id = str(entry.get("id") or "")
    approval = pending_offer_for_activity(entry_id)
    approval_status = str((approval or {}).get("status") or "")
    for index, file_info in enumerate(entry.get("files") or []):
        if not isinstance(file_info, dict):
            continue
        path = resolve_activity_file_path(str(file_info.get("path") or ""))
        exists = path.exists() and path.is_file()
        is_blocked_generated_offer = (
            str(entry.get("action") or "") == "create"
            and str(file_info.get("kind") or "") == "generated"
            and approval_status == "pending"
        )
        files.append(
            {
                "label": str(file_info.get("label") or file_info.get("name") or path.name),
                "name": str(file_info.get("name") or path.name),
                "kind": str(file_info.get("kind") or ""),
                "exists": exists,
                "blocked": is_blocked_generated_offer,
                "url": (
                    request.url_for(
                        "offer-tool:download_activity_file",
                        entry_id=entry_id,
                        file_index=index,
                    )
                    if exists and not is_blocked_generated_offer
                    else ""
                ),
            }
        )
    actor = str(entry.get("actor_name") or entry.get("actor_email") or "Bilinmeyen")
    details = entry.get("details") if isinstance(entry.get("details"), dict) else {}
    created_at = str(entry.get("created_at") or "")
    return {
        "id": entry_id,
        "created_at": created_at,
        "created_at_display": created_at.replace("T", " ")[:19],
        "actor": actor,
        "actor_email": str(entry.get("actor_email") or ""),
        "action": str(entry.get("action") or ""),
        "action_label": str(entry.get("action_label") or ACTIVITY_ACTION_LABELS.get(str(entry.get("action") or ""), "")),
        "summary": str(entry.get("summary") or ""),
        "details": details,
        "approval": approval or {},
        "approval_id": str((approval or {}).get("approval_id") or ""),
        "approval_status": approval_status,
        "approval_is_pending": approval_status == "pending",
        "approval_is_approved": approval_status == "approved",
        "approval_is_rejected": approval_status == "rejected",
        "approved_mailto": approved_offer_mailto(approval, request) if approval_status == "approved" else "",
        "files": files,
        "file_count": len(files),
        "search_text": " ".join(
            [
                actor,
                str(entry.get("actor_email") or ""),
                str(entry.get("action_label") or ""),
                str(entry.get("summary") or ""),
                " ".join(file_item["name"] for file_item in files),
            ]
        ).lower(),
    }


def resolve_runtime_pdf_path(file_name: str) -> Path:
    raw_name = str(file_name or "").strip()
    if not raw_name:
        return BASE_DIR / raw_name

    raw_path = Path(raw_name)
    candidate = raw_path if raw_path.is_absolute() else BASE_DIR / raw_path
    candidate = candidate.resolve()
    if is_inside_base_dir(candidate):
        if candidate.exists() or raw_path.parent != Path("."):
            return candidate
    else:
        return BASE_DIR / raw_path.name

    for source_dir in (BASE_DIR, OFFERS_DIR, TEMPLATES_PDF_DIR):
        fallback = source_dir / raw_path.name
        if fallback.exists():
            return fallback
    return BASE_DIR / raw_path.name


def is_template_pdf_path(path: Path) -> bool:
    try:
        path.resolve().relative_to(TEMPLATES_PDF_DIR.resolve())
        return True
    except ValueError:
        return False


def is_offer_pdf_path(path: Path) -> bool:
    try:
        path.resolve().relative_to(OFFERS_DIR.resolve())
        return True
    except ValueError:
        return False


def build_app_corrected_pdf_path(offer_path: Path) -> Path:
    target_dir = BASE_DIR / OUTPUT_ROOT_DIRNAME / CORRECTED_PDFS_DIRNAME
    target_dir.mkdir(parents=True, exist_ok=True)
    cleaned_stem = sanitize_filename_part(offer_path.stem) or "TEKLIF"
    candidate = target_dir / f"{cleaned_stem}_duzeltilmis.pdf"
    if not candidate.exists():
        return candidate

    for index in range(2, 1000):
        candidate = target_dir / f"{cleaned_stem}_duzeltilmis_{index:02d}.pdf"
        if not candidate.exists():
            return candidate
    raise RuntimeError("Duzeltilmis PDF icin bos dosya adi bulunamadi.")


def build_batch_summary_path(batch_token: str) -> Path:
    target_dir = BASE_DIR / OUTPUT_ROOT_DIRNAME / REPORTS_DIRNAME
    target_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return target_dir / f"toplu_kontrol_{timestamp}_{batch_token[:8]}.xlsx"


def remember_last_files(*, price_file_name: str | None = None, offer_file_name: str | None = None) -> None:
    settings = load_admin_settings()
    changed = False

    if price_file_name and resolve_price_file_path(price_file_name).exists():
        settings["active_price_file"] = price_file_name
        changed = True

    if offer_file_name:
        offer_path = resolve_runtime_pdf_path(offer_file_name)
        if offer_path.exists() and is_offer_pdf_path(offer_path):
            settings["last_offer_file"] = offer_file_name
            changed = True

    if changed:
        save_admin_settings(settings)


def list_pdf_files_from(*source_dirs: Path) -> list[Path]:
    ensure_runtime_folders()
    files_by_key: dict[str, Path] = {}
    for source_dir in source_dirs:
        if not source_dir.exists():
            continue
        for path in source_dir.glob("*.pdf"):
            if path.name.lower().endswith("_duzeltilmis.pdf"):
                continue
            files_by_key.setdefault(relative_runtime_path(path), path)
    return sorted(files_by_key.values(), key=lambda path: relative_runtime_path(path).lower())


def list_offer_files() -> list[Path]:
    return list_pdf_files_from(OFFERS_DIR)


def list_template_files() -> list[Path]:
    return list_pdf_files_from(TEMPLATES_PDF_DIR)


def default_price_file() -> str:
    settings = load_admin_settings()
    active_name = settings.get("active_price_file") or ""
    if active_name and resolve_price_file_path(active_name).exists():
        return active_name
    files = list_price_files()
    return files[0].name if files else ""


def default_offer_file() -> str:
    settings = load_admin_settings()
    last_offer_file = settings.get("last_offer_file") or ""
    if last_offer_file:
        last_offer_path = resolve_runtime_pdf_path(last_offer_file)
        if last_offer_path.exists() and is_offer_pdf_path(last_offer_path):
            return last_offer_file

    files = list_offer_files()
    return relative_runtime_path(files[0]) if files else ""


def default_template_file() -> str:
    settings = load_admin_settings()
    active_template_file = settings.get("active_template_file") or ""
    if active_template_file:
        active_template_path = resolve_runtime_pdf_path(active_template_file)
        if active_template_path.exists() and is_template_pdf_path(active_template_path):
            return relative_runtime_path(active_template_path)

    try:
        files = list_template_files()
        return relative_runtime_path(files[0]) if files else ""
    except Exception:
        return ""


def load_catalog_options(price_file_name: str) -> list[dict]:
    if not price_file_name:
        return []
    price_file_path = resolve_price_file_path(price_file_name)
    if not price_file_path.exists():
        return []
    try:
        price_rows, headers = load_price_rows(price_file_path)
        price_columns = get_price_columns(headers)
        if not price_columns:
            return []
        resolved_columns: dict[str, str | None] = {}
        standard_columns: dict[str, str | None] = {}
        for mode, header in PRICE_MODE_TO_HEADER.items():
            try:
                resolved_columns[mode] = resolve_price_column(header, price_columns)
            except ValueError:
                resolved_columns[mode] = None
            standard_columns[mode] = (
                derive_standard_price_header(resolved_columns[mode], price_columns)
                if resolved_columns[mode]
                else None
            )
        return [
            {
                "value": str(row.row_number),
                "label": row.product_name,
                "prices": {
                    mode: {
                        "included": (
                            resolve_price_for_vat_mode(
                                row,
                                resolved_header,
                                vat_included=True,
                                vat_rate=DEFAULT_VAT_RATE,
                            )[0]
                            if resolved_header
                            else None
                        ),
                        "excluded": (
                            resolve_price_for_vat_mode(
                                row,
                                resolved_header,
                                vat_included=False,
                                vat_rate=DEFAULT_VAT_RATE,
                            )[0]
                            if resolved_header
                            else None
                        ),
                        "excluded_source": (
                            resolve_price_for_vat_mode(
                                row,
                                resolved_header,
                                vat_included=False,
                                vat_rate=DEFAULT_VAT_RATE,
                            )[2]
                            if resolved_header
                            else "missing"
                        ),
                        "standard_included": (
                            resolve_price_for_vat_mode(
                                row,
                                standard_columns.get(mode),
                                vat_included=True,
                                vat_rate=DEFAULT_VAT_RATE,
                            )[0]
                            if standard_columns.get(mode)
                            else None
                        ),
                        "standard_excluded": (
                            resolve_price_for_vat_mode(
                                row,
                                standard_columns.get(mode),
                                vat_included=False,
                                vat_rate=DEFAULT_VAT_RATE,
                            )[0]
                            if standard_columns.get(mode)
                            else None
                        ),
                    }
                    for mode, resolved_header in resolved_columns.items()
                },
            }
            for row in price_rows
        ]
    except Exception:
        return []


def default_create_state(price_file_name: str | None = None) -> dict:
    today = date.today()
    settings = load_admin_settings()
    return {
        "template_file": default_template_file(),
        "price_file": price_file_name or default_price_file(),
        "price_mode": settings.get("default_create_mode") or "kurumsal_6",
        "vat_mode": settings.get("default_vat_mode") or VAT_MODE_INCLUDED,
        "offer_number": default_offer_number(today, BASE_DIR),
        "offer_date": today.isoformat(),
        "valid_until": default_valid_until(today).isoformat(),
        "company_name": "",
        "contact_name": "",
        "email": "",
        "gsm": "",
        "note": "",
        "items": [
            {
                "product_row_id": "",
                "quantity": "1",
                "manual_price": "",
                "discount_type": DISCOUNT_TYPE_NONE,
                "discount_value": "",
            }
            for _ in range(CREATE_FORM_MIN_ROWS)
        ],
    }


def summarize_note(result: MatchResult) -> str:
    note = result.note.lower()
    if "elle seçildi" in note:
        return "Elle secildi"
    if "bedelsiz" in note:
        return "Bedelsiz"
    if "fiyat yok" in note:
        return "Kolonda fiyat yok"
    if "skoru dusuk" in note:
        return "Düşük eşleşme"
    if result.status == "DUZELT":
        return "Fark var"
    if result.status == "ONAY":
        return "Uygun"
    if result.status == "ESLESMEDI":
        return "Eşleşmedi"
    return "İncele"


def grouped_price_modes() -> list[dict]:
    groups: list[dict] = []
    for group_title, modes in PRICE_MODE_GROUPS:
        groups.append(
            {
                "title": group_title,
                "options": [{"value": mode, "label": PRICE_MODE_LABELS[mode]} for mode in modes],
            }
        )
    return groups


def infer_price_mode_from_selected_column(selected_column: str) -> str:
    normalized_column = normalize_text(selected_column)
    for mode, header in PRICE_MODE_TO_HEADER.items():
        if normalize_text(header) in normalized_column:
            return mode
    return PRICE_MODE_AUTO


def describe_result_next_step(result: MatchResult, *, can_apply: bool, manual_selected: bool) -> str:
    if manual_selected:
        if can_apply:
            return "Secilen urune gore fiyat farkli. Kutuyu isaretleyip PDF hazirlama kuyruguna al."
        if result.status == "ONAY":
            return "Secilen urune gore teklif fiyati dogru. Bu satir tamam."
        if result.status == "INCELE":
            return "Secilen urunle karsilastirma yapildi ama otomatik onay verilmedi. Notu kontrol et."
        return "Elle urun secildi ama sonuc netlesmedi. Baska bir urun secmeyi dene."

    if can_apply:
        return "Fiyat farki bulundu. Kutuyu isaretlersen yeni PDF hazirlama kuyruguna alinir."
    if result.status == "ONAY":
        return "Bu satir zaten uyumlu. Islem gerekmez."
    if result.status == "INCELE":
        return "Bu satir sadece manuel kontrol icindir."
    return "Excel listesinden urun secip yeniden karsilastir."


def product_match_suggestions(
    result: MatchResult,
    price_rows: list,
    *,
    selected_column: str,
    vat_included: bool,
    vat_rate: float,
    limit: int = 3,
) -> list[dict]:
    if result.bundle_components:
        return [
            {
                "value": str(component.matched_row.row_number),
                "label": component.matched_row.product_name,
                "score": int(round(component.score * 100)),
                "price": format_money(component.reference_unit_price),
                "is_current": True,
            }
            for component in result.bundle_components[:limit]
        ]

    candidates: list[dict] = []
    for row in price_rows:
        score = similarity_score(result.offer_item.product_name, row.product_name)
        if result.matched_row and row.row_number == result.matched_row.row_number:
            score = max(score, result.score)
        if score <= 0:
            continue
        reference_price, _, _ = resolve_price_for_vat_mode(
            row,
            selected_column,
            vat_included=vat_included,
            vat_rate=vat_rate,
        )
        candidates.append(
            {
                "value": str(row.row_number),
                "label": row.product_name,
                "score": int(round(score * 100)),
                "price": format_money(reference_price),
                "is_current": bool(result.matched_row and row.row_number == result.matched_row.row_number),
            }
        )

    candidates.sort(key=lambda item: (item["is_current"], item["score"]), reverse=True)
    filtered: list[dict] = []
    seen: set[str] = set()
    for candidate in candidates:
        if candidate["value"] in seen:
            continue
        seen.add(candidate["value"])
        if candidate["is_current"] or candidate["score"] >= 28:
            filtered.append(candidate)
        if len(filtered) >= limit:
            break
    return filtered


def describe_result_action(result: MatchResult, *, can_apply: bool, suggestions: list[dict]) -> dict:
    if can_apply:
        if result.bundle_components:
            return {
                "tone": "danger",
                "title": "Birleşik fiyat düzeltmesi hazır",
                "body": (
                    f"Teklifte {format_money(result.offer_item.discounted_price)} görünüyor; "
                    f"{len(result.bundle_components)} ürünün liste toplamı {format_money(result.suggested_unit_price)}."
                ),
            }
        return {
            "tone": "danger",
            "title": "Fiyat düzeltmesi hazır",
            "body": (
                f"Teklifte {format_money(result.offer_item.discounted_price)} görünüyor; "
                f"listeye göre {format_money(result.suggested_unit_price)} olmalı."
            ),
        }
    if result.status == "ESLESMEDI":
        if suggestions:
            return {
                "tone": "warn",
                "title": "Ürün eşleşmedi, yakın aday var",
                "body": f"Sistem en yakın aday olarak {suggestions[0]['label']} ürününü görüyor.",
            }
        return {
            "tone": "warn",
            "title": "Ürün eşleşmedi",
            "body": "Fiyat listesinden ürün seçilmeden bu satır PDF düzeltmesine alınamaz.",
        }
    if result.status == "INCELE":
        if suggestions:
            return {
                "tone": "warn",
                "title": "Eşleşme güveni düşük",
                "body": f"Önce {suggestions[0]['label']} adayını kontrol et.",
            }
        return {
            "tone": "warn",
            "title": "Manuel kontrol gerekiyor",
            "body": "Satır otomatik onay için yeterince net değil.",
        }
    if result.status == "ONAY":
        return {
            "tone": "ok",
            "title": "Satır uyumlu",
            "body": "Ürün ve fiyat kontrolünden geçti.",
        }
    return {
        "tone": "info",
        "title": "Kontrol notu",
        "body": result.note,
    }


def describe_price_difference(result: MatchResult) -> dict:
    if result.difference is None or result.reference_unit_price is None:
        return {
            "tone": "neutral",
            "label": "Fark net değil",
            "text": "Liste fiyatı veya güvenilir ürün eşleşmesi olmadığı için fiyat farkı hesaplanamadı.",
        }

    absolute_difference = abs(result.difference)
    if absolute_difference <= 1:
        return {
            "tone": "even",
            "label": "Fiyat uyumlu",
            "text": "Teklif fiyatı seçilen liste fiyatıyla uyumlu görünüyor.",
        }
    if result.difference > 0:
        return {
            "tone": "high",
            "label": "Teklif pahalı",
            "text": f"Teklifteki birim fiyat liste fiyatından {format_money(absolute_difference)} yüksek.",
        }
    return {
        "tone": "low",
        "label": "Teklif ucuz",
        "text": f"Teklifteki birim fiyat liste fiyatından {format_money(absolute_difference)} düşük.",
    }


def describe_match_state(result: MatchResult, *, suggestions: list[dict]) -> dict:
    if result.status == "ESLESMEDI":
        if suggestions:
            return {
                "tone": "warn",
                "label": "Ürün eşleşmedi",
                "text": f"Sistem bu satırı kesin eşleştiremedi; en yakın aday {suggestions[0]['label']} olarak görünüyor.",
            }
        return {
            "tone": "warn",
            "label": "Ürün eşleşmedi",
            "text": "Excel fiyat listesinde güvenilir bir ürün karşılığı bulunamadı.",
        }
    if result.bundle_components:
        return {
            "tone": "info",
            "label": "Birleşik ürün",
            "text": f"Satır {len(result.bundle_components)} bileşen üzerinden değerlendirildi; bileşen eşleşmelerini kontrol etmek gerekir.",
        }
    if "düşük" in result.note.lower() or "dusuk" in result.note.lower():
        return {
            "tone": "warn",
            "label": "Eşleşme düşük güvenli",
            "text": "Ürün adı benziyor ama otomatik onay için güven seviyesi düşük kaldı.",
        }
    if result.note.startswith("Ürün elle seçildi."):
        return {
            "tone": "info",
            "label": "Elle seçildi",
            "text": "Karşılaştırma, kullanıcının elle seçtiği Excel ürünüyle yeniden yapıldı.",
        }
    if result.matched_row is not None:
        return {
            "tone": "ok",
            "label": "Ürün eşleşti",
            "text": f"PDF satırı Excel'deki '{result.matched_row.product_name}' ürünüyle eşleştirildi.",
        }
    return {
        "tone": "warn",
        "label": "Eşleşme belirsiz",
        "text": "Ürün karşılığı netleşmediği için manuel kontrol gerekir.",
    }


def describe_reference_context(result: MatchResult) -> str:
    if "KDV haric referans fiyat brut listeden hesaplandi" in result.note:
        return "Referans fiyat KDV dahil listeden nete çevrilerek hesaplandı."
    if "KDV dahil referans fiyat net listeden hesaplandi" in result.note:
        return "Referans fiyat KDV hariç listeden brüte çevrilerek hesaplandı."
    if "KDV haric bilesen fiyati brut listeden hesaplandi" in result.note:
        return "Bileşen fiyatları KDV dahil listeden nete çevrilerek hesaplandı."
    if "KDV dahil bilesen fiyati net listeden hesaplandi" in result.note:
        return "Bileşen fiyatları KDV hariç listeden brüte çevrilerek hesaplandı."
    return f"Karşılaştırma '{result.selected_column}' fiyat tipiyle yapıldı."


def build_result_explanation(
    result: MatchResult,
    *,
    can_apply: bool,
    suggestions: list[dict],
) -> dict:
    price_state = describe_price_difference(result)
    match_state = describe_match_state(result, suggestions=suggestions)
    reference_context = describe_reference_context(result)

    if can_apply:
        title = "Liste fiyatı uyuşmuyor, düzeltme hazır."
        next_step = "Bu satırı PDF düzeltme kuyruğuna alıp önerilen liste fiyatını yazdır."
        cta = "PDF düzeltmesine al"
        tone = "danger"
    elif result.status == "ESLESMEDI":
        title = "Ürün bulunamadığı için fiyat kararı verilemiyor."
        next_step = "Excel listesinden doğru ürünü seç; sistem fiyatı tekrar hesaplasın."
        cta = "Ürün seç"
        tone = "warn"
    elif result.status == "INCELE":
        title = "Otomatik karar için güven yeterli değil."
        next_step = "Ürün eşleşmesini ve KDV/fiyat tipini kontrol ettikten sonra karar ver."
        cta = "İncele"
        tone = "warn"
    elif result.status == "ONAY":
        title = "Bu kalemde fiyat ve eşleşme uyumlu."
        next_step = "Ek işlem gerekmez; sadece raporda saklanır."
        cta = "Tamam"
        tone = "ok"
    else:
        title = "Bu kalem için sistem notunu kontrol et."
        next_step = result.note
        cta = "Kontrol et"
        tone = "info"

    return {
        "tone": tone,
        "title": title,
        "spoken": f"{match_state['text']} {price_state['text']} {next_step}",
        "match": match_state,
        "price": price_state,
        "reference": reference_context,
        "next_step": next_step,
        "cta": cta,
    }


def result_view_model(
    result: MatchResult,
    index: int,
    *,
    price_rows: list | None = None,
    selected_column: str = "",
    vat_included: bool = True,
    vat_rate: float = DEFAULT_VAT_RATE,
) -> dict:
    bundle_components = [
        {
            "requested_name": component.requested_name,
            "row_value": str(component.matched_row.row_number),
            "matched_name": component.matched_row.product_name,
            "score": int(round(component.score * 100)),
            "price": format_money(component.reference_unit_price),
            "ambiguous": component.ambiguous,
        }
        for component in result.bundle_components
    ]
    matched_name = (
        " + ".join(component["matched_name"] for component in bundle_components)
        if bundle_components
        else result.matched_row.product_name if result.matched_row else "-"
    )
    manual_selected = result.note.startswith("Ürün elle seçildi.")
    can_apply = (
        result.status == "DUZELT"
        and result.suggested_unit_price is not None
        and result.suggested_total_price is not None
    )
    suggestions = product_match_suggestions(
        result,
        price_rows or [],
        selected_column=selected_column or result.selected_column,
        vat_included=vat_included,
        vat_rate=vat_rate,
    )
    apply_hint = describe_result_next_step(result, can_apply=can_apply, manual_selected=manual_selected)
    action = describe_result_action(result, can_apply=can_apply, suggestions=suggestions)
    explanation = build_result_explanation(result, can_apply=can_apply, suggestions=suggestions)
    return {
        "index": index,
        "status": result.status,
        "status_class": STATUS_CLASS.get(result.status, "info"),
        "row_class": f"status-{result.status.lower()}",
        "product_name": result.offer_item.product_name,
        "matched_name": matched_name,
        "is_bundle": bool(bundle_components),
        "bundle_components": bundle_components,
        "manual_selected": manual_selected,
        "manual_match_row_id": str(result.matched_row.row_number) if result.matched_row and manual_selected else "",
        "offer_price": format_money(result.offer_item.discounted_price),
        "reference_price": format_money(result.reference_unit_price),
        "difference": format_money(result.difference),
        "suggested_price": format_money(result.suggested_unit_price),
        "quantity_value": result.offer_item.quantity,
        "offer_unit_price_value": result.offer_item.discounted_price,
        "offer_total_price_value": result.offer_item.total_price,
        "note_badge": summarize_note(result),
        "note": result.note,
        "selected_column": result.selected_column,
        "apply_hint": apply_hint,
        "action": action,
        "explanation": explanation,
        "suggestions": suggestions,
        "suggestion_count": len(suggestions),
        "can_apply": can_apply,
    }


def apply_manual_match_overrides(
    session: ComparisonSession,
    manual_match_row_ids: list[str] | None,
) -> int:
    if not manual_match_row_ids or not any(str(value or "").strip() for value in manual_match_row_ids):
        return 0

    price_rows, _ = load_price_rows(session.price_list_path)
    rows_by_number = {row.row_number: row for row in price_rows}
    updated_count = 0

    for index, result in enumerate(session.results):
        raw_row_id = manual_match_row_ids[index] if index < len(manual_match_row_ids) else ""
        if not str(raw_row_id or "").strip():
            continue
        try:
            row_number = int(str(raw_row_id).strip())
        except ValueError as exc:
            raise ValueError("Elle secilen urun satirlarindan biri gecersiz.") from exc

        selected_row = rows_by_number.get(row_number)
        if selected_row is None:
            raise ValueError(f"Elle secilen urun satiri bulunamadi: {row_number}")

        session.results[index] = build_match_result(
            result.offer_item,
            selected_row,
            session.selected_column,
            tolerance=1.0,
            vat_included=session.financial_review.vat_included,
            vat_rate=session.financial_review.vat_rate,
            manual_override=True,
        )
        updated_count += 1

    return updated_count


def _build_manual_bundle_result(
    result: MatchResult,
    component_matches: list[BundleComponentMatch],
    *,
    selected_column: str,
    vat_included: bool,
    vat_rate: float,
    tolerance: float = 1.0,
) -> MatchResult:
    reference_unit_price = round(sum(component.reference_unit_price for component in component_matches), 2)
    reference_total_price = round(reference_unit_price * result.offer_item.quantity, 2)
    difference = round(result.offer_item.discounted_price - reference_unit_price, 2)
    total_difference = round(result.offer_item.total_price - reference_total_price, 2)
    status = "ONAY" if abs(difference) <= tolerance and abs(total_difference) <= tolerance else "DUZELT"
    component_summary = "; ".join(
        f"{component.requested_name} -> {component.matched_row.product_name} ({format_money(component.reference_unit_price)})"
        for component in component_matches
    )
    note = (
        "Bileşenler elle güncellendi; toplam fiyat teklif satırıyla uyumlu."
        if status == "ONAY"
        else "Bileşenler elle güncellendi; bileşen toplamı teklif satırıyla uyuşmuyor."
    )
    note = f"{note} Bileşenler: {component_summary}"
    return MatchResult(
        offer_item=result.offer_item,
        matched_row=component_matches[0].matched_row if component_matches else result.matched_row,
        score=min((component.score for component in component_matches), default=result.score),
        status=status,
        selected_column=selected_column,
        reference_unit_price=reference_unit_price,
        reference_total_price=reference_total_price,
        suggested_unit_price=reference_unit_price if status == "DUZELT" else None,
        suggested_total_price=reference_total_price if status == "DUZELT" else None,
        difference=difference,
        note=note,
        bundle_components=component_matches,
    )


def apply_bundle_match_overrides(
    session: ComparisonSession,
    bundle_match_overrides: list[str] | None,
) -> int:
    override_values = [str(value or "").strip() for value in (bundle_match_overrides or []) if str(value or "").strip()]
    if not override_values:
        return 0

    price_rows, _ = load_price_rows(session.price_list_path)
    rows_by_number = {row.row_number: row for row in price_rows}
    grouped: dict[int, dict[int, int]] = {}
    for raw_value in override_values:
        parts = raw_value.split(":", 2)
        if len(parts) != 3:
            raise ValueError("Bileşen ürün seçimlerinden biri geçersiz.")
        try:
            result_index = int(parts[0])
            component_index = int(parts[1])
            row_number = int(parts[2])
        except ValueError as exc:
            raise ValueError("Bileşen ürün seçimlerinden biri geçersiz.") from exc
        grouped.setdefault(result_index, {})[component_index] = row_number

    updated_count = 0
    for result_index, component_overrides in grouped.items():
        if result_index < 0 or result_index >= len(session.results):
            raise ValueError(f"Bileşen seçimi satırı bulunamadı: {result_index + 1}")
        result = session.results[result_index]
        if not result.bundle_components:
            continue

        next_components: list[BundleComponentMatch] = []
        changed = False
        for component_index, component in enumerate(result.bundle_components):
            row_number = component_overrides.get(component_index)
            if row_number is None:
                next_components.append(component)
                continue
            selected_row = rows_by_number.get(row_number)
            if selected_row is None:
                raise ValueError(f"Bileşen için seçilen ürün satırı bulunamadı: {row_number}")
            reference_unit_price, reference_column, reference_source = resolve_price_for_vat_mode(
                selected_row,
                session.selected_column,
                vat_included=session.financial_review.vat_included,
                vat_rate=session.financial_review.vat_rate,
            )
            if reference_unit_price is None:
                raise ValueError(f"Seçilen bileşen için fiyat bulunamadı: {selected_row.product_name}")
            next_components.append(
                BundleComponentMatch(
                    requested_name=component.requested_name,
                    matched_row=selected_row,
                    score=similarity_score(component.requested_name, selected_row.product_name),
                    reference_unit_price=reference_unit_price,
                    reference_column=reference_column,
                    reference_source=reference_source,
                )
            )
            changed = True

        if changed:
            session.results[result_index] = _build_manual_bundle_result(
                result,
                next_components,
                selected_column=session.selected_column,
                vat_included=session.financial_review.vat_included,
                vat_rate=session.financial_review.vat_rate,
            )
            updated_count += 1

    return updated_count


def format_rate_label(value: float) -> str:
    if float(value).is_integer():
        return f"%{int(value)}"
    return f"%{value:.2f}".replace(".", ",")


def build_financial_explanation(check: FinancialCheck) -> dict:
    if check.status == "ONAY":
        return {
            "tone": "ok",
            "title": f"{check.label} uyumlu.",
            "text": check.note,
            "next_step": "Finansal tarafta ek işlem gerekmiyor.",
        }
    if check.status == "DUZELT":
        difference_text = ""
        if check.difference is not None:
            direction = "yüksek" if check.difference > 0 else "düşük"
            difference_text = f" PDF değeri hesaplanan değerden {format_money(abs(check.difference))} {direction}."
        return {
            "tone": "danger",
            "title": f"{check.label} hesabı uyuşmuyor.",
            "text": f"{check.note}{difference_text}",
            "next_step": "PDF özetindeki toplam, KDV veya kalem toplamlarını yeniden kontrol et.",
        }
    return {
        "tone": "warn",
        "title": f"{check.label} için manuel kontrol gerekli.",
        "text": check.note,
        "next_step": "PDF'te ilgili alan okunamadıysa teklif şablonundaki yazımı netleştir.",
    }


def financial_check_view_model(check: FinancialCheck) -> dict:
    return {
        "label": check.label,
        "status": check.status,
        "status_class": STATUS_CLASS.get(check.status, "info"),
        "offer_value": format_money(check.offer_value),
        "calculated_value": format_money(check.calculated_value),
        "difference": format_money(check.difference),
        "note": check.note,
        "explanation": build_financial_explanation(check),
    }


def relative_display_path(path: Path | None) -> str:
    if path is None:
        return "-"
    try:
        return path.relative_to(BASE_DIR).as_posix()
    except ValueError:
        return path.name


def result_metrics(results: list[MatchResult]) -> dict[str, int]:
    counts = Counter(result.status for result in results)
    return {status: int(counts.get(status, 0)) for status in STATUS_CLASS}


def build_ai_action_board(
    result_rows: list[dict],
    financial_checks: list[dict],
    metrics: dict[str, int],
    *,
    financial_overall_status: str,
) -> dict:
    problem_rows = [row for row in result_rows if row["status"] != "ONAY"]
    financial_problems = [check for check in financial_checks if check["status"] != "ONAY"]
    action_rows = [row for row in result_rows if row["can_apply"]]
    waiting_rows = [row for row in result_rows if row["status"] in {"INCELE", "ESLESMEDI"}]

    if not problem_rows and not financial_problems and result_rows:
        headline = "Teklif temiz görünüyor."
        summary = (
            f"{metrics.get('ONAY', 0)} kalem liste fiyatıyla uyumlu. "
            "Finansal özet de sorun üretmediği için teklif onaya hazır."
        )
        tone = "ok"
        next_focus = "Raporu arşivle veya teklif sürecini kapat."
    elif action_rows:
        first = action_rows[0]
        headline = "Düzeltmeye hazır fiyat farkı var."
        summary = (
            f"Öncelik Kalem {first['index'] + 1}: {first['explanation']['price']['text']} "
            f"Sistem önerisi: {first['suggested_price']}."
        )
        tone = "danger"
        next_focus = "Hazır aksiyonları Karar Merkezi'nde onaylayıp yeni PDF üret."
    elif waiting_rows:
        first = waiting_rows[0]
        headline = "Önce ürün eşleşmesi netleşmeli."
        summary = (
            f"Kalem {first['index'] + 1} için otomatik karar güvenli değil. "
            f"{first['explanation']['match']['text']}"
        )
        tone = "warn"
        next_focus = "Doğru Excel ürününü seç; sistem fiyatı yeniden açıklasın."
    elif financial_problems:
        first_check = financial_problems[0]
        headline = "Kalemler temiz olsa bile finansal özet uyarıyor."
        summary = first_check["explanation"]["text"]
        tone = first_check["explanation"]["tone"]
        next_focus = first_check["explanation"]["next_step"]
    else:
        headline = "Teklifte kontrol edilmesi gereken alan var."
        summary = "Sistem net bir otomatik karar üretemedi; sonuç kartlarını sırayla kontrol et."
        tone = "info"
        next_focus = "Problemli satırları açıp ürün ve fiyat tipini doğrula."

    cards: list[dict] = []
    for row in problem_rows[:3]:
        cards.append(
            {
                "kind": "Kalem",
                "title": f"Kalem {row['index'] + 1}: {row['explanation']['title']}",
                "body": row["explanation"]["spoken"],
                "tone": row["explanation"]["tone"],
                "target": row["index"],
            }
        )
    for check in financial_problems[:2]:
        cards.append(
            {
                "kind": "Finans",
                "title": check["explanation"]["title"],
                "body": check["explanation"]["text"],
                "tone": check["explanation"]["tone"],
                "target": "",
            }
        )

    return {
        "tone": tone,
        "headline": headline,
        "summary": summary,
        "next_focus": next_focus,
        "cards": cards[:4],
        "has_actions": bool(cards),
        "financial_status": financial_overall_status,
    }


def create_comparison_session(price_list_path: Path, offer_path: Path, price_mode: str) -> ComparisonSession:
    price_mode = normalize_compare_mode(price_mode)
    selected_column = None
    if price_mode != PRICE_MODE_AUTO:
        selected_column = PRICE_MODE_TO_HEADER.get(price_mode)
        if selected_column is None:
            raise ValueError("Geçersiz fiyat tipi seçimi.")

    results, used_column, final_output_path, _, financial_review = run_comparison(
        price_list_path=price_list_path,
        offer_path=offer_path,
        price_column=selected_column,
        output_path=build_report_output_path(BASE_DIR, offer_path),
    )
    session = ComparisonSession(
        token=uuid.uuid4().hex,
        price_list_path=price_list_path,
        offer_path=offer_path,
        output_path=final_output_path,
        selected_column=used_column,
        price_mode=price_mode,
        results=results,
        financial_review=financial_review,
    )
    SESSIONS[session.token] = session
    return session


def create_demo_comparison_session() -> ComparisonSession:
    selected_column = PRICE_MODE_TO_HEADER["kurumsal_nakit"]
    price_list_path = resolve_price_file_path(default_price_file())
    offer_path = OFFERS_DIR / "AI_ORNEK_TEKLIF.pdf"
    output_path = BASE_DIR / OUTPUT_ROOT_DIRNAME / REPORTS_DIRNAME / "AI_ORNEK_RAPOR.xlsx"
    matched_filter = PriceRow(
        row_number=2,
        product_name="Rainwater Black Serisi 3 Aşama Filtre Seti",
        prices={selected_column: 18500.0},
    )
    matched_softener = PriceRow(
        row_number=3,
        product_name="Rainwater Compact Yumuşatma Sistemi",
        prices={selected_column: 42000.0},
    )
    results = [
        MatchResult(
            offer_item=OfferItem(
                product_name="Rainwater Black Serisi 3 Aşama Filtre Seti",
                quantity=1,
                unit_price=21500.0,
                discounted_price=21500.0,
                total_price=21500.0,
            ),
            matched_row=matched_filter,
            score=0.96,
            status="DUZELT",
            selected_column=selected_column,
            reference_unit_price=18500.0,
            reference_total_price=18500.0,
            suggested_unit_price=18500.0,
            suggested_total_price=18500.0,
            difference=3000.0,
            note="Teklif fiyatı seçilen liste fiyatından farklı.",
        ),
        MatchResult(
            offer_item=OfferItem(
                product_name="Rainwater Compact Yumuşatma Sistemi",
                quantity=1,
                unit_price=42000.0,
                discounted_price=42000.0,
                total_price=42000.0,
            ),
            matched_row=matched_softener,
            score=0.94,
            status="ONAY",
            selected_column=selected_column,
            reference_unit_price=42000.0,
            reference_total_price=42000.0,
            suggested_unit_price=None,
            suggested_total_price=None,
            difference=0.0,
            note="Teklif fiyatı liste ile uyumlu.",
        ),
        MatchResult(
            offer_item=OfferItem(
                product_name="Rainwater Premium Pompalı Arıtmalı Sistem",
                quantity=1,
                unit_price=28500.0,
                discounted_price=28500.0,
                total_price=28500.0,
            ),
            matched_row=None,
            score=0.42,
            status="ESLESMEDI",
            selected_column=selected_column,
            reference_unit_price=None,
            reference_total_price=None,
            suggested_unit_price=None,
            suggested_total_price=None,
            difference=None,
            note="Excel listesinde güvenilir eşleşme bulunamadı.",
        ),
    ]
    financial_review = FinancialReview(
        vat_rate=20.0,
        vat_rate_source="PDF",
        vat_included=True,
        item_gross_total=92000.0,
        expected_net_total=76666.67,
        expected_vat_total=15333.33,
        expected_gross_total=92000.0,
        expected_summary_total=92000.0,
        checks=[
            FinancialCheck(
                label="Ürünlerin Toplam Tutarı",
                status="ONAY",
                offer_value=92000.0,
                calculated_value=92000.0,
                difference=0.0,
                note="PDF'teki ürün toplam tutarları birim fiyat x adet hesabıyla uyumlu.",
            ),
            FinancialCheck(
                label="KDV",
                status="DUZELT",
                offer_value=14750.0,
                calculated_value=15333.33,
                difference=-583.33,
                note="KDV tutarı kalem toplamlarına göre farklı.",
            ),
            FinancialCheck(
                label="Toplam Yatırım Maliyeti",
                status="ONAY",
                offer_value=92000.0,
                calculated_value=92000.0,
                difference=0.0,
                note="Kalem toplamı ile uyumlu.",
            ),
        ],
    )
    return ComparisonSession(
        token="demo",
        price_list_path=price_list_path,
        offer_path=offer_path,
        output_path=output_path,
        selected_column=selected_column,
        price_mode="kurumsal_nakit",
        results=results,
        financial_review=financial_review,
    )


def batch_item_from_session(session: ComparisonSession) -> BatchComparisonItem:
    item = BatchComparisonItem(
        offer_path=session.offer_path,
        status="completed",
        session_token=session.token,
        output_path=session.output_path,
        selected_column=session.selected_column,
        metrics=result_metrics(session.results),
        financial_status=session.financial_review.overall_status,
    )
    enrich_batch_item(item)
    return item


def batch_item_issue_count(item: BatchComparisonItem) -> int:
    return item.metrics.get("DUZELT", 0) + item.metrics.get("INCELE", 0) + item.metrics.get("ESLESMEDI", 0)


def batch_item_needs_action(item: BatchComparisonItem) -> bool:
    if item.status == "error":
        return True
    return bool(batch_item_issue_count(item) or item.financial_status != "ONAY")


def batch_item_category(item: BatchComparisonItem) -> str:
    if item.status == "error":
        return "error"
    if batch_item_needs_action(item):
        return "action"
    return "clean"


def batch_item_problem_summary(item: BatchComparisonItem) -> str:
    if item.status == "error":
        return item.error or "Bu teklif işlenemedi."
    problems: list[str] = []
    if item.metrics.get("DUZELT", 0):
        problems.append(f"{item.metrics.get('DUZELT', 0)} fiyat farkı")
    if item.metrics.get("INCELE", 0):
        problems.append(f"{item.metrics.get('INCELE', 0)} satır inceleme")
    if item.metrics.get("ESLESMEDI", 0):
        problems.append(f"{item.metrics.get('ESLESMEDI', 0)} eşleşmeyen ürün")
    if item.financial_status != "ONAY":
        problems.append("finansal özet uyarısı")
    return ", ".join(problems) if problems else "İşlem gerekmiyor."


def batch_item_next_step(item: BatchComparisonItem) -> str:
    if item.status == "error":
        return "Dosyayı, PDF tipini veya fiyat listesini kontrol edip yeniden çalıştır."
    if item.metrics.get("ESLESMEDI", 0) or item.metrics.get("INCELE", 0):
        return "İncele ekranında ürün eşleştir ve farkı yeniden değerlendir."
    if item.metrics.get("DUZELT", 0):
        return "Karar Merkezi'nde PDF'e yazılacak satırları onayla."
    if item.financial_status != "ONAY":
        return "Finansal kontrol detayını açıp toplam ve KDV değerlerini kontrol et."
    return "Temiz; raporu arşivle veya ZIP içinde indir."


def enrich_batch_item(item: BatchComparisonItem) -> BatchComparisonItem:
    item.problem_summary = item.problem_summary or batch_item_problem_summary(item)
    item.next_step = item.next_step or batch_item_next_step(item)
    return item


def batch_item_status_label(item: BatchComparisonItem) -> str:
    if item.status == "error":
        return "Hata"
    if batch_item_issue_count(item):
        return "İnceleme gerekli"
    if item.financial_status != "ONAY":
        return "Finansal uyarı"
    return "Temiz"


def batch_item_status_class(item: BatchComparisonItem) -> str:
    if item.status == "error":
        return "danger"
    if item.metrics.get("DUZELT", 0):
        return "danger"
    if item.metrics.get("INCELE", 0) or item.metrics.get("ESLESMEDI", 0) or item.financial_status != "ONAY":
        return "warn"
    return "ok"


def batch_job_storage_path(batch_token: str) -> Path:
    return BATCH_JOBS_DIR / f"{batch_token}.json"


def batch_item_to_payload(item: BatchComparisonItem) -> dict:
    enrich_batch_item(item)
    return {
        "offer_path": relative_runtime_path(item.offer_path),
        "status": item.status,
        "session_token": item.session_token,
        "output_path": relative_runtime_path(item.output_path) if item.output_path else "",
        "selected_column": item.selected_column,
        "metrics": {key: int(value) for key, value in item.metrics.items()},
        "financial_status": item.financial_status,
        "error": item.error,
        "problem_summary": item.problem_summary,
        "next_step": item.next_step,
    }


def batch_item_from_payload(payload: dict) -> BatchComparisonItem:
    output_path_value = str(payload.get("output_path") or "").strip()
    item = BatchComparisonItem(
        offer_path=resolve_runtime_pdf_path(str(payload.get("offer_path") or "")),
        status=str(payload.get("status") or "error"),
        session_token=str(payload.get("session_token") or ""),
        output_path=resolve_activity_file_path(output_path_value) if output_path_value else None,
        selected_column=str(payload.get("selected_column") or ""),
        metrics={str(key): int(value or 0) for key, value in dict(payload.get("metrics") or {}).items()},
        financial_status=str(payload.get("financial_status") or "-"),
        error=str(payload.get("error") or ""),
        problem_summary=str(payload.get("problem_summary") or ""),
        next_step=str(payload.get("next_step") or ""),
    )
    return enrich_batch_item(item)


def batch_job_to_payload(job: BatchComparisonJob) -> dict:
    return {
        "token": job.token,
        "price_list_path": relative_runtime_path(job.price_list_path),
        "price_mode": job.price_mode,
        "created_at": job.created_at.isoformat(),
        "summary_path": relative_runtime_path(job.summary_path),
        "items": [batch_item_to_payload(item) for item in job.items],
    }


def batch_job_from_payload(payload: dict) -> BatchComparisonJob:
    created_at_raw = str(payload.get("created_at") or "")
    try:
        created_at = datetime.fromisoformat(created_at_raw)
    except ValueError:
        created_at = datetime.now()
    summary_path_value = str(payload.get("summary_path") or "")
    job = BatchComparisonJob(
        token=str(payload.get("token") or ""),
        price_list_path=resolve_activity_file_path(str(payload.get("price_list_path") or "")),
        price_mode=normalize_compare_mode(str(payload.get("price_mode") or PRICE_MODE_AUTO)),
        created_at=created_at,
        summary_path=resolve_activity_file_path(summary_path_value) if summary_path_value else build_batch_summary_path(str(payload.get("token") or uuid.uuid4().hex)),
        items=[batch_item_from_payload(item) for item in list(payload.get("items") or [])],
    )
    return job


def save_batch_job(job: BatchComparisonJob) -> None:
    ensure_admin_storage()
    batch_job_storage_path(job.token).write_text(
        json.dumps(batch_job_to_payload(job), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_batch_job(batch_token: str) -> BatchComparisonJob | None:
    token = str(batch_token or "").strip()
    if not token:
        return None
    cached = BATCHES.get(token)
    if cached is not None:
        return cached
    path = batch_job_storage_path(token)
    if not path.exists():
        return None
    try:
        job = batch_job_from_payload(json.loads(path.read_text(encoding="utf-8")))
    except (OSError, json.JSONDecodeError, TypeError, ValueError):
        return None
    if job.token:
        BATCHES[job.token] = job
    return job


def register_batch_job(job: BatchComparisonJob) -> None:
    BATCHES[job.token] = job
    save_batch_job(job)


def write_batch_summary(job: BatchComparisonJob) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TopluKontrol"
    headers = [
        "Teklif",
        "Durum",
        "Referans",
        "Onay",
        "Duzelt",
        "Incele",
        "Eslesmedi",
        "Finansal",
        "Problem",
        "Siradaki Islem",
        "Rapor",
        "Hata",
    ]
    sheet.append(headers)
    for item in job.items:
        sheet.append(
            [
                item.offer_path.name,
                batch_item_status_label(item),
                item.selected_column,
                item.metrics.get("ONAY", 0),
                item.metrics.get("DUZELT", 0),
                item.metrics.get("INCELE", 0),
                item.metrics.get("ESLESMEDI", 0),
                item.financial_status,
                enrich_batch_item(item).problem_summary,
                enrich_batch_item(item).next_step,
                relative_display_path(item.output_path) if item.output_path else "",
                item.error,
            ]
    )
    for cell in sheet[1]:
        font = copy(cell.font)
        font.bold = True
        cell.font = font
    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 42)
    job.summary_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(job.summary_path)


def batch_job_view_model(request: Request, job: BatchComparisonJob | None) -> dict:
    if job is None:
        return {"show": False, "rows": [], "metrics": {}, "token": ""}
    completed = sum(1 for item in job.items if item.status == "completed")
    failed = sum(1 for item in job.items if item.status == "error")
    needs_review = sum(1 for item in job.items if item.status == "completed" and batch_item_needs_action(item))
    clean = sum(1 for item in job.items if item.status == "completed" and not batch_item_needs_action(item))
    price_diff_count = sum(item.metrics.get("DUZELT", 0) for item in job.items)
    unmatched_count = sum(item.metrics.get("ESLESMEDI", 0) + item.metrics.get("INCELE", 0) for item in job.items)
    row_views = []
    for index, item in enumerate(job.items, start=1):
        enrich_batch_item(item)
        row_views.append(
            {
                "index": index,
                "offer_name": item.offer_path.name,
                "status": batch_item_status_label(item),
                "status_class": batch_item_status_class(item),
                "category": batch_item_category(item),
                "needs_action": batch_item_needs_action(item),
                "session_token": item.session_token,
                "selected_column": item.selected_column or "-",
                "metrics": item.metrics,
                "financial_status": item.financial_status,
                "error": item.error,
                "problem_summary": item.problem_summary,
                "next_step": item.next_step,
                "report_name": item.output_path.name if item.output_path else "-",
                "view_url": (
                    request.url_for("offer-tool:view_session", token=item.session_token)
                    if item.session_token and item.session_token in SESSIONS
                    else ""
                ),
                "report_url": (
                    request.url_for("offer-tool:download_batch_item_report", batch_token=job.token, item_index=index)
                    if item.output_path
                    else ""
                ),
            }
        )
    if failed:
        command_title = "Önce hata veren PDF'leri ayıkla"
        command_copy = f"{failed} teklif işlenemedi. Hatalı satırları filtreleyip dosyaları yeniden seç."
    elif needs_review:
        command_title = "Aksiyon gereken teklifleri incele"
        command_copy = f"{needs_review} teklif karar bekliyor. İncele butonuyla tekli karar ekranına geç."
    else:
        command_title = "Toplu kontrol temiz"
        command_copy = "Excel özeti veya ZIP dosyasını indirip arşivleyebilirsin."
    return {
        "show": True,
        "token": job.token,
        "price_file": job.price_list_path.name,
        "price_mode": PRICE_MODE_LABELS.get(job.price_mode, job.price_mode),
        "created_at": job.created_at.strftime("%d.%m.%Y %H:%M"),
        "summary_url": request.url_for("offer-tool:download_batch_summary", batch_token=job.token),
        "zip_url": request.url_for("offer-tool:download_batch_reports_zip", batch_token=job.token),
        "rows": row_views,
        "command": {"title": command_title, "copy": command_copy},
        "metrics": {
            "total": len(job.items),
            "completed": completed,
            "failed": failed,
            "needs_review": needs_review,
            "clean": clean,
            "price_diff": price_diff_count,
            "unmatched": unmatched_count,
        },
    }


def build_feedback_state(
    *,
    notice: str | None,
    error: str | None,
    metrics: dict[str, int],
    has_results: bool,
    financial_overall_status: str,
    has_corrected_pdf: bool,
) -> dict:
    if error:
        return {
            "show": True,
            "tone": "error",
            "title": "Islem tamamlanamadi",
            "message": error,
        }

    if not notice:
        return {"show": False, "tone": "idle", "title": "", "message": ""}

    if has_corrected_pdf:
        return {
            "show": True,
            "tone": "success",
            "title": "Düzeltmeler hazırlandı",
            "message": f"{notice} Yeni PDF duzeltilmis_teklifler klasorune kaydedildi; indirme manuel onayla baslar.",
        }

    if has_results:
        issue_count = metrics.get("DUZELT", 0) + metrics.get("INCELE", 0) + metrics.get("ESLESMEDI", 0)
        if issue_count == 0 and financial_overall_status == "ONAY":
            message = f"{metrics.get('ONAY', 0)} kalem ve finansal ozet uyumlu. Teklif onaya hazir."
            return {"show": True, "tone": "success", "title": "Kontrol olumlu", "message": message}

        details: list[str] = []
        if metrics.get("DUZELT", 0):
            details.append(f"{metrics['DUZELT']} duzeltme")
        if metrics.get("INCELE", 0):
            details.append(f"{metrics['INCELE']} inceleme")
        if metrics.get("ESLESMEDI", 0):
            details.append(f"{metrics['ESLESMEDI']} eslesmeyen")
        if financial_overall_status != "ONAY":
            details.append("finansal kontrolde dikkat")
        detail_text = ", ".join(details) if details else "inceleme gerekiyor"
        return {
            "show": True,
            "tone": "warning",
            "title": "Kontrol tamamlandi",
            "message": f"{detail_text}. Sonuc tablosunu ve raporu kontrol et.",
        }

    return {
        "show": True,
        "tone": "success",
        "title": "Islem tamamlandi",
        "message": notice,
    }


def default_admin_state() -> dict:
    settings = load_admin_settings()
    active_price_file = settings.get("active_price_file") or default_price_file()
    active_template_file = settings.get("active_template_file") or default_template_file()
    return {
        "selected_price_file": active_price_file,
        "selected_template_file": active_template_file,
        "default_compare_mode": settings.get("default_compare_mode") or PRICE_MODE_AUTO,
        "default_create_mode": settings.get("default_create_mode") or "kurumsal_6",
        "default_vat_mode": settings.get("default_vat_mode") or VAT_MODE_INCLUDED,
        "activate_after_import": True,
        "activate_template_after_import": True,
    }


def list_recent_files(directory: Path, pattern: str, *, limit: int = 5) -> list[str]:
    if not directory.exists():
        return []
    files = sorted(
        directory.glob(pattern),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    return [relative_display_path(path) for path in files[:limit]]


def build_workspace_snapshot() -> dict:
    ensure_runtime_folders()
    reports_dir = BASE_DIR / OUTPUT_ROOT_DIRNAME / REPORTS_DIRNAME
    corrected_dir = BASE_DIR / OUTPUT_ROOT_DIRNAME / CORRECTED_PDFS_DIRNAME
    generated_dir = BASE_DIR / OUTPUT_ROOT_DIRNAME / GENERATED_OFFERS_DIRNAME
    workspace_cards = [
        {"label": "Fiyat listesi", "count": len(list_price_files())},
        {"label": "Teklif PDF", "count": len(list_offer_files())},
        {"label": "Sablon", "count": len(list_template_files())},
        {"label": "Rapor", "count": len(list(reports_dir.glob("*.xlsx"))) if reports_dir.exists() else 0},
        {"label": "Duzeltilmis PDF", "count": len(list(corrected_dir.glob("*.pdf"))) if corrected_dir.exists() else 0},
        {"label": "Olusturulan teklif", "count": len(list(generated_dir.glob("*.pdf"))) if generated_dir.exists() else 0},
    ]
    workspace_folders = [
        {"label": "Teklif klasoru", "path": relative_display_path(OFFERS_DIR), "count": len(list_offer_files())},
        {"label": "Sablon klasoru", "path": relative_display_path(TEMPLATES_PDF_DIR), "count": len(list_template_files())},
        {"label": "Fiyat listeleri", "path": relative_display_path(PRICE_LISTS_DIR), "count": len(list_price_files())},
        {"label": "Raporlar", "path": relative_display_path(reports_dir), "count": workspace_cards[3]["count"]},
        {"label": "Duzeltilmisler", "path": relative_display_path(corrected_dir), "count": workspace_cards[4]["count"]},
        {"label": "Olusturulanlar", "path": relative_display_path(generated_dir), "count": workspace_cards[5]["count"]},
    ]
    recent_files = {
        "offers": list_recent_files(OFFERS_DIR, "*.pdf"),
        "templates": list_recent_files(TEMPLATES_PDF_DIR, "*.pdf"),
        "reports": list_recent_files(reports_dir, "*.xlsx"),
        "corrected": list_recent_files(corrected_dir, "*.pdf"),
        "generated": list_recent_files(generated_dir, "*.pdf"),
    }
    return {
        "workspace_cards": workspace_cards,
        "workspace_folders": workspace_folders,
        "recent_files": recent_files,
    }


def build_context(
    request: Request,
    *,
    selected_price_file: str | None = None,
    selected_offer_file: str | None = None,
    selected_mode: str | None = None,
    session: ComparisonSession | None = None,
    create_state: dict | None = None,
    admin_state: dict | None = None,
    batch_job: BatchComparisonJob | None = None,
    notice: str | None = None,
    error: str | None = None,
    active_workspace: str | None = None,
    play_result_sound: bool = False,
) -> dict:
    portal_user = get_offer_portal_user(request)
    offer_is_admin = bool(portal_user and portal_user.is_admin)
    admin_settings = load_admin_settings()
    selected_price_file = selected_price_file or default_price_file()
    selected_offer_file = selected_offer_file or default_offer_file()
    selected_mode = normalize_compare_mode(selected_mode or admin_settings.get("default_compare_mode"))
    create_state = create_state or default_create_state(selected_price_file)
    admin_state = {**default_admin_state(), **(admin_state or {})}

    price_files = [path.name for path in list_price_files()]
    offer_files = [relative_runtime_path(path) for path in list_offer_files()]
    template_files = [relative_runtime_path(path) for path in list_template_files()]
    catalog_options = load_catalog_options(create_state.get("price_file", selected_price_file))
    result_catalog_options = load_catalog_options(session.price_list_path.name if session else selected_price_file)
    workspace_snapshot = build_workspace_snapshot()
    batch_view = batch_job_view_model(request, batch_job)

    metrics = {"ONAY": 0, "DUZELT": 0, "INCELE": 0, "ESLESMEDI": 0}
    results: list[dict] = []
    financial_checks: list[dict] = []
    reference_column = "-"
    report_name = "Excel raporu henüz üretilmedi."
    corrected_name = "Düzeltilmiş PDF henüz oluşturulmadı."
    token = ""
    has_corrected_pdf = False
    vat_rate_label = "-"
    vat_rate_source = ""
    financial_overall_status = "-"
    apply_ready_count = 0
    review_needed_count = 0
    result_effective_price_mode = ""
    activity_entries: list[dict] = []
    ai_action_board = {
        "tone": "info",
        "headline": "Henüz kontrol sonucu yok.",
        "summary": "Teklif kontrolü çalışınca sistem satır satır Türkçe açıklama üretir.",
        "next_focus": "Önce teklif PDF'ini kontrol et.",
        "cards": [],
        "has_actions": False,
        "financial_status": financial_overall_status,
    }

    if session is not None:
        metrics = Counter(result.status for result in session.results)
        try:
            session_price_rows, _ = load_price_rows(session.price_list_path)
        except Exception:
            session_price_rows = []
        results = [
            result_view_model(
                result,
                index,
                price_rows=session_price_rows,
                selected_column=session.selected_column,
                vat_included=session.financial_review.vat_included,
                vat_rate=session.financial_review.vat_rate,
            )
            for index, result in enumerate(session.results)
        ]
        apply_ready_count = sum(1 for row in results if row["can_apply"])
        review_needed_count = metrics.get("INCELE", 0) + metrics.get("ESLESMEDI", 0)
        financial_checks = [financial_check_view_model(check) for check in session.financial_review.checks]
        reference_column = session.selected_column
        report_name = relative_display_path(session.output_path)
        corrected_name = relative_display_path(session.corrected_pdf_path) if session.corrected_pdf_path else corrected_name
        token = session.token
        has_corrected_pdf = session.corrected_pdf_path is not None and session.corrected_pdf_path.exists()
        vat_rate_label = format_rate_label(session.financial_review.vat_rate)
        vat_rate_source = (
            "KDV oran\u0131 PDF i\u00e7inden okundu."
            if session.financial_review.vat_rate_source == "PDF"
            else "KDV oran\u0131 PDF i\u00e7inde bulunamad\u0131, varsay\u0131lan oran kullan\u0131ld\u0131."
        )
        financial_overall_status = session.financial_review.overall_status
        result_effective_price_mode = (
            session.price_mode
            if session.price_mode != PRICE_MODE_AUTO
            else infer_price_mode_from_selected_column(session.selected_column)
        )
        ai_action_board = build_ai_action_board(
            results,
            financial_checks,
            metrics,
            financial_overall_status=financial_overall_status,
        )

    if offer_is_admin:
        activity_entries = [activity_entry_view_model(request, entry) for entry in load_activity_log()]

    feedback = build_feedback_state(
        notice=notice,
        error=error,
        metrics=metrics,
        has_results=bool(results),
        financial_overall_status=financial_overall_status,
        has_corrected_pdf=has_corrected_pdf,
    )

    valid_workspaces = {"compare", "batch", "results", "apply", "create"}
    if offer_is_admin:
        valid_workspaces.add("settings")
    selected_offer_path = resolve_runtime_pdf_path(selected_offer_file) if selected_offer_file else None
    if active_workspace not in valid_workspaces:
        if selected_offer_path is not None and is_template_pdf_path(selected_offer_path):
            active_workspace = "create"
        else:
            active_workspace = "results" if results else "compare"
    if not results and active_workspace in {"results", "apply"}:
        active_workspace = "compare"

    return {
        "request": request,
        "portal_user": portal_user,
        "offer_is_admin": offer_is_admin,
        "price_files": price_files,
        "offer_files": offer_files,
        "template_files": template_files,
        "catalog_options": catalog_options,
        "result_catalog_options": result_catalog_options,
        "create_state": create_state,
        "admin_state": admin_state,
        "selected_price_file": selected_price_file,
        "selected_offer_file": selected_offer_file,
        "selected_mode": selected_mode,
        "price_mode_groups": grouped_price_modes(),
        "price_mode_labels": PRICE_MODE_LABELS,
        "price_mode_payment_info": PRICE_MODE_PAYMENT_INFO,
        "vat_mode_labels": VAT_MODE_LABELS,
        "result_price_mode": session.price_mode if session else "",
        "result_effective_price_mode": result_effective_price_mode,
        "result_selected_column": session.selected_column if session else "",
        "result_vat_included": session.financial_review.vat_included if session else True,
        "metrics": metrics,
        "results": results,
        "financial_checks": financial_checks,
        "reference_column": reference_column,
        "report_name": report_name,
        "corrected_name": corrected_name,
        "apply_ready_count": apply_ready_count,
        "review_needed_count": review_needed_count,
        "vat_rate_label": vat_rate_label,
        "vat_rate_source": vat_rate_source,
        "financial_overall_status": financial_overall_status,
        "ai_action_board": ai_action_board,
        "token": token,
        "notice": notice,
        "error": error,
        "storage_hint": STORAGE_HINT,
        "admin_storage_hint": ADMIN_STORAGE_HINT,
        "active_price_file": admin_settings.get("active_price_file") or default_price_file(),
        "active_template_file": admin_settings.get("active_template_file") or default_template_file(),
        "default_compare_mode": admin_settings.get("default_compare_mode") or PRICE_MODE_AUTO,
        "default_create_mode": admin_settings.get("default_create_mode") or "kurumsal_6",
        "default_vat_mode": admin_settings.get("default_vat_mode") or VAT_MODE_INCLUDED,
        "offer_activity_entries": activity_entries,
        "offer_activity_latest": activity_entries[0] if activity_entries else None,
        "offer_activity_action_labels": ACTIVITY_ACTION_LABELS,
        "batch_result": batch_view,
        "has_batch_results": bool(batch_view.get("show")),
        "workspace_cards": workspace_snapshot["workspace_cards"],
        "workspace_folders": workspace_snapshot["workspace_folders"],
        "recent_files": workspace_snapshot["recent_files"],
        "feedback": feedback,
        "has_results": bool(results),
        "has_corrected_pdf": has_corrected_pdf,
        "active_workspace": active_workspace,
        "play_result_sound": bool(play_result_sound and results and token),
        "result_sound_key": token if play_result_sound and results else "",
    }


@app.get("/", response_class=HTMLResponse)
async def index(request: Request) -> HTMLResponse:
    return templates.TemplateResponse("index.html", build_context(request))


@app.get("/demo", response_class=HTMLResponse)
async def demo(request: Request) -> HTMLResponse:
    session = create_demo_comparison_session()
    context = build_context(
        request,
        selected_price_file=session.price_list_path.name,
        selected_offer_file=session.offer_path.name,
        selected_mode=session.price_mode,
        session=session,
        notice="AI açıklamalı örnek teklif kontrolü hazırlandı.",
        active_workspace="results",
        play_result_sound=True,
    )
    return templates.TemplateResponse("index.html", context)


@app.post("/admin/import-price-list", response_class=HTMLResponse)
async def admin_import_price_list(
    request: Request,
    admin_pin: str = Form(""),
    price_file_upload: UploadFile = File(...),
    activate_after_import: str | None = Form(default="on"),
    active_workspace: str = Form(default="settings"),
) -> HTMLResponse:
    admin_state = {
        "selected_price_file": default_price_file(),
        "activate_after_import": activate_after_import is not None,
    }

    try:
        require_offer_admin(request)
    except PermissionError as exc:
        context = build_context(
            request,
            admin_state=admin_state,
            error=str(exc),
        )
        return templates.TemplateResponse("index.html", context, status_code=403)

    try:
        imported_path, row_count, price_columns = import_price_file(
            price_file_upload,
            activate_after_import=activate_after_import is not None,
        )
    except Exception as exc:
        context = build_context(
            request,
            admin_state=admin_state,
            error=str(exc),
        )
        return templates.TemplateResponse("index.html", context, status_code=400)

    selected_price_file = imported_path.name if activate_after_import is not None else default_price_file()
    admin_state["selected_price_file"] = imported_path.name
    notice = (
        f"Yeni fiyat listesi içe aktarıldı: {imported_path.name}. "
        f"{row_count} ürün ve {len(price_columns)} fiyat kolonu doğrulandı."
    )
    if activate_after_import is not None:
        notice += " Aktif fiyat listesi güncellendi."

    context = build_context(
        request,
        selected_price_file=selected_price_file,
        create_state=default_create_state(selected_price_file),
        admin_state=admin_state,
        notice=notice,
    )
    return templates.TemplateResponse("index.html", context)


@app.post("/admin/set-active-price-file", response_class=HTMLResponse)
async def admin_set_active_price_file(
    request: Request,
    admin_pin: str = Form(""),
    admin_price_file: str = Form(""),
) -> HTMLResponse:
    admin_state = {
        "selected_price_file": admin_price_file or default_price_file(),
        "activate_after_import": True,
    }

    try:
        require_offer_admin(request)
    except PermissionError as exc:
        context = build_context(
            request,
            admin_state=admin_state,
            error=str(exc),
        )
        return templates.TemplateResponse("index.html", context, status_code=403)

    try:
        active_path = activate_price_file(admin_price_file)
    except Exception as exc:
        context = build_context(
            request,
            admin_state=admin_state,
            error=str(exc),
        )
        return templates.TemplateResponse("index.html", context, status_code=400)

    selected_price_file = active_path.name
    notice = f"Aktif fiyat listesi güncellendi: {active_path.name}"
    context = build_context(
        request,
        selected_price_file=selected_price_file,
        create_state=default_create_state(selected_price_file),
        admin_state=admin_state,
        notice=notice,
    )
    return templates.TemplateResponse("index.html", context)


@app.post("/admin/save-settings", response_class=HTMLResponse)
async def admin_save_settings(
    request: Request,
    admin_pin: str = Form(""),
    active_price_file: str = Form(""),
    active_template_file: str = Form(""),
    default_compare_mode: str = Form(PRICE_MODE_AUTO),
    default_create_mode: str = Form("kurumsal_6"),
    default_vat_mode: str = Form(VAT_MODE_INCLUDED),
) -> HTMLResponse:
    admin_state = {
        "selected_price_file": active_price_file or default_price_file(),
        "selected_template_file": active_template_file or default_template_file(),
        "default_compare_mode": normalize_compare_mode(default_compare_mode),
        "default_create_mode": normalize_create_mode(default_create_mode),
        "default_vat_mode": normalize_vat_mode(default_vat_mode),
    }

    try:
        require_offer_admin(request)
    except PermissionError as exc:
        context = build_context(
            request,
            admin_state=admin_state,
            error=str(exc),
        )
        return templates.TemplateResponse("index.html", context, status_code=403)

    try:
        resolved_price_file = active_price_file or default_price_file()
        if resolved_price_file:
            active_price_path = activate_price_file(resolved_price_file)
            admin_state["selected_price_file"] = active_price_path.name
        resolved_template_file = active_template_file or default_template_file()
        if resolved_template_file:
            active_template_path = activate_template_file(resolved_template_file)
            admin_state["selected_template_file"] = relative_runtime_path(active_template_path)
    except Exception as exc:
        context = build_context(
            request,
            admin_state=admin_state,
            error=str(exc),
        )
        return templates.TemplateResponse("index.html", context, status_code=400)

    settings = load_admin_settings()
    settings["default_compare_mode"] = admin_state["default_compare_mode"]
    settings["default_create_mode"] = admin_state["default_create_mode"]
    settings["default_vat_mode"] = admin_state["default_vat_mode"]
    save_admin_settings(settings)

    context = build_context(
        request,
        selected_price_file=admin_state["selected_price_file"],
        admin_state=admin_state,
        create_state=default_create_state(admin_state["selected_price_file"]),
        notice="Sistem ayarlari kaydedildi.",
    )
    return templates.TemplateResponse("index.html", context)


@app.post("/admin/import-template", response_class=HTMLResponse)
async def admin_import_template(
    request: Request,
    admin_pin: str = Form(""),
    template_file_upload: UploadFile = File(...),
    activate_after_import: str | None = Form(default="on"),
    active_workspace: str = Form(default="settings"),
) -> HTMLResponse:
    admin_state = {
        "selected_template_file": default_template_file(),
        "activate_template_after_import": activate_after_import is not None,
    }

    try:
        require_offer_admin(request)
    except PermissionError as exc:
        context = build_context(
            request,
            admin_state=admin_state,
            error=str(exc),
            active_workspace=active_workspace,
        )
        return templates.TemplateResponse("index.html", context, status_code=403)

    try:
        imported_path = import_template_file(
            template_file_upload,
            activate_after_import=activate_after_import is not None,
        )
    except Exception as exc:
        context = build_context(
            request,
            admin_state=admin_state,
            error=str(exc),
            active_workspace=active_workspace,
        )
        return templates.TemplateResponse("index.html", context, status_code=400)

    imported_template_file = relative_runtime_path(imported_path)
    admin_state["selected_template_file"] = imported_template_file
    create_state = default_create_state(default_price_file())
    create_state["template_file"] = imported_template_file
    context = build_context(
        request,
        admin_state=admin_state,
        create_state=create_state,
        notice=(
            f"Yeni şablon eklendi: {imported_path.name}."
            + (" Aktif şablon güncellendi." if activate_after_import is not None else "")
        ),
        active_workspace=active_workspace,
    )
    return templates.TemplateResponse("index.html", context)


@app.post("/compare", response_class=HTMLResponse)
async def compare(
    request: Request,
    price_file: str = Form(""),
    offer_file: str = Form(""),
    price_mode: str = Form(PRICE_MODE_AUTO),
    active_workspace: str = Form(default="compare"),
    price_file_upload: UploadFile | None = File(default=None),
    offer_file_upload: UploadFile | None = File(default=None),
) -> HTMLResponse:
    price_mode = normalize_compare_mode(price_mode)
    try:
        if price_file_upload is not None and price_file_upload.filename:
            try:
                require_offer_admin(request)
            except PermissionError as exc:
                raise ValueError("Fiyat listesi yukleme yetkisi sadece adminde.") from exc
            imported_path, _, _ = import_price_file(price_file_upload, activate_after_import=True)
            price_file = imported_path.name
        if offer_file_upload is not None and offer_file_upload.filename:
            imported_offer_path = import_offer_file(offer_file_upload)
            offer_file = relative_runtime_path(imported_offer_path)
    except Exception as exc:
        context = build_context(
            request,
            selected_price_file=price_file,
            selected_offer_file=offer_file,
            selected_mode=price_mode,
            active_workspace=active_workspace,
        )
        context["error"] = str(exc)
        return templates.TemplateResponse("index.html", context, status_code=400)

    price_list_path = resolve_price_file_path(price_file)
    offer_path = resolve_runtime_pdf_path(offer_file)

    if not price_list_path.exists():
        context = build_context(
            request,
            selected_price_file=price_file,
            selected_offer_file=offer_file,
            selected_mode=price_mode,
            active_workspace=active_workspace,
        )
        context["error"] = "Fiyat listesi bulunamadı."
        return templates.TemplateResponse("index.html", context, status_code=400)
    if not offer_path.exists():
        context = build_context(
            request,
            selected_price_file=price_file,
            selected_offer_file=offer_file,
            selected_mode=price_mode,
            active_workspace=active_workspace,
        )
        context["error"] = "Teklif PDF bulunamadı."
        return templates.TemplateResponse("index.html", context, status_code=400)
    if is_template_pdf_path(offer_path):
        context = build_context(
            request,
            selected_price_file=price_file,
            selected_offer_file="",
            selected_mode=price_mode,
            active_workspace=active_workspace,
        )
        context["error"] = (
            "Seçilen PDF şablon klasöründe duruyor. Kontrol için teklif PDF'i seç veya "
            "Yeni teklif PDF seç alanından dosyayı yükle."
        )
        return templates.TemplateResponse("index.html", context, status_code=400)

    remember_last_files(price_file_name=price_file, offer_file_name=relative_runtime_path(offer_path))
    try:
        session = create_comparison_session(price_list_path, offer_path, price_mode)
    except Exception as exc:
        context = build_context(
            request,
            selected_price_file=price_file,
            selected_offer_file=offer_file,
            selected_mode=price_mode,
            active_workspace=active_workspace,
        )
        context["error"] = str(exc)
        return templates.TemplateResponse("index.html", context, status_code=400)

    append_activity_log(
        request,
        action="compare",
        summary=f"{offer_path.name} kontrol edildi; referans fiyat tipi {session.selected_column}.",
        files=[
            activity_file_payload(session.output_path, "Excel raporu", "report"),
            activity_file_payload(offer_path, "Kontrol edilen teklif", "source"),
        ],
        details={
            "price_file": price_list_path.name,
            "offer_file": offer_path.name,
            "reference_column": session.selected_column,
            "row_count": len(session.results),
        },
    )

    context = build_context(
        request,
        selected_price_file=price_file,
        selected_offer_file=offer_file,
        selected_mode=price_mode,
        session=session,
        notice=f"Kontrol tamamlandı. Referans fiyat tipi: {session.selected_column}",
        active_workspace="results",
        play_result_sound=True,
    )
    return templates.TemplateResponse("index.html", context)


@app.get("/session/{token}", response_class=HTMLResponse)
async def view_session(request: Request, token: str) -> HTMLResponse:
    session = SESSIONS.get(token)
    if session is None:
        raise HTTPException(status_code=404, detail="Oturum bulunamadı.")
    context = build_context(
        request,
        selected_price_file=session.price_list_path.name,
        selected_offer_file=relative_runtime_path(session.offer_path),
        selected_mode=session.price_mode,
        session=session,
        active_workspace="results",
    )
    return templates.TemplateResponse("index.html", context)


@app.post("/batch-compare", response_class=HTMLResponse)
async def batch_compare(
    request: Request,
    price_file: str = Form(""),
    price_mode: str = Form(PRICE_MODE_AUTO),
    active_workspace: str = Form(default="batch"),
    offer_files: list[str] | None = Form(default=None),
    price_file_upload: UploadFile | None = File(default=None),
    offer_file_uploads: list[UploadFile] | None = File(default=None),
) -> HTMLResponse:
    price_mode = normalize_compare_mode(price_mode)
    try:
        if price_file_upload is not None and price_file_upload.filename:
            try:
                require_offer_admin(request)
            except PermissionError as exc:
                raise ValueError("Fiyat listesi yukleme yetkisi sadece adminde.") from exc
            imported_path, _, _ = import_price_file(price_file_upload, activate_after_import=True)
            price_file = imported_path.name
    except Exception as exc:
        context = build_context(request, selected_price_file=price_file, selected_mode=price_mode, active_workspace=active_workspace)
        context["error"] = str(exc)
        return templates.TemplateResponse("index.html", context, status_code=400)

    price_list_path = resolve_price_file_path(price_file)
    if not price_list_path.exists():
        context = build_context(request, selected_price_file=price_file, selected_mode=price_mode, active_workspace=active_workspace)
        context["error"] = "Fiyat listesi bulunamadı."
        return templates.TemplateResponse("index.html", context, status_code=400)

    offer_paths: list[Path] = []
    seen_offer_paths: set[str] = set()

    def add_offer_path(path: Path) -> None:
        key = str(path.resolve()) if path.exists() else str(path)
        if key in seen_offer_paths:
            return
        seen_offer_paths.add(key)
        offer_paths.append(path)

    for file_name in offer_files or []:
        if str(file_name or "").strip():
            add_offer_path(resolve_runtime_pdf_path(file_name))

    try:
        for upload in offer_file_uploads or []:
            if upload is not None and upload.filename:
                add_offer_path(import_offer_file(upload))
    except Exception as exc:
        context = build_context(
            request,
            selected_price_file=price_file,
            selected_mode=price_mode,
            active_workspace=active_workspace,
        )
        context["error"] = str(exc)
        return templates.TemplateResponse("index.html", context, status_code=400)

    if not offer_paths:
        context = build_context(request, selected_price_file=price_file, selected_mode=price_mode, active_workspace=active_workspace)
        context["error"] = "Toplu kontrol için en az bir teklif PDF seç veya yükle."
        return templates.TemplateResponse("index.html", context, status_code=400)

    if len(offer_paths) > BATCH_COMPARE_LIMIT:
        context = build_context(request, selected_price_file=price_file, selected_mode=price_mode, active_workspace=active_workspace)
        context["error"] = f"Toplu kontrolde tek seferde en fazla {BATCH_COMPARE_LIMIT} PDF işlenebilir."
        return templates.TemplateResponse("index.html", context, status_code=400)

    batch_token = uuid.uuid4().hex
    job = BatchComparisonJob(
        token=batch_token,
        price_list_path=price_list_path,
        price_mode=price_mode,
        created_at=datetime.now(),
        summary_path=build_batch_summary_path(batch_token),
    )

    for offer_path in offer_paths:
        if not offer_path.exists():
            job.items.append(
                BatchComparisonItem(
                    offer_path=offer_path,
                    status="error",
                    error="Teklif PDF bulunamadı.",
                )
            )
            enrich_batch_item(job.items[-1])
            continue
        if is_template_pdf_path(offer_path):
            job.items.append(
                BatchComparisonItem(
                    offer_path=offer_path,
                    status="error",
                    error="Şablon PDF kontrol listesine alınamaz.",
                )
            )
            enrich_batch_item(job.items[-1])
            continue
        try:
            session = create_comparison_session(price_list_path, offer_path, price_mode)
            job.items.append(batch_item_from_session(session))
        except Exception as exc:
            job.items.append(
                BatchComparisonItem(
                    offer_path=offer_path,
                    status="error",
                    error=str(exc),
                )
            )
            enrich_batch_item(job.items[-1])

    write_batch_summary(job)
    register_batch_job(job)

    first_completed = next((item for item in job.items if item.status == "completed"), None)
    remember_last_files(
        price_file_name=price_file,
        offer_file_name=relative_runtime_path(first_completed.offer_path) if first_completed else None,
    )
    completed_count = sum(1 for item in job.items if item.status == "completed")
    failed_count = sum(1 for item in job.items if item.status == "error")
    review_count = sum(
        1
        for item in job.items
        if item.status == "completed"
        and (
            item.metrics.get("DUZELT", 0)
            or item.metrics.get("INCELE", 0)
            or item.metrics.get("ESLESMEDI", 0)
            or item.financial_status != "ONAY"
        )
    )
    append_activity_log(
        request,
        action="batch_compare",
        summary=f"{len(job.items)} teklif toplu kontrol edildi; {completed_count} tamam, {failed_count} hata.",
        files=[
            activity_file_payload(job.summary_path, "Toplu özet Excel", "batch_summary"),
            activity_file_payload(price_list_path, "Fiyat listesi", "price"),
        ],
        details={
            "price_file": price_list_path.name,
            "price_mode": PRICE_MODE_LABELS.get(price_mode, price_mode),
            "total_count": len(job.items),
            "completed_count": completed_count,
            "failed_count": failed_count,
            "review_count": review_count,
        },
    )

    context = build_context(
        request,
        selected_price_file=price_file,
        selected_mode=price_mode,
        batch_job=job,
        notice=f"Toplu kontrol tamamlandı. {completed_count} teklif işlendi, {failed_count} hata var.",
        active_workspace="batch",
    )
    return templates.TemplateResponse("index.html", context)


@app.post("/apply", response_class=HTMLResponse)
async def apply_corrections(
    request: Request,
    token: str = Form(...),
    selected_indexes: list[int] | None = Form(default=None),
    manual_match_row_ids: list[str] = Form(default=[]),
    bundle_match_overrides: list[str] = Form(default=[]),
    action: str = Form(default="apply"),
    active_workspace: str = Form(default="apply"),
) -> HTMLResponse:
    session = SESSIONS.get(token)
    if session is None:
        raise HTTPException(status_code=404, detail="Oturum bulunamadı. Lütfen tekrar kontrol yap.")

    action = str(action or "apply").strip().lower()

    try:
        manual_override_count = apply_manual_match_overrides(session, manual_match_row_ids)
        manual_override_count += apply_bundle_match_overrides(session, bundle_match_overrides)
    except Exception as exc:
        context = build_context(
            request,
            selected_price_file=session.price_list_path.name,
            selected_offer_file=relative_runtime_path(session.offer_path),
            selected_mode=session.price_mode,
            session=session,
            error=str(exc),
            active_workspace=active_workspace,
        )
        return templates.TemplateResponse("index.html", context, status_code=400)

    if action == "preview":
        if manual_override_count == 0:
            context = build_context(
                request,
                selected_price_file=session.price_list_path.name,
                selected_offer_file=relative_runtime_path(session.offer_path),
                selected_mode=session.price_mode,
                session=session,
                error="Yeniden karsilastirmak icin en az bir satirda elle urun sec.",
                active_workspace=active_workspace,
            )
            return templates.TemplateResponse("index.html", context, status_code=400)

        session.corrected_pdf_path = None
        session.applied_indexes.clear()
        context = build_context(
            request,
            selected_price_file=session.price_list_path.name,
            selected_offer_file=relative_runtime_path(session.offer_path),
            selected_mode=session.price_mode,
            session=session,
            notice=f"{manual_override_count} satirda elle urun secimiyle sonuc yeniden hesaplandi.",
            active_workspace="results",
        )
        return templates.TemplateResponse("index.html", context)

    if action != "apply":
        context = build_context(
            request,
            selected_price_file=session.price_list_path.name,
            selected_offer_file=relative_runtime_path(session.offer_path),
            selected_mode=session.price_mode,
            session=session,
            error="Gecersiz islem tipi secildi.",
            active_workspace=active_workspace,
        )
        return templates.TemplateResponse("index.html", context, status_code=400)

    if not selected_indexes:
        context = build_context(
            request,
            selected_price_file=session.price_list_path.name,
            selected_offer_file=relative_runtime_path(session.offer_path),
            selected_mode=session.price_mode,
            session=session,
            error="PDF hazırlamak için en az bir satır seç.",
            active_workspace=active_workspace,
        )
        return templates.TemplateResponse("index.html", context, status_code=400)

    actionable_indexes = [
        index
        for index in selected_indexes
        if 0 <= index < len(session.results)
        and session.results[index].status == "DUZELT"
        and session.results[index].suggested_unit_price is not None
    ]
    if not actionable_indexes:
        context = build_context(
            request,
            selected_price_file=session.price_list_path.name,
            selected_offer_file=relative_runtime_path(session.offer_path),
            selected_mode=session.price_mode,
            session=session,
            error=(
                "Seçilen satırlar için hazırlanacak fiyat farkı kalmadı. "
                "Elle ürün seçtiysen sonucu kontrol et; sadece DÜZELT olan satırlar yeni PDF'e yazılabilir."
            ),
            active_workspace=active_workspace,
        )
        return templates.TemplateResponse("index.html", context, status_code=400)

    try:
        corrected_path = apply_approved_corrections_to_pdf(
            offer_path=session.offer_path,
            results=session.results,
            approved_indexes=actionable_indexes,
            output_path=build_app_corrected_pdf_path(session.offer_path),
        )
    except Exception as exc:
        context = build_context(
            request,
            selected_price_file=session.price_list_path.name,
            selected_offer_file=relative_runtime_path(session.offer_path),
            selected_mode=session.price_mode,
            session=session,
            error=str(exc),
            active_workspace=active_workspace,
        )
        return templates.TemplateResponse("index.html", context, status_code=400)

    session.corrected_pdf_path = corrected_path
    append_activity_log(
        request,
        action="correct",
        summary=f"{session.offer_path.name} için {len(actionable_indexes)} satırla düzeltilmiş PDF hazırlandı.",
        files=[
            activity_file_payload(corrected_path, "Düzenlenmiş PDF", "corrected"),
            activity_file_payload(session.output_path, "Excel raporu", "report"),
            activity_file_payload(session.offer_path, "Kaynak teklif", "source"),
        ],
        details={
            "selected_count": len(actionable_indexes),
            "manual_override_count": manual_override_count,
            "price_file": session.price_list_path.name,
            "reference_column": session.selected_column,
        },
    )
    context = build_context(
        request,
        selected_price_file=session.price_list_path.name,
        selected_offer_file=relative_runtime_path(session.offer_path),
        selected_mode=session.price_mode,
        session=session,
        notice=(
            f"{manual_override_count} satırda elle ürün seçimi kullanıldı. " if manual_override_count else ""
        )
        + f"Onaylı düzeltmeler hazırlandı. Dosyayı kontrol edip manuel indirebilirsin: {corrected_path.name}",
        active_workspace="apply",
    )
    return templates.TemplateResponse("index.html", context)


@app.post("/create-offer")
async def create_offer(
    request: Request,
    template_file: str = Form(...),
    price_file: str = Form(...),
    price_mode: str = Form(...),
    vat_mode: str = Form(VAT_MODE_INCLUDED),
    offer_number: str = Form(...),
    offer_date: str = Form(...),
    valid_until: str = Form(...),
    company_name: str = Form(""),
    contact_name: str = Form(""),
    email: str = Form(""),
    gsm: str = Form(""),
    note: str = Form(""),
    compare_token: str | None = Form(default=None),
    product_row_ids: list[str] = Form(default=[]),
    quantities: list[str] = Form(default=[]),
    manual_prices: list[str] = Form(default=[]),
    discount_types: list[str] = Form(default=[]),
    discount_values: list[str] = Form(default=[]),
):
    price_mode = normalize_create_mode(price_mode)
    vat_mode = normalize_vat_mode(vat_mode)
    session = SESSIONS.get(compare_token or "") if compare_token else None
    item_count = max(
        len(product_row_ids),
        len(quantities),
        len(manual_prices),
        len(discount_types),
        len(discount_values),
        CREATE_FORM_MIN_ROWS,
    )
    create_state = {
        "template_file": template_file,
        "price_file": price_file,
        "price_mode": price_mode,
        "vat_mode": vat_mode,
        "offer_number": offer_number,
        "offer_date": offer_date,
        "valid_until": valid_until,
        "company_name": company_name,
        "contact_name": contact_name,
        "email": email,
        "gsm": gsm,
        "note": note,
        "items": [
            {
                "product_row_id": product_row_ids[index] if index < len(product_row_ids) else "",
                "quantity": quantities[index] if index < len(quantities) else "1",
                "manual_price": manual_prices[index] if index < len(manual_prices) else "",
                "discount_type": discount_types[index] if index < len(discount_types) else DISCOUNT_TYPE_NONE,
                "discount_value": discount_values[index] if index < len(discount_values) else "",
            }
            for index in range(item_count)
        ],
    }
    portal_user = get_offer_portal_user(request)
    offer_is_admin = bool(portal_user and portal_user.is_admin)
    if not offer_is_admin:
        has_manual_price = any(str(value or "").strip() for value in manual_prices)
        has_discount = any(str(value or DISCOUNT_TYPE_NONE).strip().lower() != DISCOUNT_TYPE_NONE for value in discount_types)
        has_discount_value = any(str(value or "").strip() for value in discount_values)
        if has_manual_price or has_discount or has_discount_value:
            context = build_context(
                request,
                selected_price_file=price_file,
                selected_offer_file=relative_runtime_path(session.offer_path) if session else template_file,
                selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
                session=session,
                create_state=create_state,
                error="Manuel fiyat ve iskonto sadece admin tarafindan girilebilir.",
                active_workspace="create",
            )
            return templates.TemplateResponse("index.html", context, status_code=403)

    template_path = resolve_runtime_pdf_path(template_file)
    price_list_path = resolve_price_file_path(price_file)
    selected_column = PRICE_MODE_TO_HEADER.get(price_mode)
    payment_info = PRICE_MODE_PAYMENT_INFO.get(price_mode)
    price_label = PRICE_MODE_LABELS.get(price_mode, selected_column or price_mode)
    vat_included = vat_mode != VAT_MODE_EXCLUDED

    if not template_path.exists():
        context = build_context(
            request,
            selected_price_file=session.price_list_path.name if session else price_file,
            selected_offer_file=relative_runtime_path(session.offer_path) if session else template_file,
            selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
            session=session,
            create_state=create_state,
            error="Şablon PDF bulunamadı.",
        )
        return templates.TemplateResponse("index.html", context, status_code=400)
    if not is_template_pdf_path(template_path):
        context = build_context(
            request,
            selected_price_file=session.price_list_path.name if session else price_file,
            selected_offer_file=relative_runtime_path(session.offer_path) if session else template_file,
            selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
            session=session,
            create_state=create_state,
            error="Teklif olusturmak icin sablonlar klasorundeki sabit bir PDF sec.",
        )
        return templates.TemplateResponse("index.html", context, status_code=400)
    if not price_list_path.exists():
        context = build_context(
            request,
            selected_price_file=price_file,
            selected_offer_file=relative_runtime_path(session.offer_path) if session else template_file,
            selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
            session=session,
            create_state=create_state,
            error="Fiyat listesi bulunamadı.",
        )
        return templates.TemplateResponse("index.html", context, status_code=400)
    if selected_column is None:
        context = build_context(
            request,
            selected_price_file=price_file,
            selected_offer_file=relative_runtime_path(session.offer_path) if session else template_file,
            selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
            session=session,
            create_state=create_state,
            error="Teklif oluşturmak için geçerli bir fiyat tipi seç.",
        )
        return templates.TemplateResponse("index.html", context, status_code=400)
    if vat_mode not in VAT_MODE_LABELS:
        context = build_context(
            request,
            selected_price_file=price_file,
            selected_offer_file=relative_runtime_path(session.offer_path) if session else template_file,
            selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
            session=session,
            create_state=create_state,
            error="Teklif olusturmak icin gecerli bir KDV modu sec.",
        )
        return templates.TemplateResponse("index.html", context, status_code=400)

    try:
        parsed_offer_date = date.fromisoformat(offer_date)
        parsed_valid_until = date.fromisoformat(valid_until)
    except ValueError:
        context = build_context(
            request,
            selected_price_file=price_file,
            selected_offer_file=relative_runtime_path(session.offer_path) if session else template_file,
            selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
            session=session,
            create_state=create_state,
            error="Teklif tarihi veya geçerlilik tarihi geçerli değil.",
        )
        return templates.TemplateResponse("index.html", context, status_code=400)

    selected_entries: list[OfferSelection] = []
    for index in range(item_count):
        row_id = product_row_ids[index] if index < len(product_row_ids) else ""
        quantity_text = quantities[index] if index < len(quantities) else ""
        manual_price_text = manual_prices[index] if index < len(manual_prices) else ""
        discount_type_text = discount_types[index] if index < len(discount_types) else DISCOUNT_TYPE_NONE
        discount_value_text = discount_values[index] if index < len(discount_values) else ""
        if not str(row_id).strip():
            continue
        quantity = parse_money(quantity_text)
        if quantity is None or quantity <= 0:
            context = build_context(
                request,
                selected_price_file=price_file,
                selected_offer_file=relative_runtime_path(session.offer_path) if session else template_file,
                selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
                session=session,
                create_state=create_state,
                error="Seçilen her ürün için sıfırdan büyük adet gir.",
            )
            return templates.TemplateResponse("index.html", context, status_code=400)

        discount_type = str(discount_type_text or DISCOUNT_TYPE_NONE).strip().lower()
        if discount_type not in {DISCOUNT_TYPE_NONE, DISCOUNT_TYPE_AMOUNT, DISCOUNT_TYPE_PERCENT}:
            context = build_context(
                request,
                selected_price_file=price_file,
                selected_offer_file=relative_runtime_path(session.offer_path) if session else template_file,
                selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
                session=session,
                create_state=create_state,
                error="İskonto tipi geçersiz.",
            )
            return templates.TemplateResponse("index.html", context, status_code=400)

        discount_value = None
        if discount_type != DISCOUNT_TYPE_NONE:
            discount_value = parse_money(discount_value_text)
            if discount_value is None or discount_value < 0:
                context = build_context(
                    request,
                    selected_price_file=price_file,
                    selected_offer_file=relative_runtime_path(session.offer_path) if session else template_file,
                    selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
                    session=session,
                    create_state=create_state,
                    error="İskonto değeri geçersiz.",
                )
                return templates.TemplateResponse("index.html", context, status_code=400)
            if discount_type == DISCOUNT_TYPE_PERCENT and discount_value > 100:
                context = build_context(
                    request,
                    selected_price_file=price_file,
                    selected_offer_file=relative_runtime_path(session.offer_path) if session else template_file,
                    selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
                    session=session,
                    create_state=create_state,
                    error="İskonto yüzdesi 100'ü aşamaz.",
                )
                return templates.TemplateResponse("index.html", context, status_code=400)

        manual_price = None
        if discount_type == DISCOUNT_TYPE_NONE and str(manual_price_text).strip():
            manual_price = parse_money(manual_price_text)
            if manual_price is None or manual_price <= 0:
                context = build_context(
                    request,
                    selected_price_file=price_file,
                    selected_offer_file=relative_runtime_path(session.offer_path) if session else template_file,
                    selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
                    session=session,
                    create_state=create_state,
                    error="Elle girilen birim fiyat varsa sifirdan buyuk olmali.",
                )
                return templates.TemplateResponse("index.html", context, status_code=400)
        try:
            selected_entries.append(
                OfferSelection(
                    row_number=int(row_id),
                    quantity=quantity,
                    manual_price=manual_price,
                    discount_type=discount_type,
                    discount_value=discount_value,
                )
            )
        except ValueError:
            context = build_context(
                request,
                selected_price_file=price_file,
                selected_offer_file=relative_runtime_path(session.offer_path) if session else template_file,
                selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
                session=session,
                create_state=create_state,
                error="Urun seciminde gecersiz bir satir bulundu.",
            )
            return templates.TemplateResponse("index.html", context, status_code=400)

    if not selected_entries:
        context = build_context(
            request,
            selected_price_file=price_file,
            selected_offer_file=relative_runtime_path(session.offer_path) if session else template_file,
            selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
            session=session,
            create_state=create_state,
            error="Teklif oluşturmak için en az bir ürün seç.",
        )
        return templates.TemplateResponse("index.html", context, status_code=400)

    if not company_name.strip() and not contact_name.strip():
        context = build_context(
            request,
            selected_price_file=price_file,
            selected_offer_file=relative_runtime_path(session.offer_path) if session else template_file,
            selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
            session=session,
            create_state=create_state,
            error="Firma/Bireysel veya Yetkili Adı alanlarından en az biri dolu olmalı.",
        )
        return templates.TemplateResponse("index.html", context, status_code=400)

    requested_offer_number = offer_number.strip()
    if not requested_offer_number:
        resolved_offer_number = default_offer_number(parsed_offer_date, BASE_DIR)
        create_state["offer_number"] = resolved_offer_number
    elif offer_number_exists(BASE_DIR, requested_offer_number):
        if AUTO_OFFER_NUMBER_PATTERN.fullmatch(requested_offer_number):
            resolved_offer_number = default_offer_number(parsed_offer_date, BASE_DIR)
            create_state["offer_number"] = resolved_offer_number
        else:
            context = build_context(
                request,
                selected_price_file=price_file,
                selected_offer_file=relative_runtime_path(session.offer_path) if session else template_file,
                selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
                session=session,
                create_state=create_state,
                error="Bu teklif numarası daha önce kullanılmış. Lütfen farklı bir teklif numarası gir.",
            )
            return templates.TemplateResponse("index.html", context, status_code=400)
    else:
        resolved_offer_number = requested_offer_number

    try:
        generated_offer_path = create_offer_from_catalog(
            template_path=template_path,
            price_list_path=price_list_path,
            selected_column=selected_column,
            selected_entries=selected_entries,
            vat_included=vat_included,
            vat_rate=DEFAULT_VAT_RATE,
            offer_number=resolved_offer_number,
            offer_date=parsed_offer_date,
            valid_until=parsed_valid_until,
            company_name=company_name.strip(),
            contact_name=contact_name.strip(),
            email=email.strip(),
            gsm=gsm.strip(),
            note_text=note.strip(),
            payment_info=payment_info,
            price_label=price_label,
        )
    except Exception as exc:
        context = build_context(
            request,
            selected_price_file=price_file,
            selected_offer_file=relative_runtime_path(session.offer_path) if session else template_file,
            selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
            session=session,
            create_state=create_state,
            error=str(exc),
        )
        return templates.TemplateResponse("index.html", context, status_code=400)

    settings = load_admin_settings()
    settings["active_template_file"] = relative_runtime_path(template_path)
    save_admin_settings(settings)
    activity_entry = append_activity_log(
        request,
        action="create",
        summary=f"{generated_offer_path.name} oluşturuldu; müşteri {company_name.strip() or contact_name.strip() or '-'}",
        files=[
            activity_file_payload(generated_offer_path, "Oluşturulan teklif", "generated"),
            activity_file_payload(template_path, "Kullanılan şablon", "template"),
            activity_file_payload(price_list_path, "Fiyat listesi", "price"),
        ],
        details={
            "offer_number": resolved_offer_number,
            "company_name": company_name.strip(),
            "contact_name": contact_name.strip(),
            "item_count": len(selected_entries),
            "price_mode": price_label,
            "vat_mode": VAT_MODE_LABELS.get(vat_mode, vat_mode),
        },
    )

    register_pending_offer_approval(
        activity_entry,
        generated_offer_path,
        {
            "offer_number": resolved_offer_number,
            "company_name": company_name.strip(),
            "contact_name": contact_name.strip(),
            "customer_email": email.strip(),
            "item_count": len(selected_entries),
            "price_mode": price_label,
            "vat_mode": VAT_MODE_LABELS.get(vat_mode, vat_mode),
        },
    )

    context = build_context(
        request,
        selected_price_file=price_file,
        selected_offer_file=template_file,
        selected_mode=session.price_mode if session else PRICE_MODE_AUTO,
        session=session,
        create_state=default_create_state(price_file),
        notice="Teklif admin onayina gonderildi. Admin onaylamadan indirilemez veya mail aktarimina acilmaz.",
        active_workspace="create",
    )
    return templates.TemplateResponse("index.html", context)


@app.get("/admin/activity-file/{entry_id}/{file_index}")
async def download_activity_file(request: Request, entry_id: str, file_index: int) -> FileResponse:
    try:
        require_offer_admin(request)
    except PermissionError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc

    entry = next((item for item in load_activity_log() if str(item.get("id") or "") == entry_id), None)
    if entry is None:
        raise HTTPException(status_code=404, detail="Log kaydı bulunamadı.")
    files = entry.get("files") if isinstance(entry.get("files"), list) else []
    if file_index < 0 or file_index >= len(files) or not isinstance(files[file_index], dict):
        raise HTTPException(status_code=404, detail="Log dosyası bulunamadı.")
    approval = pending_offer_for_activity(str(entry.get("id") or ""))
    if (
        str(entry.get("action") or "") == "create"
        and str(files[file_index].get("kind") or "") == "generated"
        and str((approval or {}).get("status") or "") != "approved"
    ):
        raise HTTPException(status_code=403, detail="Bu teklif admin onayi olmadan indirilemez.")
    path = resolve_activity_file_path(str(files[file_index].get("path") or ""))
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Dosya artık bulunamıyor.")
    return FileResponse(path, filename=path.name)


@app.get("/admin/activity-log.csv")
async def download_activity_log_csv(request: Request) -> Response:
    try:
        require_offer_admin(request)
    except PermissionError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Tarih", "Kullanıcı", "E-posta", "İşlem", "Özet", "Dosyalar"])
    for entry in load_activity_log():
        files = entry.get("files") if isinstance(entry.get("files"), list) else []
        writer.writerow(
            [
                str(entry.get("created_at") or "").replace("T", " ")[:19],
                str(entry.get("actor_name") or entry.get("actor_email") or ""),
                str(entry.get("actor_email") or ""),
                str(entry.get("action_label") or ACTIVITY_ACTION_LABELS.get(str(entry.get("action") or ""), "")),
                str(entry.get("summary") or ""),
                ", ".join(str(file_info.get("name") or file_info.get("path") or "") for file_info in files if isinstance(file_info, dict)),
            ]
        )
    csv_bytes = output.getvalue().encode("utf-8-sig")
    return Response(
        csv_bytes,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="teklif-islem-logu.csv"'},
    )


@app.post("/admin/offers/{approval_id}/approve", response_class=HTMLResponse)
async def approve_generated_offer(request: Request, approval_id: str) -> HTMLResponse:
    try:
        admin = require_offer_admin(request)
    except PermissionError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc

    approvals = load_pending_offer_approvals()
    approval = approvals.get(approval_id)
    if approval is None:
        raise HTTPException(status_code=404, detail="Onay bekleyen teklif bulunamadi.")
    if str(approval.get("status") or "") == "rejected":
        raise HTTPException(status_code=400, detail="Reddedilen teklif tekrar onaylanamaz.")

    offer_path = resolve_activity_file_path(str(approval.get("generated_path") or ""))
    if not offer_path.exists() or not offer_path.is_file():
        raise HTTPException(status_code=404, detail="Teklif PDF dosyasi bulunamadi.")

    approval["status"] = "approved"
    approval["approved_at"] = datetime.now().isoformat(timespec="seconds")
    approval["approved_by"] = admin.email
    approvals[approval_id] = approval
    save_pending_offer_approvals(approvals)
    update_activity_entry_details(
        str(approval.get("activity_entry_id") or ""),
        {
            "approval_status": "approved",
            "approved_by": admin.email,
            "approved_at": approval["approved_at"],
        },
    )
    append_activity_log(
        request,
        action="approve",
        summary=f"{approval.get('generated_name') or offer_path.name} admin tarafindan onaylandi.",
        files=[activity_file_payload(offer_path, "Onayli teklif", "generated")],
        details={
            "approval_id": approval_id,
            "offer_number": approval.get("offer_number") or "",
            "company_name": approval.get("company_name") or "",
            "customer_email": approval.get("customer_email") or "",
        },
    )
    context = build_context(
        request,
        notice="Teklif onaylandi; artik indirilebilir ve mail aktarimi icin hazir.",
        active_workspace="create",
    )
    return templates.TemplateResponse("index.html", context)


@app.post("/admin/offers/{approval_id}/reject", response_class=HTMLResponse)
async def reject_generated_offer(request: Request, approval_id: str) -> HTMLResponse:
    try:
        admin = require_offer_admin(request)
    except PermissionError as exc:
        raise HTTPException(status_code=403, detail=str(exc)) from exc

    approvals = load_pending_offer_approvals()
    approval = approvals.get(approval_id)
    if approval is None:
        raise HTTPException(status_code=404, detail="Onay bekleyen teklif bulunamadi.")
    if str(approval.get("status") or "") == "approved":
        raise HTTPException(status_code=400, detail="Onaylanmis teklif reddedilemez.")

    offer_path = resolve_activity_file_path(str(approval.get("generated_path") or ""))
    if offer_path.exists() and offer_path.is_file():
        try:
            offer_path.unlink()
        except OSError:
            pass

    approval["status"] = "rejected"
    approval["rejected_at"] = datetime.now().isoformat(timespec="seconds")
    approval["rejected_by"] = admin.email
    approvals[approval_id] = approval
    save_pending_offer_approvals(approvals)
    update_activity_entry_details(
        str(approval.get("activity_entry_id") or ""),
        {
            "approval_status": "rejected",
            "rejected_by": admin.email,
            "rejected_at": approval["rejected_at"],
        },
    )
    append_activity_log(
        request,
        action="reject",
        summary=f"{approval.get('generated_name') or 'Teklif'} admin tarafindan reddedildi; final PDF kapatildi.",
        details={
            "approval_id": approval_id,
            "offer_number": approval.get("offer_number") or "",
            "company_name": approval.get("company_name") or "",
        },
    )
    context = build_context(
        request,
        notice="Teklif reddedildi; indirilebilir final teklif olusturulmadi.",
        active_workspace="create",
    )
    return templates.TemplateResponse("index.html", context)


@app.get("/batch/{batch_token}", response_class=HTMLResponse)
async def view_batch(request: Request, batch_token: str) -> HTMLResponse:
    job = load_batch_job(batch_token)
    if job is None:
        raise HTTPException(status_code=404, detail="Toplu kontrol bulunamadı.")
    return templates.TemplateResponse(
        "index.html",
        build_context(
            request,
            selected_price_file=job.price_list_path.name,
            selected_mode=job.price_mode,
            batch_job=job,
            active_workspace="batch",
        ),
    )


@app.get("/batch/{batch_token}/summary.xlsx")
async def download_batch_summary(batch_token: str) -> FileResponse:
    job = load_batch_job(batch_token)
    if job is None:
        raise HTTPException(status_code=404, detail="Toplu kontrol bulunamadı.")
    if not job.summary_path.exists():
        raise HTTPException(status_code=404, detail="Toplu kontrol özeti bulunamadı.")
    return FileResponse(job.summary_path, filename=job.summary_path.name)


@app.get("/batch/{batch_token}/reports.zip")
async def download_batch_reports_zip(batch_token: str) -> Response:
    job = load_batch_job(batch_token)
    if job is None:
        raise HTTPException(status_code=404, detail="Toplu kontrol bulunamadı.")

    buffer = io.BytesIO()
    with ZipFile(buffer, "w", ZIP_DEFLATED) as archive:
        if job.summary_path.exists():
            archive.write(job.summary_path, arcname=job.summary_path.name)
        for item in job.items:
            if item.output_path and item.output_path.exists():
                archive.write(item.output_path, arcname=f"raporlar/{item.output_path.name}")
    buffer.seek(0)
    return Response(
        buffer.getvalue(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="toplu-kontrol-{batch_token[:8]}.zip"'},
    )


@app.get("/batch/{batch_token}/report/{item_index}")
async def download_batch_item_report(batch_token: str, item_index: int) -> FileResponse:
    job = load_batch_job(batch_token)
    if job is None:
        raise HTTPException(status_code=404, detail="Toplu kontrol bulunamadı.")
    if item_index < 1 or item_index > len(job.items):
        raise HTTPException(status_code=404, detail="Toplu kontrol satırı bulunamadı.")
    item = job.items[item_index - 1]
    if item.output_path is None or not item.output_path.exists():
        raise HTTPException(status_code=404, detail="Teklif raporu bulunamadı.")
    return FileResponse(item.output_path, filename=item.output_path.name)


@app.get("/download/{token}/{kind}")
async def download_file(token: str, kind: str) -> FileResponse:
    session = SESSIONS.get(token)
    if session is None:
        raise HTTPException(status_code=404, detail="Oturum bulunamadı.")

    if kind == "report":
        path = session.output_path
    elif kind == "corrected":
        path = session.corrected_pdf_path
    else:
        raise HTTPException(status_code=404, detail="Geçersiz dosya tipi.")

    if path is None or not path.exists():
        raise HTTPException(status_code=404, detail="Dosya bulunamadı.")

    return FileResponse(path, filename=path.name)


def find_free_port(start: int = 8765) -> int:
    for port in range(start, start + 20):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            if sock.connect_ex(("127.0.0.1", port)) != 0:
                return port
    raise RuntimeError("Boş port bulunamadı.")


def wait_for_server(port: int, timeout: float = 12.0) -> bool:
    deadline = time.monotonic() + timeout
    while time.monotonic() < deadline:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.settimeout(0.25)
            if sock.connect_ex(("127.0.0.1", port)) == 0:
                return True
        time.sleep(0.1)
    return False


def open_browser_when_ready(url: str, port: int) -> None:
    if wait_for_server(port):
        try:
            webbrowser.open(url)
        except Exception:
            pass


def create_offer_app() -> FastAPI:
    ensure_runtime_folders()
    return app


def launch() -> None:
    configure_logging()
    ensure_runtime_folders()
    port = find_free_port()
    url = f"http://127.0.0.1:{port}"
    threading.Thread(target=open_browser_when_ready, args=(url, port), daemon=True).start()
    print(f"Rainwater Teklif Kontrol başladı: {url}")
    uvicorn.run(app, host="127.0.0.1", port=port, log_level="warning", log_config=None, access_log=False)


if __name__ == "__main__":
    launch()
