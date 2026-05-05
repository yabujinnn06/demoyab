from __future__ import annotations

import csv
import importlib
import json
import sqlite3
import sys
import uuid
from io import StringIO
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from fastapi.testclient import TestClient


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


def load_fresh_app():
    for module_name in (
        "backend.security",
        "backend.database",
        "backend.xlsx_reader",
        "backend.offer_module.teklif_kontrol",
        "backend.offer_module.portal_auth",
        "backend.offer_module.webapp",
        "backend.offer_module",
        "backend.app",
    ):
        if module_name in sys.modules:
            importlib.reload(sys.modules[module_name])
        else:
            importlib.import_module(module_name)
    return sys.modules["backend.app"].app


def make_test_db_path() -> Path:
    root = PROJECT_ROOT / ".tmp_testdata"
    root.mkdir(parents=True, exist_ok=True)
    return root / f"{uuid.uuid4().hex}.db"


def build_xlsx_bytes(rows: list[list[str]]) -> bytes:
    shared_values: list[str] = []
    shared_lookup: dict[str, int] = {}

    def shared_index(value: str) -> int:
        if value not in shared_lookup:
            shared_lookup[value] = len(shared_values)
            shared_values.append(value)
        return shared_lookup[value]

    def column_name(index: int) -> str:
        result = ""
        while index > 0:
            index, remainder = divmod(index - 1, 26)
            result = chr(65 + remainder) + result
        return result

    sheet_rows = []
    for row_number, row in enumerate(rows, start=1):
        cells = []
        for column_number, value in enumerate(row, start=1):
            ref = f"{column_name(column_number)}{row_number}"
            idx = shared_index(value)
            cells.append(f'<c r="{ref}" t="s"><v>{idx}</v></c>')
        sheet_rows.append(f'<row r="{row_number}">{"".join(cells)}</row>')

    shared_strings_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        f'count="{len(shared_values)}" uniqueCount="{len(shared_values)}">'
        + "".join(f"<si><t>{value}</t></si>" for value in shared_values)
        + "</sst>"
    )
    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f"<sheetData>{''.join(sheet_rows)}</sheetData>"
        "</worksheet>"
    )
    workbook_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<sheets><sheet name="Sayfa1" sheetId="1" r:id="rId1" /></sheets>'
        "</workbook>"
    )
    workbook_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        'Target="worksheets/sheet1.xml" />'
        '<Relationship Id="rId2" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" '
        'Target="sharedStrings.xml" />'
        "</Relationships>"
    )
    content_types_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml" />'
        '<Default Extension="xml" ContentType="application/xml" />'
        '<Override PartName="/xl/workbook.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml" />'
        '<Override PartName="/xl/worksheets/sheet1.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml" />'
        '<Override PartName="/xl/sharedStrings.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml" />'
        "</Types>"
    )
    root_rels_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="xl/workbook.xml" />'
        "</Relationships>"
    )

    from io import BytesIO

    buffer = BytesIO()
    with ZipFile(buffer, "w", ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", content_types_xml)
        archive.writestr("_rels/.rels", root_rels_xml)
        archive.writestr("xl/workbook.xml", workbook_xml)
        archive.writestr("xl/_rels/workbook.xml.rels", workbook_rels_xml)
        archive.writestr("xl/worksheets/sheet1.xml", sheet_xml)
        archive.writestr("xl/sharedStrings.xml", shared_strings_xml)
    return buffer.getvalue()


def test_login_import_assign_and_agent_update(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()

    workbook = build_xlsx_bytes(
        [
            ["İsim", "Adres", "Telefon", "Website", "Email"],
            ["Alem Bar", "Alsancak", "+90 530 237 14 74", "https://ornek.com", "mail@ornek.com"],
            ["Kordon Cafe", "Konak", "0530 111 22 33", "https://kordon.example", "iletisim@kordon.com"],
        ]
    )

    with TestClient(app) as client:
        login = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin12345!"},
        )
        assert login.status_code == 200
        admin_token = login.json()["access_token"]

        create_user = client.post(
            "/api/users",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "full_name": "Operator Bir",
                "email": "operator@test.local",
                "password": "Operator123!",
                "role": "agent",
            },
        )
        assert create_user.status_code == 201
        agent_id = create_user.json()["id"]

        imported = client.post(
            "/api/lists/import",
            headers={
                "Authorization": f"Bearer {admin_token}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "X-File-Name": "izmir.xlsx",
                "X-List-Name": "Izmir Arama",
            },
            content=workbook,
        )
        assert imported.status_code == 201
        call_list_id = imported.json()["id"]
        assert imported.json()["summary"]["total"] == 2

        assigned = client.post(
            f"/api/lists/{call_list_id}/assign-evenly",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={"user_ids": [agent_id], "mode": "all"},
        )
        assert assigned.status_code == 200
        assert assigned.json()["assigned_count"] == 2

        agent_login = client.post(
            "/api/auth/login",
            json={"email": "operator@test.local", "password": "Operator123!"},
        )
        assert agent_login.status_code == 200
        agent_token = agent_login.json()["access_token"]

        records = client.get(
            f"/api/records?call_list_id={call_list_id}",
            headers={"Authorization": f"Bearer {agent_token}"},
        )
        assert records.status_code == 200
        assert records.json()["total"] == 2
        record_id = records.json()["items"][0]["id"]

        updated = client.patch(
            f"/api/records/{record_id}",
            headers={"Authorization": f"Bearer {agent_token}"},
            json={
                "call_status": "CALLING",
                "result_status": "PENDING",
                "note": "Arama basladi",
            },
        )
        assert updated.status_code == 200
        assert updated.json()["call_status"] == "CALLING"
        assert updated.json()["assigned_user_id"] == agent_id

        processed_records = client.get(
            f"/api/records?call_list_id={call_list_id}&assigned_user_id={agent_id}&processed=true",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert processed_records.status_code == 200
        assert processed_records.json()["total"] == 1
        assert processed_records.json()["items"][0]["id"] == record_id

        health = client.get("/health")
        assert health.status_code == 200
        assert health.json() == {"status": "ok", "db": "ok"}
    db_path.unlink(missing_ok=True)


def test_processed_records_are_added_to_contact_pool(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()

    workbook = build_xlsx_bytes(
        [
            ["İsim", "Adres", "Telefon", "Website", "Email"],
            ["Havuz Klinik", "Bayrakli", "05301112233", "https://havuz.example", "info@havuz.example"],
        ]
    )

    with TestClient(app) as client:
        admin_login = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin12345!"},
        )
        admin_token = admin_login.json()["access_token"]
        agent_id = client.post(
            "/api/users",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "full_name": "Operator Bir",
                "email": "operator@test.local",
                "password": "Operator123!",
                "role": "agent",
            },
        ).json()["id"]
        imported = client.post(
            "/api/lists/import",
            headers={
                "Authorization": f"Bearer {admin_token}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "X-File-Name": "havuz.xlsx",
                "X-List-Name": "Havuz Test",
            },
            content=workbook,
        )
        call_list_id = imported.json()["id"]
        client.post(
            f"/api/lists/{call_list_id}/assign-evenly",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={"user_ids": [agent_id], "mode": "all"},
        )
        agent_login = client.post(
            "/api/auth/login",
            json={"email": "operator@test.local", "password": "Operator123!"},
        )
        agent_token = agent_login.json()["access_token"]
        records = client.get(
            f"/api/records?call_list_id={call_list_id}",
            headers={"Authorization": f"Bearer {agent_token}"},
        )
        record_id = records.json()["items"][0]["id"]

        updated = client.patch(
            f"/api/records/{record_id}",
            headers={"Authorization": f"Bearer {agent_token}"},
            json={
                "call_status": "CALLED",
                "result_status": "POSITIVE",
                "note": "Yetkiliye ulasildi",
            },
        )
        assert updated.status_code == 200

        pool = client.get(
            f"/api/contact-pool?call_list_id={call_list_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert pool.status_code == 200
        assert pool.json()["total"] == 1
        entry = pool.json()["items"][0]
        assert entry["company_name"] == "Havuz Klinik"
        assert entry["reach_status"] == "REACHED"
        assert entry["record_note"] == "Yetkiliye ulasildi"

        operator_pool = client.get(
            "/api/contact-pool",
            headers={"Authorization": f"Bearer {agent_token}"},
        )
        assert operator_pool.status_code == 403

        edited = client.patch(
            f"/api/contact-pool/{entry['id']}",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={"reach_status": "UNREACHED", "admin_note": "Tekrar teyit edilecek", "is_active": True},
        )
        assert edited.status_code == 200
        assert edited.json()["reach_status"] == "UNREACHED"
        assert edited.json()["admin_note"] == "Tekrar teyit edilecek"

        exported = client.get(
            "/api/contact-pool/export.csv",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert exported.status_code == 200
        exported_text = exported.content.decode("utf-8-sig")
        assert "Havuz Klinik" in exported_text
        assert "Tekrar teyit edilecek" in exported_text

        stats = client.get(
            f"/api/operator-stats?call_list_id={call_list_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert stats.status_code == 200
        operator_stat = next(item for item in stats.json() if item["user_id"] == agent_id)
        assert operator_stat["assigned_count"] == 1
        assert operator_stat["processed_count"] == 1
        assert operator_stat["reached_count"] == 1
        assert operator_stat["positive_count"] == 1
        assert operator_stat["negative_count"] == 0

        filtered_records = client.get(
            f"/api/records?call_list_id={call_list_id}&assigned_user_id={agent_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert filtered_records.status_code == 200
        assert filtered_records.json()["total"] == 1

        operator_stats_denied = client.get(
            "/api/operator-stats",
            headers={"Authorization": f"Bearer {agent_token}"},
        )
        assert operator_stats_denied.status_code == 403

    db_path.unlink(missing_ok=True)


def test_operation_summary_tracks_targets_and_due_callbacks(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()
    workbook = build_xlsx_bytes(
        [
            ["İsim", "Adres", "Telefon", "Website", "Email"],
            ["Takip Firma", "Konak", "05301112233", "https://takip.example", "takip@example.com"],
        ]
    )

    with TestClient(app) as client:
        admin_token = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin12345!"},
        ).json()["access_token"]
        agent_id = client.post(
            "/api/users",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "full_name": "Hedefli Operator",
                "email": "hedef@test.local",
                "password": "Operator123!",
                "role": "agent",
                "daily_target": 10,
            },
        ).json()["id"]
        call_list_id = client.post(
            "/api/lists/import",
            headers={
                "Authorization": f"Bearer {admin_token}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "X-File-Name": "takip.xlsx",
                "X-List-Name": "Takip Listesi",
            },
            content=workbook,
        ).json()["id"]
        client.post(
            f"/api/lists/{call_list_id}/assign-evenly",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={"user_ids": [agent_id], "mode": "all"},
        )

        record_id = client.get(
            f"/api/records?call_list_id={call_list_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
        ).json()["items"][0]["id"]
        callback_at = "2026-04-26T09:00"
        updated = client.patch(
            f"/api/records/{record_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={"call_status": "CALLBACK", "result_status": "PENDING", "callback_at": callback_at},
        )
        assert updated.status_code == 200
        assert updated.json()["callback_at"].startswith("2026-04-26T09:00")

        due_records = client.get(
            f"/api/records?call_list_id={call_list_id}&due_callbacks=true",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert due_records.status_code == 200
        assert due_records.json()["total"] == 1

        summary = client.get(
            f"/api/operation-summary?call_list_id={call_list_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert summary.status_code == 200
        assert summary.json()["due_callback_count"] == 1
        assert summary.json()["total_daily_target"] == 10

        stats = client.get(
            f"/api/operator-stats?call_list_id={call_list_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert stats.status_code == 200
        operator_stat = next(item for item in stats.json() if item["user_id"] == agent_id)
        assert operator_stat["daily_target"] == 10
        assert operator_stat["callback_count"] == 1

    db_path.unlink(missing_ok=True)


def test_agent_export_only_returns_assigned_rows(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()

    workbook = build_xlsx_bytes(
        [
            ["İsim", "Adres", "Telefon", "Website", "Email"],
            ["Birinci Firma", "Konak", "05301112233", "https://bir.example", "bir@example.com"],
            ["İkinci Firma", "Bornova", "05302223344", "https://iki.example", "iki@example.com"],
        ]
    )

    with TestClient(app) as client:
        admin_login = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin12345!"},
        )
        admin_token = admin_login.json()["access_token"]

        first_agent = client.post(
            "/api/users",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "full_name": "Operator Bir",
                "email": "operator1@test.local",
                "password": "Operator123!",
                "role": "agent",
            },
        ).json()["id"]

        second_agent = client.post(
            "/api/users",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "full_name": "Operator İki",
                "email": "operator2@test.local",
                "password": "Operator123!",
                "role": "agent",
            },
        ).json()["id"]

        imported = client.post(
            "/api/lists/import",
            headers={
                "Authorization": f"Bearer {admin_token}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "X-File-Name": "izmir.xlsx",
                "X-List-Name": "Izmir Arama",
            },
            content=workbook,
        )
        call_list_id = imported.json()["id"]

        assigned = client.post(
            f"/api/lists/{call_list_id}/assign-evenly",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={"user_ids": [first_agent, second_agent], "mode": "all"},
        )
        assert assigned.status_code == 200

        agent_login = client.post(
            "/api/auth/login",
            json={"email": "operator1@test.local", "password": "Operator123!"},
        )
        agent_token = agent_login.json()["access_token"]

        exported = client.get(
            f"/api/lists/{call_list_id}/export.csv",
            headers={"Authorization": f"Bearer {agent_token}"},
        )
        assert exported.status_code == 200

        rows = list(csv.reader(StringIO(exported.content.decode("utf-8-sig"))))
        assert len(rows) == 2
        assert rows[1][1] == "Birinci Firma"
    db_path.unlink(missing_ok=True)


def test_records_endpoint_filters_and_summarizes_on_server(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()

    workbook = build_xlsx_bytes(
        [
            ["İsim", "Adres", "Telefon", "Website", "Email"],
            ["Email Firma", "Konak", "05301112233", "https://bir.example", "bir@example.com"],
            ["Telefonsuz Firma", "Bornova", "", "https://iki.example", ""],
        ]
    )

    with TestClient(app) as client:
        admin_login = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin12345!"},
        )
        admin_token = admin_login.json()["access_token"]

        imported = client.post(
            "/api/lists/import",
            headers={
                "Authorization": f"Bearer {admin_token}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "X-File-Name": "izmir.xlsx",
                "X-List-Name": "Izmir Arama",
            },
            content=workbook,
        )
        call_list_id = imported.json()["id"]

        filtered = client.get(
            f"/api/records?call_list_id={call_list_id}&has_email=true&limit=100",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert filtered.status_code == 200
        assert filtered.json()["total"] == 1
        assert filtered.json()["summary"]["total"] == 1
        assert filtered.json()["items"][0]["company_name"] == "Email Firma"

    db_path.unlink(missing_ok=True)


def test_render_startup_requires_nondefault_admin_password(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("RENDER", "true")
    monkeypatch.setenv("CALL_PORTAL_SECRET_KEY", "super-secret-key-for-render-1234567890")
    monkeypatch.delenv("CALL_PORTAL_ADMIN_PASSWORD", raising=False)
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")

    app = load_fresh_app()

    with pytest.raises(RuntimeError, match="CALL_PORTAL_ADMIN_PASSWORD"):
        with TestClient(app):
            pass
    db_path.unlink(missing_ok=True)


def test_custom_assign_respects_requested_counts(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()

    workbook = build_xlsx_bytes(
        [
            ["İsim", "Adres", "Telefon", "Website", "Email"],
            ["Firma 1", "Konak", "05300000001", "", "f1@example.com"],
            ["Firma 2", "Konak", "05300000002", "", "f2@example.com"],
            ["Firma 3", "Konak", "05300000003", "", "f3@example.com"],
            ["Firma 4", "Konak", "05300000004", "", "f4@example.com"],
            ["Firma 5", "Konak", "05300000005", "", "f5@example.com"],
        ]
    )

    with TestClient(app) as client:
        admin_login = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin12345!"},
        )
        admin_token = admin_login.json()["access_token"]

        first_agent = client.post(
            "/api/users",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "full_name": "Operator Bir",
                "email": "operator1@test.local",
                "password": "Operator123!",
                "role": "agent",
            },
        ).json()["id"]

        second_agent = client.post(
            "/api/users",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "full_name": "Operator İki",
                "email": "operator2@test.local",
                "password": "Operator123!",
                "role": "agent",
            },
        ).json()["id"]

        imported = client.post(
            "/api/lists/import",
            headers={
                "Authorization": f"Bearer {admin_token}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "X-File-Name": "izmir.xlsx",
                "X-List-Name": "Izmir Arama",
            },
            content=workbook,
        )
        call_list_id = imported.json()["id"]

        assigned = client.post(
            f"/api/lists/{call_list_id}/assign-custom",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "mode": "all",
                "allocations": [
                    {"user_id": first_agent, "count": 2},
                    {"user_id": second_agent, "count": 1},
                ],
            },
        )
        assert assigned.status_code == 200
        assert assigned.json()["assigned_count"] == 3
        assert assigned.json()["remaining_count"] == 2

        records = client.get(
            f"/api/records?call_list_id={call_list_id}&limit=100",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert records.status_code == 200
        items = records.json()["items"]
        first_count = sum(1 for item in items if item["assigned_user_id"] == first_agent)
        second_count = sum(1 for item in items if item["assigned_user_id"] == second_agent)
        unassigned_count = sum(1 for item in items if item["assigned_user_id"] is None)
        assert first_count == 2
        assert second_count == 1
        assert unassigned_count == 2

    db_path.unlink(missing_ok=True)


def test_assignment_rejects_admin_users(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()
    workbook = build_xlsx_bytes(
        [
            ["İsim", "Adres", "Telefon", "Website", "Email"],
            ["Firma 1", "Konak", "05300000001", "", "f1@example.com"],
        ]
    )

    with TestClient(app) as client:
        admin_login = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin12345!"},
        )
        admin_token = admin_login.json()["access_token"]

        second_admin = client.post(
            "/api/users",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "full_name": "Admin Iki",
                "email": "admin2@test.local",
                "password": "Admin23456!",
                "role": "admin",
            },
        ).json()["id"]

        imported = client.post(
            "/api/lists/import",
            headers={
                "Authorization": f"Bearer {admin_token}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "X-File-Name": "izmir.xlsx",
                "X-List-Name": "Izmir Arama",
            },
            content=workbook,
        )
        call_list_id = imported.json()["id"]

        assigned = client.post(
            f"/api/lists/{call_list_id}/assign-evenly",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={"user_ids": [second_admin], "mode": "all"},
        )
        assert assigned.status_code == 422

    db_path.unlink(missing_ok=True)


def test_agent_cannot_update_unassigned_record(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()
    workbook = build_xlsx_bytes(
        [
            ["İsim", "Adres", "Telefon", "Website", "Email"],
            ["Firma 1", "Konak", "05300000001", "", "f1@example.com"],
        ]
    )

    with TestClient(app) as client:
        admin_login = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin12345!"},
        )
        admin_token = admin_login.json()["access_token"]

        client.post(
            "/api/users",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "full_name": "Operator Bir",
                "email": "operator@test.local",
                "password": "Operator123!",
                "role": "agent",
            },
        )

        imported = client.post(
            "/api/lists/import",
            headers={
                "Authorization": f"Bearer {admin_token}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "X-File-Name": "izmir.xlsx",
                "X-List-Name": "Izmir Arama",
            },
            content=workbook,
        )
        call_list_id = imported.json()["id"]
        records = client.get(
            f"/api/records?call_list_id={call_list_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        record_id = records.json()["items"][0]["id"]

        agent_login = client.post(
            "/api/auth/login",
            json={"email": "operator@test.local", "password": "Operator123!"},
        )
        agent_token = agent_login.json()["access_token"]

        updated = client.patch(
            f"/api/records/{record_id}",
            headers={"Authorization": f"Bearer {agent_token}"},
            json={"call_status": "CALLING", "result_status": "PENDING"},
        )
        assert updated.status_code == 403

    db_path.unlink(missing_ok=True)


def test_agent_cannot_export_inactive_list(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()
    workbook = build_xlsx_bytes(
        [
            ["İsim", "Adres", "Telefon", "Website", "Email"],
            ["Firma 1", "Konak", "05300000001", "", "f1@example.com"],
        ]
    )

    with TestClient(app) as client:
        admin_login = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin12345!"},
        )
        admin_token = admin_login.json()["access_token"]

        agent_id = client.post(
            "/api/users",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "full_name": "Operator Bir",
                "email": "operator@test.local",
                "password": "Operator123!",
                "role": "agent",
            },
        ).json()["id"]

        imported = client.post(
            "/api/lists/import",
            headers={
                "Authorization": f"Bearer {admin_token}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "X-File-Name": "izmir.xlsx",
                "X-List-Name": "Izmir Arama",
            },
            content=workbook,
        )
        call_list_id = imported.json()["id"]
        client.post(
            f"/api/lists/{call_list_id}/assign-evenly",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={"user_ids": [agent_id], "mode": "all"},
        )
        client.patch(
            f"/api/lists/{call_list_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={"is_active": False},
        )

        agent_login = client.post(
            "/api/auth/login",
            json={"email": "operator@test.local", "password": "Operator123!"},
        )
        agent_token = agent_login.json()["access_token"]
        exported = client.get(
            f"/api/lists/{call_list_id}/export.csv",
            headers={"Authorization": f"Bearer {agent_token}"},
        )
        assert exported.status_code == 403

    db_path.unlink(missing_ok=True)


def test_password_change_invalidates_existing_token(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()

    with TestClient(app) as client:
        admin_login = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin12345!"},
        )
        admin_token = admin_login.json()["access_token"]

        agent_id = client.post(
            "/api/users",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "full_name": "Operator Bir",
                "email": "operator@test.local",
                "password": "Operator123!",
                "role": "agent",
            },
        ).json()["id"]

        agent_login = client.post(
            "/api/auth/login",
            json={"email": "operator@test.local", "password": "Operator123!"},
        )
        old_token = agent_login.json()["access_token"]

        changed = client.patch(
            f"/api/users/{agent_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={"full_name": "Operator Bir", "password": "Operator456!", "role": "agent", "is_active": True},
        )
        assert changed.status_code == 200

        stale_session = client.get(
            "/api/auth/me",
            headers={"Authorization": f"Bearer {old_token}"},
        )
        assert stale_session.status_code == 401

        new_login = client.post(
            "/api/auth/login",
            json={"email": "operator@test.local", "password": "Operator456!"},
        )
        assert new_login.status_code == 200

    db_path.unlink(missing_ok=True)


def test_user_delete_is_soft_delete_and_invalidates_session(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()

    with TestClient(app) as client:
        admin_login = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin12345!"},
        )
        admin_token = admin_login.json()["access_token"]

        agent_id = client.post(
            "/api/users",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "full_name": "Operator Bir",
                "email": "operator@test.local",
                "password": "Operator123!",
                "role": "agent",
            },
        ).json()["id"]

        agent_login = client.post(
            "/api/auth/login",
            json={"email": "operator@test.local", "password": "Operator123!"},
        )
        old_token = agent_login.json()["access_token"]

        deleted = client.delete(
            f"/api/users/{agent_id}",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        assert deleted.status_code == 200

        stale_session = client.get(
            "/api/auth/me",
            headers={"Authorization": f"Bearer {old_token}"},
        )
        assert stale_session.status_code == 401

        users = client.get(
            "/api/users",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        soft_deleted_user = next(item for item in users.json() if item["id"] == agent_id)
        assert soft_deleted_user["is_active"] is False

    db_path.unlink(missing_ok=True)


def test_csv_export_escapes_formula_values(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()
    workbook = build_xlsx_bytes(
        [
            ["İsim", "Adres", "Telefon", "Website", "Email"],
            ["=FORMULA", "+Adres", "05300000001", "https://example.com", "f1@example.com"],
        ]
    )

    with TestClient(app) as client:
        admin_login = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin12345!"},
        )
        admin_token = admin_login.json()["access_token"]
        imported = client.post(
            "/api/lists/import",
            headers={
                "Authorization": f"Bearer {admin_token}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "X-File-Name": "izmir.xlsx",
                "X-List-Name": "Izmir Arama",
            },
            content=workbook,
        )
        call_list_id = imported.json()["id"]
        exported = client.get(
            f"/api/lists/{call_list_id}/export.csv",
            headers={"Authorization": f"Bearer {admin_token}"},
        )
        rows = list(csv.reader(StringIO(exported.content.decode("utf-8-sig"))))
        assert rows[1][1] == "'=FORMULA"
        assert rows[1][2] == "'+Adres"

    db_path.unlink(missing_ok=True)


def test_user_create_validation_returns_readable_detail(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()

    with TestClient(app) as client:
        admin_login = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin12345!"},
        )
        admin_token = admin_login.json()["access_token"]
        response = client.post(
            "/api/users",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "email": "operator@test.local",
                "full_name": "Operator",
                "password": "operator123!",
                "role": "agent",
            },
        )

    assert response.status_code == 422
    assert response.json()["detail"] == "Şifre: Şifre en az bir büyük harf içermeli."
    db_path.unlink(missing_ok=True)


def test_favicon_route_uses_brand_mark(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()

    with TestClient(app) as client:
        response = client.get("/favicon.ico")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("image/svg+xml")
    db_path.unlink(missing_ok=True)


def test_ping_route_is_public_and_lightweight(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()

    with TestClient(app) as client:
        response = client.get("/ping")

    assert response.status_code == 200
    assert response.json() == {"status": "ok"}
    assert response.headers["cache-control"] == "no-store"
    db_path.unlink(missing_ok=True)


def test_head_requests_work_for_root_ping_and_health(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()

    with TestClient(app) as client:
        root_response = client.head("/")
        ping_response = client.head("/ping")
        health_response = client.head("/health")

    assert root_response.status_code == 200
    assert ping_response.status_code == 200
    assert health_response.status_code == 200
    assert root_response.text == ""
    assert ping_response.text == ""
    assert health_response.text == ""
    db_path.unlink(missing_ok=True)


def test_xlsx_parser_rejects_oversized_zip_structure() -> None:
    from io import BytesIO

    from backend.xlsx_reader import parse_xlsx_records

    buffer = BytesIO()
    with ZipFile(buffer, "w", ZIP_DEFLATED) as archive:
        for index in range(251):
            archive.writestr(f"xl/extra{index}.xml", "<x />")

    with pytest.raises(ValueError, match="fazla parca"):
        parse_xlsx_records(buffer.getvalue())


def test_login_rate_limit_persists_in_database(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()

    with TestClient(app) as client:
        for _ in range(8):
            response = client.post(
                "/api/auth/login",
                json={"email": "admin@test.local", "password": "yanlis-sifre"},
            )
            assert response.status_code == 401

        blocked = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "yanlis-sifre"},
        )
        assert blocked.status_code == 429

    connection = sqlite3.connect(db_path)
    try:
        total = connection.execute("SELECT COUNT(*) FROM login_attempts").fetchone()[0]
        assert total == 8
    finally:
        connection.close()
    db_path.unlink(missing_ok=True)


def test_startup_creates_schema_migration_state(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()

    with TestClient(app):
        pass

    connection = sqlite3.connect(db_path)
    try:
        applied_versions = {
            row[0]
            for row in connection.execute("SELECT version FROM schema_migrations ORDER BY version").fetchall()
        }
        assert applied_versions == {1, 2, 3, 4, 5, 6}
        login_attempts_table = connection.execute(
            "SELECT name FROM sqlite_master WHERE type = 'table' AND name = 'login_attempts'"
        ).fetchone()
        assert login_attempts_table is not None
        token_version_column = connection.execute("PRAGMA table_info(users)").fetchall()
        user_columns = {row[1] for row in token_version_column}
        assert "token_version" in user_columns
        assert "can_access_offer_tool" in user_columns
        contact_pool_table = connection.execute(
            "SELECT name FROM sqlite_master WHERE type = 'table' AND name = 'contact_pool_entries'"
        ).fetchone()
        assert contact_pool_table is not None
    finally:
        connection.close()
    db_path.unlink(missing_ok=True)


def test_offer_module_requires_permission_and_uses_session_cookie(monkeypatch) -> None:
    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()

    with TestClient(app) as client:
        denied_redirect = client.get("/teklif/", follow_redirects=False)
        assert denied_redirect.status_code in {303, 307, 401}

        admin_login = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin12345!"},
        )
        assert admin_login.status_code == 200
        assert "call_portal_session" in admin_login.cookies
        me_from_cookie = client.get("/api/auth/me")
        assert me_from_cookie.status_code == 200
        assert me_from_cookie.json()["email"] == "admin@test.local"

        admin_offer = client.get("/teklif/")
        assert admin_offer.status_code == 200
        assert "Rainwater Teklif Ofisi" in admin_offer.text
        assert "Teklif akışlarını tek merkezden yönet" in admin_offer.text
        assert "Şablon PDF yükle" in admin_offer.text
        assert "/teklif/create-offer" in admin_offer.text
        assert "/teklif/admin/import-template" in admin_offer.text
        assert "/teklif/admin/create-offer" not in admin_offer.text
        assert "/teklif/static/vendor/bootstrap/bootstrap.min.css" in admin_offer.text
        assert "/teklif/static/styles.css" in admin_offer.text

        offer_bootstrap = client.get("/teklif/static/vendor/bootstrap/bootstrap.min.css")
        assert offer_bootstrap.status_code == 200
        offer_styles = client.get("/teklif/static/styles.css")
        assert offer_styles.status_code == 200
        offer_sound = client.get("/teklif/static/audio/choose2.wav")
        assert offer_sound.status_code == 200
        assert len(offer_sound.content) > 0
        offer_app_js = client.get("/teklif/static/app.js")
        assert offer_app_js.status_code == 200
        assert "preferServerWorkspace" in offer_app_js.text
        assert "data-manual-match" in offer_app_js.text

        created_user = client.post(
            "/api/users",
            headers={"Authorization": f"Bearer {admin_login.json()['access_token']}"},
            json={
                "full_name": "Teklifsiz Operator",
                "email": "agent@test.local",
                "password": "Operator123!",
                "role": "agent",
                "can_access_offer_tool": False,
            },
        )
        assert created_user.status_code == 201

        agent_client = TestClient(app)
        with agent_client:
            agent_login = agent_client.post(
                "/api/auth/login",
                json={"email": "agent@test.local", "password": "Operator123!"},
            )
            assert agent_login.status_code == 200
            denied_offer = agent_client.get("/teklif/", follow_redirects=False)
            assert denied_offer.status_code == 403

        updated_user = client.patch(
            f"/api/users/{created_user.json()['id']}",
            headers={"Authorization": f"Bearer {admin_login.json()['access_token']}"},
            json={"can_access_offer_tool": True},
        )
        assert updated_user.status_code == 200
        assert updated_user.json()["can_access_offer_tool"] is True

        grant_client = TestClient(app)
        with grant_client:
            grant_login = grant_client.post(
                "/api/auth/login",
                json={"email": "agent@test.local", "password": "Operator123!"},
            )
            assert grant_login.status_code == 200
        granted_offer = grant_client.get("/teklif/")
        assert granted_offer.status_code == 200
        assert "Teklif akışlarını tek merkezden yönet" in granted_offer.text
        assert "Şablon PDF yükle" not in granted_offer.text
        assert "/teklif/static/styles.css" in granted_offer.text


def test_offer_notifications_include_pending_for_creator_without_offer_access(tmp_path, monkeypatch) -> None:
    offer_base = tmp_path / "offer_data"
    offer_data_dir = offer_base / "veri"
    offer_dir = offer_base / "teklifler"
    offer_data_dir.mkdir(parents=True)
    offer_dir.mkdir(parents=True)

    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.setenv("CALL_PORTAL_OFFER_DATA_DIR", str(offer_base))
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()

    with TestClient(app) as client:
        admin_login = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin12345!"},
        )
        assert admin_login.status_code == 200
        admin_token = admin_login.json()["access_token"]

        created_user = client.post(
            "/api/users",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={
                "full_name": "Bildirim Operator",
                "email": "bildirim@test.local",
                "password": "Operator123!",
                "role": "agent",
                "can_access_offer_tool": False,
            },
        )
        assert created_user.status_code == 201
        creator_id = created_user.json()["id"]
        generated_offer = offer_dir / "RW-TEST-1.pdf"
        generated_offer.write_bytes(b"%PDF-1.4\n")

        (offer_data_dir / "pending_offer_approvals.json").write_text(
            json.dumps(
                {
                    "approval-1": {
                        "approval_id": "approval-1",
                        "activity_entry_id": "entry-create",
                        "creator_user_id": creator_id,
                        "creator_email": "bildirim@test.local",
                        "creator_name": "Bildirim Operator",
                        "status": "pending",
                        "offer_number": "RW-TEST-1",
                        "company_name": "Test Firma",
                        "generated_path": "teklifler/RW-TEST-1.pdf",
                        "generated_name": "RW-TEST-1.pdf",
                    }
                }
            ),
            encoding="utf-8",
        )
        (offer_data_dir / "offer_activity_log.json").write_text(
            json.dumps(
                [
                    {
                        "id": "entry-create",
                        "created_at": "2026-05-05T10:00:00",
                        "actor_id": creator_id,
                        "actor_email": "bildirim@test.local",
                        "actor_name": "Bildirim Operator",
                        "action": "create",
                        "action_label": "Teklif oluşturuldu",
                        "summary": "RW-TEST-1.pdf oluşturuldu; müşteri Test Firma",
                        "files": [
                            {
                                "label": "Oluşturulan teklif",
                                "kind": "generated",
                                "path": "teklifler/RW-TEST-1.pdf",
                                "name": "RW-TEST-1.pdf",
                            }
                        ],
                        "details": {
                            "approval_id": "approval-1",
                            "offer_number": "RW-TEST-1",
                            "company_name": "Test Firma",
                        },
                    }
                ]
            ),
            encoding="utf-8",
        )

        agent_login = client.post(
            "/api/auth/login",
            json={"email": "bildirim@test.local", "password": "Operator123!"},
        )
        assert agent_login.status_code == 200
        agent_token = agent_login.json()["access_token"]

        response = client.get(
            "/api/offer-notifications",
            headers={"Authorization": f"Bearer {agent_token}"},
        )

        assert response.status_code == 200
        assert response.json() == [
            {
                "id": "approval-1",
                "status": "pending",
                "offer_number": "RW-TEST-1",
                "company_name": "Test Firma",
                "contact_name": "",
                "generated_name": "RW-TEST-1.pdf",
                "approved_at": "",
                "rejected_at": "",
                "rejection_reason": "",
                "download_url": "",
            }
        ]

        activity_response = client.get(
            "/api/offer-activity",
            headers={"Authorization": f"Bearer {agent_token}"},
        )
        assert activity_response.status_code == 200
        activity = activity_response.json()
        assert len(activity) == 1
        assert activity[0]["approval_status"] == "pending"
        assert activity[0]["creator_name"] == "Bildirim Operator"
        assert activity[0]["creator_email"] == "bildirim@test.local"
        assert activity[0]["files"][0]["blocked"] is True
        assert activity[0]["files"][0]["url"] == ""

        approvals = json.loads((offer_data_dir / "pending_offer_approvals.json").read_text(encoding="utf-8"))
        approvals["approval-1"]["status"] = "approved"
        approvals["approval-1"]["approved_at"] = "2026-05-05T10:05:00"
        (offer_data_dir / "pending_offer_approvals.json").write_text(json.dumps(approvals), encoding="utf-8")

        approved_activity = client.get(
            "/api/offer-activity",
            headers={"Authorization": f"Bearer {agent_token}"},
        ).json()
        assert approved_activity[0]["approval_status"] == "approved"
        assert approved_activity[0]["files"][0]["blocked"] is False
        assert approved_activity[0]["files"][0]["url"] == "/api/offer-activity/entry-create/files/0"

        download = client.get(
            approved_activity[0]["files"][0]["url"],
            headers={"Authorization": f"Bearer {agent_token}"},
        )
        assert download.status_code == 200
        assert download.content.startswith(b"%PDF-1.4")


def test_offer_manual_match_preview_makes_row_actionable(tmp_path) -> None:
    from openpyxl import Workbook

    from backend.offer_module import webapp as offer_webapp
    from backend.offer_module.teklif_kontrol import FinancialReview, MatchResult, OfferItem

    price_path = tmp_path / "fiyat.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["URUN", "2026 KURUMSAL NAKIT"])
    sheet.append(["Rainwater Superior Paslanmaz Tank", 100])
    workbook.save(price_path)

    session = offer_webapp.ComparisonSession(
        token="manual-test",
        price_list_path=price_path,
        offer_path=tmp_path / "teklif.pdf",
        output_path=tmp_path / "rapor.xlsx",
        selected_column="2026 KURUMSAL NAKIT",
        price_mode="kurumsal_nakit",
        results=[
            MatchResult(
                offer_item=OfferItem(
                    product_name="Bilinmeyen tank",
                    quantity=1,
                    unit_price=80,
                    discounted_price=80,
                    total_price=80,
                ),
                matched_row=None,
                score=0,
                status="ESLESMEDI",
                selected_column="2026 KURUMSAL NAKIT",
                reference_unit_price=None,
                reference_total_price=None,
                suggested_unit_price=None,
                suggested_total_price=None,
                difference=None,
                note="Güvenilir eşleşme bulunamadı.",
            )
        ],
        financial_review=FinancialReview(
            vat_rate=20,
            vat_rate_source="default",
            vat_included=True,
            item_gross_total=80,
            expected_net_total=66.67,
            expected_vat_total=13.33,
            expected_gross_total=80,
            expected_summary_total=80,
            checks=[],
        ),
    )

    changed_count = offer_webapp.apply_manual_match_overrides(session, ["2"])
    row = offer_webapp.result_view_model(session.results[0], 0)

    assert changed_count == 1
    assert row["manual_selected"] is True
    assert row["manual_match_row_id"] == "2"
    assert row["status"] == "DUZELT"
    assert row["can_apply"] is True


def test_batch_comparison_session_and_summary_are_reusable(tmp_path, monkeypatch) -> None:
    from backend.offer_module import webapp as offer_webapp
    from backend.offer_module.teklif_kontrol import FinancialCheck, FinancialReview, MatchResult, OfferItem

    price_path = tmp_path / "fiyat.xlsx"
    offer_path = tmp_path / "teklif.pdf"
    final_report_path = tmp_path / "teklif_rapor.xlsx"
    price_path.write_bytes(b"price")
    offer_path.write_bytes(b"%PDF-1.4\n")

    fake_review = FinancialReview(
        vat_rate=20,
        vat_rate_source="PDF",
        vat_included=True,
        item_gross_total=100,
        expected_net_total=83.33,
        expected_vat_total=16.67,
        expected_gross_total=100,
        expected_summary_total=100,
        checks=[FinancialCheck("Toplam", "ONAY", 100, 100, 0, "Uygun")],
    )
    fake_result = MatchResult(
        offer_item=OfferItem(
            product_name="Rainwater Test",
            quantity=1,
            unit_price=100,
            discounted_price=100,
            total_price=100,
        ),
        matched_row=None,
        score=100,
        status="ONAY",
        selected_column="2026 KURUMSAL NAKIT",
        reference_unit_price=100,
        reference_total_price=100,
        suggested_unit_price=None,
        suggested_total_price=None,
        difference=0,
        note="Uygun",
    )

    def fake_run_comparison(**_kwargs):
        final_report_path.write_bytes(b"report")
        return [fake_result], "2026 KURUMSAL NAKIT", final_report_path, ["2026 KURUMSAL NAKIT"], fake_review

    monkeypatch.setattr(offer_webapp, "BASE_DIR", tmp_path)
    monkeypatch.setattr(offer_webapp, "BATCH_JOBS_DIR", tmp_path / "batch_jobs")
    monkeypatch.setattr(offer_webapp, "run_comparison", fake_run_comparison)
    offer_webapp.SESSIONS.clear()
    offer_webapp.BATCHES.clear()

    session = offer_webapp.create_comparison_session(price_path, offer_path, "kurumsal_nakit")
    item = offer_webapp.batch_item_from_session(session)
    job = offer_webapp.BatchComparisonJob(
        token="batch-test-token",
        price_list_path=price_path,
        price_mode="kurumsal_nakit",
        created_at=offer_webapp.datetime.now(),
        summary_path=tmp_path / "toplu_ozet.xlsx",
        items=[item],
    )
    offer_webapp.write_batch_summary(job)
    offer_webapp.register_batch_job(job)
    offer_webapp.BATCHES.clear()
    reloaded_job = offer_webapp.load_batch_job(job.token)

    assert session.token in offer_webapp.SESSIONS
    assert item.metrics["ONAY"] == 1
    assert item.financial_status == "ONAY"
    assert item.problem_summary == "İşlem gerekmiyor."
    assert job.summary_path.exists()
    assert reloaded_job is not None
    assert reloaded_job.items[0].offer_path.name == "teklif.pdf"
    assert reloaded_job.items[0].problem_summary == "İşlem gerekmiyor."


def test_batch_compare_route_renders_results_without_server_error(tmp_path, monkeypatch) -> None:
    from openpyxl import Workbook

    offer_base = tmp_path / "offer_data"
    price_dir = offer_base / "veri" / "fiyat_listeleri"
    offer_dir = offer_base / "teklifler"
    price_dir.mkdir(parents=True)
    offer_dir.mkdir(parents=True)

    price_path = price_dir / "fiyat.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["URUN", "2026 KURUMSAL NAKIT"])
    sheet.append(["Rainwater Test", 100])
    workbook.save(price_path)

    (offer_dir / "a.pdf").write_bytes(b"%PDF-1.4\n")
    (offer_dir / "b.pdf").write_bytes(b"%PDF-1.4\n")

    db_path = make_test_db_path()
    monkeypatch.setenv("CALL_PORTAL_DB_PATH", str(db_path))
    monkeypatch.setenv("CALL_PORTAL_ADMIN_EMAIL", "admin@test.local")
    monkeypatch.setenv("CALL_PORTAL_ADMIN_PASSWORD", "Admin12345!")
    monkeypatch.setenv("CALL_PORTAL_OFFER_DATA_DIR", str(offer_base))
    monkeypatch.delenv("RENDER", raising=False)

    app = load_fresh_app()

    from backend.offer_module import webapp as offer_webapp
    from backend.offer_module.teklif_kontrol import FinancialCheck, FinancialReview, MatchResult, OfferItem

    def fake_run_comparison(**kwargs):
        output_path = kwargs["output_path"]
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(b"report")
        review = FinancialReview(
            vat_rate=20,
            vat_rate_source="PDF",
            vat_included=True,
            item_gross_total=100,
            expected_net_total=83.33,
            expected_vat_total=16.67,
            expected_gross_total=100,
            expected_summary_total=100,
            checks=[FinancialCheck("Toplam", "ONAY", 100, 100, 0, "Uygun")],
        )
        result = MatchResult(
            offer_item=OfferItem("Rainwater Test", 1, 100, 100, 100),
            matched_row=None,
            score=100,
            status="ONAY",
            selected_column="2026 KURUMSAL NAKIT",
            reference_unit_price=100,
            reference_total_price=100,
            suggested_unit_price=None,
            suggested_total_price=None,
            difference=0,
            note="Uygun",
        )
        return [result], "2026 KURUMSAL NAKIT", output_path, ["2026 KURUMSAL NAKIT"], review

    monkeypatch.setattr(offer_webapp, "run_comparison", fake_run_comparison)

    with TestClient(app) as client:
        login = client.post(
            "/api/auth/login",
            json={"email": "admin@test.local", "password": "Admin12345!"},
        )
        assert login.status_code == 200

        response = client.post(
            "/teklif/batch-compare",
            data={
                "price_file": "fiyat.xlsx",
                "price_mode": "kurumsal_nakit",
                "offer_files": ["teklifler/a.pdf", "teklifler/b.pdf"],
            },
        )

        assert response.status_code == 200
        assert "Toplu teklif kontrolü" in response.text
        assert "Akıllı toplu komut" in response.text
        assert "a.pdf" in response.text
        assert "b.pdf" in response.text
        assert "Raporları ZIP indir" in response.text
        batch_token = next(iter(offer_webapp.BATCHES))
        offer_webapp.BATCHES.clear()
        batch_response = client.get(f"/teklif/batch/{batch_token}")
        assert batch_response.status_code == 200
        assert "batch-issue-table" in batch_response.text


def test_rainwater_pdf_price_list_converts_to_readable_workbook(tmp_path) -> None:
    from backend.offer_module import webapp as offer_webapp
    from backend.offer_module.teklif_kontrol import load_price_rows

    pdf_path = (
        PROJECT_ROOT
        / "backend"
        / "offer_module"
        / "veri"
        / "fiyat_listeleri"
        / "20260427_FIYAT_LISTESI_2026.pdf"
    )
    workbook_path = tmp_path / "fiyat_listesi_2026.xlsx"

    converted_count = offer_webapp.convert_price_pdf_to_workbook(pdf_path, workbook_path)
    rows, headers = load_price_rows(workbook_path)

    assert converted_count == 75
    assert len(rows) == 75
    assert "2026 KURUMSAL NAKIT" in headers
    assert "2026 PERAKENDE 6 TAKSIT" in headers
    assert rows[0].product_name == "Rainwater Superior"
    assert rows[0].prices["2026 KURUMSAL NAKIT"] == 42950
    assert next(row for row in rows if row.product_name == "Rnw 3100 Arıtmasız Su Sebili").prices[
        "2026 KURUMSAL 6 TAKSIT"
    ] == 68950
    assert sum(1 for row in rows if row.product_name == "Rainwater 40 LT Emaye Tank") == 1
    emaye_80 = next(row for row in rows if row.product_name == "Rainwater 80 LT Emaye Tank")
    assert emaye_80.prices["2026 KURUMSAL NAKIT"] == 19950
    assert emaye_80.prices["2026 PERAKENDE 6 TAKSIT"] == 28950


def test_product_match_suggestions_surface_nearest_catalog_item() -> None:
    from backend.offer_module import webapp as offer_webapp
    from backend.offer_module.teklif_kontrol import MatchResult, OfferItem, PriceRow

    price_rows = [
        PriceRow(row_number=2, product_name="Rainwater Superior", prices={"2026 KURUMSAL 6 TAKSIT": 44950}),
        PriceRow(row_number=3, product_name="Rnw 3100 Arıtmasız Su Sebili", prices={"2026 KURUMSAL 6 TAKSIT": 68950}),
    ]
    result = MatchResult(
        offer_item=OfferItem(
            product_name="RNW 3100 Aritmasiz Su Sebili",
            quantity=1,
            unit_price=92950,
            discounted_price=92950,
            total_price=92950,
        ),
        matched_row=None,
        score=0,
        status="ESLESMEDI",
        selected_column="2026 KURUMSAL 6 TAKSIT",
        reference_unit_price=None,
        reference_total_price=None,
        suggested_unit_price=None,
        suggested_total_price=None,
        difference=None,
        note="Güvenilir eşleşme bulunamadı.",
    )

    suggestions = offer_webapp.product_match_suggestions(
        result,
        price_rows,
        selected_column="2026 KURUMSAL 6 TAKSIT",
        vat_included=True,
        vat_rate=20,
    )
    view = offer_webapp.result_view_model(
        result,
        0,
        price_rows=price_rows,
        selected_column="2026 KURUMSAL 6 TAKSIT",
        vat_included=True,
        vat_rate=20,
    )

    assert suggestions[0]["value"] == "3"
    assert suggestions[0]["price"] == "68.950 TL"
    assert view["action"]["title"] == "Ürün eşleşmedi, yakın aday var"
    assert view["suggestions"][0]["label"] == "Rnw 3100 Arıtmasız Su Sebili"


def test_bundle_match_view_model_shows_all_components() -> None:
    from backend.offer_module import webapp as offer_webapp
    from backend.offer_module.teklif_kontrol import BundleComponentMatch, MatchResult, OfferItem, PriceRow

    first = PriceRow(row_number=4, product_name="Rainwater RO-500 (1600 lt/gün) 20 INCH", prices={})
    second = PriceRow(row_number=5, product_name="Rainwater 80 LT Fiber Tank", prices={})
    result = MatchResult(
        offer_item=OfferItem(
            product_name="RAINWATER RO-500 20” + 80 LT TANK",
            quantity=1,
            unit_price=68900,
            discounted_price=68900,
            total_price=68900,
        ),
        matched_row=first,
        score=0.88,
        status="DUZELT",
        selected_column="2026 KURUMSAL NAKIT",
        reference_unit_price=81900,
        reference_total_price=81900,
        suggested_unit_price=81900,
        suggested_total_price=81900,
        difference=-13000,
        note="Satir birden fazla urunden olusuyor; bilesen fiyatlari toplami teklif satiriyla uyusmuyor.",
        bundle_components=[
            BundleComponentMatch("RAINWATER RO-500 20”", first, 0.88, 48950, "2026 KURUMSAL NAKIT", "list"),
            BundleComponentMatch("80 LT TANK", second, 0.84, 32950, "2026 KURUMSAL NAKIT", "list"),
        ],
    )

    view = offer_webapp.result_view_model(result, 0)

    assert view["is_bundle"] is True
    assert view["matched_name"] == "Rainwater RO-500 (1600 lt/gün) 20 INCH + Rainwater 80 LT Fiber Tank"
    assert [component["price"] for component in view["bundle_components"]] == ["48.950 TL", "32.950 TL"]
    assert view["action"]["title"] == "Birleşik fiyat düzeltmesi hazır"
    assert view["suggestion_count"] == 2


def test_bundle_component_override_recalculates_combined_price(tmp_path) -> None:
    from openpyxl import Workbook

    from backend.offer_module import webapp as offer_webapp
    from backend.offer_module.teklif_kontrol import BundleComponentMatch, FinancialReview, MatchResult, OfferItem, PriceRow

    price_path = tmp_path / "fiyat.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["URUN", "2026 KURUMSAL NAKIT"])
    sheet.append(["Rainwater RO-500", 48950])
    sheet.append(["Rainwater 80 LT Fiber Tank", 32950])
    sheet.append(["Rainwater 80 LT Emaye Tank", 21950])
    workbook.save(price_path)

    first = PriceRow(row_number=2, product_name="Rainwater RO-500", prices={"2026 KURUMSAL NAKIT": 48950})
    second = PriceRow(row_number=3, product_name="Rainwater 80 LT Fiber Tank", prices={"2026 KURUMSAL NAKIT": 32950})
    result = MatchResult(
        offer_item=OfferItem(
            product_name="RAINWATER RO-500 + 80 LT TANK",
            quantity=1,
            unit_price=68900,
            discounted_price=68900,
            total_price=68900,
        ),
        matched_row=first,
        score=0.88,
        status="DUZELT",
        selected_column="2026 KURUMSAL NAKIT",
        reference_unit_price=81900,
        reference_total_price=81900,
        suggested_unit_price=81900,
        suggested_total_price=81900,
        difference=-13000,
        note="Bundle.",
        bundle_components=[
            BundleComponentMatch("RAINWATER RO-500", first, 0.88, 48950, "2026 KURUMSAL NAKIT", "list"),
            BundleComponentMatch("80 LT TANK", second, 0.84, 32950, "2026 KURUMSAL NAKIT", "list"),
        ],
    )
    session = offer_webapp.ComparisonSession(
        token="bundle-override",
        price_list_path=price_path,
        offer_path=tmp_path / "teklif.pdf",
        output_path=tmp_path / "rapor.xlsx",
        selected_column="2026 KURUMSAL NAKIT",
        price_mode="kurumsal_nakit",
        results=[result],
        financial_review=FinancialReview(
            vat_rate=20,
            vat_rate_source="default",
            vat_included=True,
            item_gross_total=68900,
            expected_net_total=57416.67,
            expected_vat_total=11483.33,
            expected_gross_total=68900,
            expected_summary_total=68900,
            checks=[],
        ),
    )

    changed_count = offer_webapp.apply_bundle_match_overrides(session, ["0:1:4"])
    updated = session.results[0]

    assert changed_count == 1
    assert updated.bundle_components[1].matched_row.product_name == "Rainwater 80 LT Emaye Tank"
    assert updated.reference_unit_price == 70900
    assert updated.suggested_total_price == 70900
    assert updated.status == "DUZELT"


def test_generated_offer_pdf_keeps_turkish_text_and_template_layout(tmp_path) -> None:
    from datetime import date

    import fitz
    from openpyxl import Workbook

    from backend.offer_module.teklif_kontrol import OfferSelection, create_offer_from_catalog

    template_path = tmp_path / "sablon.pdf"
    template_doc = fitz.open()
    template_page = template_doc.new_page(width=540, height=780)
    template_page.draw_rect(
        fitz.Rect(0, 0, 540, 84),
        color=(0.2, 0.6, 0.9),
        fill=(0.2, 0.6, 0.9),
    )
    template_page.insert_textbox(
        fitz.Rect(340, 732, 520, 748),
        "444 0 420 | www.rainwater.com.tr",
        fontsize=8,
        color=(0, 0, 0),
    )
    template_doc.save(template_path)
    template_doc.close()

    price_path = tmp_path / "fiyat.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["URUN", "2026 KURUMSAL NAKIT"])
    sheet.append(["Rainwater \u00c7ift Filtre \u00d6zel \u0130\u00e7me Suyu Ar\u0131tma Sistemi", 12500])
    workbook.save(price_path)

    output_path = tmp_path / "teklif.pdf"
    create_offer_from_catalog(
        template_path=template_path,
        price_list_path=price_path,
        selected_column="2026 KURUMSAL NAKIT",
        selected_entries=[OfferSelection(2, 2, discount_type="amount", discount_value=500)],
        vat_included=False,
        offer_number="RW-TEST-001",
        offer_date=date(2026, 4, 24),
        valid_until=date(2026, 5, 1),
        company_name="\u00c7a\u011fr\u0131 \u0130leti\u015fim \u015eirketi",
        contact_name="\u0130lknur Han\u0131m",
        email="test@example.com",
        gsm="0555 111 22 33",
        note_text="T\u00fcrk\u00e7e karakter kontrol\u00fc: \u015f, \u011f, \u00fc, \u00f6, \u00e7, \u0130",
        output_path=output_path,
    )

    generated = fitz.open(output_path)
    page = generated[0]
    text = page.get_text().replace("\xa0", " ")
    assert "\u00c7a\u011fr\u0131 \u0130leti\u015fim \u015eirketi" in text
    assert "RAINWATER \u00c7\u0130FT F\u0130LTRE \u00d6ZEL \u0130\u00c7ME" in text
    assert "SUYU ARITMA S\u0130STEM\u0130" in text
    assert "T\u00fcrk\u00e7e karakter kontrol\u00fc: \u015f, \u011f, \u00fc, \u00f6, \u00e7, \u0130" in text
    assert "\u0130SKONTO TUTARI" in text
    assert "YATIRIM MAL\u0130YET\u0130" in text
    assert "KDV (%20)" in text
    assert "19.833 TL" in text
    assert "23.800 TL" in text

    bottom_signature_lines = []
    for block in page.get_text("dict").get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            line_text = "".join(span.get("text", "") for span in line.get("spans", []))
            if "YETK\u0130L\u0130" in line_text and fitz.Rect(line["bbox"]).y0 > 630:
                bottom_signature_lines.append(fitz.Rect(line["bbox"]))
    assert bottom_signature_lines
    assert bottom_signature_lines[0].x0 > 330


def test_offer_parser_uses_layout_rows_for_interleaved_price_table(tmp_path) -> None:
    import fitz

    from backend.offer_module.teklif_kontrol import (
        build_financial_review,
        parse_offer_financial_summary,
        parse_offer_items,
    )

    offer_path = tmp_path / "layout_offer.pdf"
    doc = fitz.open()
    page = doc.new_page(width=540, height=760)
    page.insert_text((110, 345), "MALZEME MIKTAR BIRIM FIYAT", fontsize=10)
    page.insert_text((473, 345), "TOPLAM TUTAR", fontsize=10)

    rows = [
        (
            380,
            406,
            ["RAINWATER RNW-2200", "ARITMALI SICAK & SOGUK SU SEBIL", "TOPLAM SU KAPASITESI 25.8 L"],
            "3 ADET 65.791 TL 54.125 TL 162.375 TL",
        ),
        (
            450,
            476,
            ["RAINWATER RNW-1600", "ARITMALI SICAK & SOGUK SU SEBIL", "TOPLAM SU KAPASITESI 9.6 L"],
            "1 ADET 62.458 TL 49.959 TL 49.959 TL",
        ),
        (
            520,
            546,
            ["RAINWATER RNW-1600'S TEZGAH USTU", "ARITMALI SICAK & SOGUK SU SEBIL", "TOPLAM SU KAPASITESI 8.8 L"],
            "1 ADET 62.458 TL 49.959 TL 49.959 TL",
        ),
        (
            590,
            606,
            ["RAINWATER RAINBOW", "ARITMA SISTEMI"],
            "1 ADET 29.125 TL 19.959 TL 19.959 TL",
        ),
    ]
    for product_y, price_y, product_lines, price_line in rows:
        for offset, line in enumerate(product_lines):
            page.insert_text((25, product_y + (offset * 13)), line, fontsize=9)
        page.insert_text((268, price_y), price_line, fontsize=9)

    page.insert_text((23, 633), "RAINWATER RO-300 20'' + 40 LT TANK", fontsize=9)
    page.insert_text((23, 646), "YUKSEK KAPASITELI ICME SUYU ARITIM SISTEMI", fontsize=9)
    page.insert_text((23, 659), "900LT/GUN-37.5LT/SAAT", fontsize=9)
    page.insert_text((268, 646), "1 ADET 65.750 TL", fontsize=9)
    page.insert_text((398, 646), "49.917 TL", fontsize=9)
    page.insert_text((474, 646), "49.917 TL", fontsize=9)
    page.insert_text((45, 690), "Fiyatlarimiza KDV (%20) dahil degildir.", fontsize=9)
    page.insert_text((351, 680), "YATIRIM MALIYETI :", fontsize=10)
    page.insert_text((478, 680), "282.210 TL", fontsize=10)
    doc.save(offer_path)
    doc.close()

    items, offer_text = parse_offer_items(offer_path)
    summary = parse_offer_financial_summary(offer_text)
    review = build_financial_review(items, offer_text)
    net_check = next(check for check in review.checks if check.label.startswith("Yat"))

    assert len(items) == 5
    assert any("RAINBOW" in item.product_name for item in items)
    assert any("RO-300" in item.product_name for item in items)
    assert round(sum(item.total_price for item in items), 2) == 332169.0
    assert summary.net_total == 282210.0
    assert net_check.status == "DUZELT"
    assert net_check.offer_value == 282210.0
    assert net_check.calculated_value == 332169.0


def test_offer_parser_does_not_limit_layout_row_count(tmp_path) -> None:
    import fitz

    from backend.offer_module.teklif_kontrol import parse_offer_items

    offer_path = tmp_path / "many_rows_offer.pdf"
    doc = fitz.open()
    page = doc.new_page(width=540, height=760)
    page.insert_text((35, 78), "MALZEME", fontsize=10)
    page.insert_text((270, 78), "MIKTAR", fontsize=10)
    page.insert_text((330, 78), "BIRIM FIYAT", fontsize=10)
    page.insert_text((402, 78), "KURUMSAL INDIRIMLI FIYAT", fontsize=10)
    page.insert_text((475, 78), "TOPLAM TUTAR", fontsize=10)

    expected_total = 0
    for index in range(8):
        y = 112 + (index * 62)
        unit_price = 10000 + (index * 1000)
        discounted_price = 9000 + (index * 1000)
        expected_total += discounted_price
        page.insert_text((25, y), f"RAINWATER TEST URUN {index + 1}", fontsize=9)
        page.insert_text((25, y + 12), f"MODEL RNW-{2200 + index}", fontsize=9)
        if index == 6:
            page.insert_text((268, y + 12), f"1 ADET {unit_price:,.0f}".replace(",", ".") + " TL", fontsize=9)
            page.insert_text(
                (395, y + 24),
                f"{discounted_price:,.0f} TL {discounted_price:,.0f} TL".replace(",", "."),
                fontsize=9,
            )
        else:
            page.insert_text(
                (268, y + 12),
                f"1 ADET {unit_price:,.0f} TL {discounted_price:,.0f} TL {discounted_price:,.0f} TL".replace(",", "."),
                fontsize=9,
            )

    page.insert_text((351, 635), "YATIRIM MALIYETI :", fontsize=10)
    page.insert_text((478, 635), f"{expected_total:,.0f} TL".replace(",", "."), fontsize=10)
    doc.save(offer_path)
    doc.close()

    items, _offer_text = parse_offer_items(offer_path)

    assert len(items) == 8
    assert round(sum(item.total_price for item in items), 2) == float(expected_total)
    assert items[6].discounted_price == 15000.0


def test_offer_parser_keeps_last_row_when_summary_label_is_joined(tmp_path) -> None:
    import fitz

    from backend.offer_module.teklif_kontrol import build_financial_review, parse_offer_items

    offer_path = tmp_path / "joined_summary_label_offer.pdf"
    doc = fitz.open()
    page = doc.new_page(width=540, height=720)
    page.insert_text((30, 330), "MALZEME", fontsize=10)
    page.insert_text((268, 330), "MIKTAR", fontsize=10)
    page.insert_text((330, 330), "BIRIM FIYAT", fontsize=10)
    page.insert_text((472, 330), "TOPLAM", fontsize=10)

    rows = [
        (370, "RAINWATER RO-300 10'' + 20 LT TANK", "1 ADET 61.583 TL 46.583 TL 46.583 TL"),
        (438, "RAINWATER RO-500 10'' + 80 LT TANK", "1 ADET 79.916 TL 59.916 TL 59.916 TL"),
        (506, "RAINWATER RO-300 20'' + 80 LT TANK", "1 ADET 70.750 TL 54.916 TL 54.916 TL"),
    ]
    for y, product_name, price_line in rows:
        page.insert_text((24, y), product_name, fontsize=9)
        page.insert_text((24, y + 13), "YUKSEK KAPASITELI ICME SUYU ARITIM SISTEMI", fontsize=9)
        page.insert_text((268, y + 13), price_line, fontsize=9)

    page.insert_text((45, 570), "Fiyatlarimiza KDV (%20) dahil degildir.", fontsize=9)
    page.insert_text((315, 586), "YATIRIMMALIYETI :", fontsize=10)
    page.insert_text((458, 586), "161.415 TL", fontsize=10)
    doc.save(offer_path)
    doc.close()

    items, offer_text = parse_offer_items(offer_path)
    review = build_financial_review(items, offer_text)

    assert len(items) == 3
    assert round(sum(item.total_price for item in items), 2) == 161415.0
    assert review.overall_status == "ONAY"
    assert [check.label for check in review.checks] == [
        "\u00dcr\u00fcnlerin Toplam Tutar\u0131",
        "KDV Modu",
        "Yat\u0131r\u0131m Maliyeti (KDV Hari\u00e7)",
    ]


def test_offer_pdf_correction_handles_bundle_row_index_mismatch(tmp_path) -> None:
    import fitz

    from backend.offer_module.teklif_kontrol import (
        MatchResult,
        OfferItem,
        PriceRow,
        apply_approved_corrections_to_pdf,
    )

    offer_path = tmp_path / "bundle_offer.pdf"
    doc = fitz.open()
    page = doc.new_page(width=540, height=760)
    row_fill = (219 / 255, 242 / 255, 247 / 255)
    page.draw_rect(fitz.Rect(10, 200, 530, 260), color=row_fill, fill=row_fill)
    page.insert_text((20, 180), "MALZEME        MİKTAR        BİRİM FİYAT        TOPLAM TUTAR", fontsize=10)
    page.insert_textbox(
        fitz.Rect(25, 205, 225, 250),
        "RAINWATER RO-300 20” + 40 LT TANK\nYÜKSEK KAPASİTELİ İÇME SUYU ARITIM SİSTEMİ",
        fontsize=9,
    )
    page.insert_text(
        fitz.Point(250, 230),
        "1 ADET      44.950 TL      44.950 TL      44.950 TL",
        fontsize=9,
    )
    doc.save(offer_path)
    doc.close()

    results = [
        MatchResult(
            offer_item=OfferItem(
                product_name="Başka satır",
                quantity=1,
                unit_price=100,
                discounted_price=100,
                total_price=100,
            ),
            matched_row=None,
            score=0,
            status="ONAY",
            selected_column="2026 KURUMSAL NAKIT",
            reference_unit_price=100,
            reference_total_price=100,
            suggested_unit_price=None,
            suggested_total_price=None,
            difference=0,
            note="Test satırı.",
        ),
        MatchResult(
            offer_item=OfferItem(
                product_name="RAINWATER RO-300 20” + 40 LT TANK YÜKSEK KAPASİTELİ İÇME SUYU ARITIM SİSTEMİ",
                quantity=1,
                unit_price=44950,
                discounted_price=44950,
                total_price=44950,
            ),
            matched_row=PriceRow(
                row_number=2,
                product_name="RAINWATER RO-300 20” + 40 LT TANK YÜKSEK KAPASİTELİ İÇME SUYU ARITIM SİSTEMİ",
                prices={"2026 KURUMSAL NAKIT": 64900},
            ),
            score=0.99,
            status="DUZELT",
            selected_column="2026 KURUMSAL NAKIT",
            reference_unit_price=64900,
            reference_total_price=64900,
            suggested_unit_price=64900,
            suggested_total_price=64900,
            difference=-19950,
            note="Satırda birden fazla ürün olabilir; fiyat farkı bulundu.",
        ),
    ]

    output_path = tmp_path / "bundle_offer_corrected.pdf"
    corrected_path = apply_approved_corrections_to_pdf(
        offer_path=offer_path,
        results=results,
        approved_indexes=[1],
        output_path=output_path,
    )

    corrected = fitz.open(corrected_path)
    corrected_text = corrected[0].get_text()
    replacement_rects = corrected[0].search_for("64.900 TL")
    assert replacement_rects
    replacement_rect = max(replacement_rects, key=lambda rect: rect.x0)
    pix = corrected[0].get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
    sample_x = min(pix.width - 1, max(0, int((replacement_rect.x1 + 3) * 2)))
    sample_y = min(pix.height - 1, max(0, int((replacement_rect.y0 + (replacement_rect.height / 2)) * 2)))
    sample_index = sample_y * getattr(pix, "stride", pix.width * pix.n) + sample_x * pix.n
    red, green, blue = pix.samples[sample_index : sample_index + 3]
    corrected.close()

    assert "64.900 TL" in corrected_text
    assert red < 245
    assert green > 220
    assert blue > 225
